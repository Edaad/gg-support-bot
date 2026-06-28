"""Build minimal Trade Record XLSX workbooks for tests."""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook

from api.trade_record_parser import (
    COL_AGENT_ID,
    COL_AGENT_NICK,
    COL_AMOUNT,
    COL_MEMBER_ID,
    COL_MEMBER_NICK,
    COL_SA_ID,
    COL_SA_NICK,
    COL_TIME,
    DATA_START_ROW,
    SHEET_NAME,
)


def build_sample_trade_record_xlsx(
    *,
    club_label: str = "Aces Table",
    audit_date: date | None = None,
    include_second_row: bool = True,
) -> bytes:
    d = audit_date or date(2026, 6, 21)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.cell(row=1, column=1, value="Club")
    ws.cell(row=1, column=2, value=club_label)
    ws.cell(row=2, column=1, value="Date")
    ws.cell(row=2, column=2, value=d.isoformat())

    row = DATA_START_ROW
    ws.cell(row=row, column=COL_TIME, value="14:30:00")
    ws.cell(row=row, column=COL_AMOUNT, value=100)
    ws.cell(row=row, column=COL_SA_ID, value="1000-1001")
    ws.cell(row=row, column=COL_SA_NICK, value="SuperOne")
    ws.cell(row=row, column=COL_AGENT_ID, value="2000-2001")
    ws.cell(row=row, column=COL_AGENT_NICK, value="AgentOne")
    ws.cell(row=row, column=COL_MEMBER_ID, value="3011-9668")
    ws.cell(row=row, column=COL_MEMBER_NICK, value="MemberOne")

    if include_second_row:
        row2 = DATA_START_ROW + 1
        ws.cell(row=row2, column=COL_TIME, value="15:00:00")
        ws.cell(row=row2, column=COL_AMOUNT, value=-50)
        ws.cell(row=row2, column=COL_MEMBER_ID, value="3011-9668")
        ws.cell(row=row2, column=COL_MEMBER_NICK, value="MemberOne")
        ws.cell(row=row2, column=COL_AGENT_ID, value="-")
        ws.cell(row=row2, column=COL_SA_ID, value="-")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
