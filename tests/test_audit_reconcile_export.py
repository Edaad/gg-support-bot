"""Tests for reconcile XLSX export."""

from __future__ import annotations

import io
import unittest
from dataclasses import replace
from datetime import date, datetime, timezone
from decimal import Decimal
from unittest.mock import patch

from openpyxl import load_workbook

from api.audit_ledger import LedgerBreakdown, LedgerLine
from api.audit_reconcile import AuditReconcilePlayerResult, AuditReconcileReport, TradeLineForMatch
from api.audit_reconcile_export import (
    CLUBGTO_CASHOUT_LABELING_ENABLED,
    MATCHING_HEADERS,
    MATCHING_WIDTHS,
    UNRESOLVED_HEADERS,
    _CURRENCY_FORMAT,
    _EXCEL_TIME_FORMAT,
    _MATCHING_BAND_FILL,
    _MATCHING_BODY_FONT,
    _MATCHING_HEADER_FONT,
    _MATCHING_ROW_HEIGHT,
    _MATCHING_SOURCE_FILL_HEX,
    build_all_clubs_matching_workbook,
)


def _player(
    *,
    gg_player_id: str,
    nickname: str,
    net_trade: str,
    net_ledger: str,
    delta: str,
    status: str,
) -> AuditReconcilePlayerResult:
    return AuditReconcilePlayerResult(
        gg_player_id=gg_player_id,
        member_nickname=nickname,
        net_trade_record=Decimal(net_trade),
        net_ledger=Decimal(net_ledger),
        delta=Decimal(delta),
        ledger_breakdown=LedgerBreakdown(
            deposits=Decimal(net_ledger),
            early_rb=Decimal("0"),
            bonuses=Decimal("0"),
            monday=Decimal("0"),
            cashouts=Decimal("0"),
        ),
        status=status,
    )


def _empty_report(*, club_slug: str, club_name: str) -> AuditReconcileReport:
    return AuditReconcileReport(
        audit_date=date(2026, 7, 3),
        club_slug=club_slug,
        club_name=club_name,
        status="pass",
        players=[],
    )


def _all_clubs_reports(
    *,
    round_table: AuditReconcileReport | None = None,
    clubgto: AuditReconcileReport | None = None,
    creator_club: AuditReconcileReport | None = None,
) -> dict[str, AuditReconcileReport]:
    return {
        "round-table": round_table
        or _empty_report(club_slug="round-table", club_name="Round Table"),
        "clubgto": clubgto or _empty_report(club_slug="clubgto", club_name="ClubGTO"),
        "creator-club": creator_club
        or _empty_report(club_slug="creator-club", club_name="Creator Club"),
    }


def _clubgto_cashout_label_report() -> AuditReconcileReport:
    occurred = datetime(2026, 7, 3, 15, 30, tzinfo=timezone.utc)
    return AuditReconcileReport(
        audit_date=date(2026, 7, 3),
        club_slug="clubgto",
        club_name="ClubGTO",
        status="pass",
        players=[],
        trade_lines=[
            TradeLineForMatch(
                line_id=1,
                occurred_at=occurred,
                amount=Decimal("50"),
                member_gg_player_id="1111-2222",
                member_nickname="P1",
                sheet_row=1,
                manager_nickname="Mgr",
                trade_club_slug="clubgto",
            ),
        ],
        ledger_lines=[
            LedgerLine(
                gg_player_id="1111-2222",
                member_nickname="P1",
                source="cashout",
                source_label="Cashout Venmo",
                amount_signed=Decimal("50"),
                occurred_at_utc=occurred,
                external_id="cashout:1",
            ),
            LedgerLine(
                gg_player_id="3333-4444",
                member_nickname="LeftOver",
                source="cashout",
                source_label="Cashout Zelle",
                amount_signed=Decimal("22"),
                occurred_at_utc=occurred,
                external_id="cashout:2",
                display_name="Charlie Kim",
            ),
        ],
    )


class ReconcileExportTestCase(unittest.TestCase):
    def test_matching_flat_headers_manager_and_time_format(self):
        occurred = datetime(2026, 7, 3, 15, 30, tzinfo=timezone.utc)
        report = AuditReconcileReport(
            audit_date=date(2026, 7, 3),
            club_slug="clubgto",
            club_name="ClubGTO",
            status="pass",
            players=[],
            trade_lines=[
                TradeLineForMatch(
                    line_id=1,
                    occurred_at=occurred,
                    amount=Decimal("-50"),
                    member_gg_player_id="1111-2222",
                    member_nickname="P1",
                    sheet_row=1,
                    manager_nickname="TrafficLight7",
                    trade_club_slug="clubgto",
                ),
            ],
            ledger_lines=[
                LedgerLine(
                    gg_player_id="1111-2222",
                    member_nickname="P1",
                    source="deposit_zelle",
                    source_label="Zelle",
                    amount_signed=Decimal("-50"),
                    occurred_at_utc=occurred,
                    external_id="deposit_zelle:1",
                    display_name="Payer",
                    variant="2133729202",
                    method_owner="vaughn",
                ),
                LedgerLine(
                    gg_player_id="1111-2222",
                    member_nickname="P1",
                    source="deposit_venmo",
                    source_label="Venmo",
                    amount_signed=Decimal("-20"),
                    occurred_at_utc=occurred,
                    external_id="deposit_venmo:1",
                    variant="@janseashells",
                    method_owner="vaughn",
                ),
            ],
        )
        wb = load_workbook(
            io.BytesIO(
                build_all_clubs_matching_workbook(_all_clubs_reports(clubgto=report))
            )
        )
        matching = wb["ClubGTO"]
        self.assertEqual(matching.cell(row=1, column=1).value, "Trade Time")
        self.assertEqual(matching.cell(row=1, column=2).value, "Manager")
        self.assertEqual(matching.cell(row=2, column=2).value, "TrafficLight7")
        self.assertEqual(matching.cell(row=2, column=10).value, "2133729202")
        self.assertEqual(
            matching.cell(row=2, column=3).number_format,
            _CURRENCY_FORMAT,
        )
        self.assertIn("[Red]", _CURRENCY_FORMAT)
        self.assertIsInstance(matching.cell(row=2, column=1).value, datetime)
        self.assertEqual(matching.cell(row=2, column=1).number_format, _EXCEL_TIME_FORMAT)
        # 15:30 UTC in July → 11:30 America/New_York (EDT), not club UTC-5 10:30.
        self.assertEqual(
            matching.cell(row=2, column=1).value,
            datetime(2026, 7, 3, 11, 30),
        )
        self.assertIsInstance(matching.cell(row=2, column=8).value, datetime)
        self.assertEqual(matching.cell(row=2, column=8).number_format, _EXCEL_TIME_FORMAT)
        self.assertEqual(
            matching.cell(row=2, column=8).value,
            datetime(2026, 7, 3, 11, 30),
        )
        self.assertEqual(matching.cell(row=2, column=6).value, "GTO Zelle")
        # Vaughn tally keys off Source = GTO …
        self.assertEqual(matching.cell(row=1, column=12).value, "Vaughn methods")
        self.assertEqual(matching.cell(row=2, column=12).value, "Method")
        self.assertEqual(matching.cell(row=3, column=12).value, "Zelle")
        self.assertEqual(matching.cell(row=3, column=13).value, "2133729202")
        self.assertEqual(matching.cell(row=2, column=9).value, 50.0)
        self.assertEqual(matching.cell(row=2, column=9).number_format, _CURRENCY_FORMAT)
        zelle_count = matching.cell(row=3, column=14).value
        zelle_total = matching.cell(row=3, column=15).value
        self.assertIsInstance(zelle_count, str)
        self.assertTrue(zelle_count.startswith("="))
        self.assertIn('COUNTIFS($F:$F,"GTO Zelle",$J:$J,"2133729202")', zelle_count)
        self.assertIsInstance(zelle_total, str)
        self.assertIn('SUMIFS($I:$I,$F:$F,"GTO Zelle",$J:$J,"2133729202")', zelle_total)
        chips_title_row = next(
            r
            for r in range(1, 30)
            if matching.cell(row=r, column=12).value == "Vaughn methods (chips)"
        )
        chips_zelle_row = chips_title_row + 2
        self.assertEqual(matching.cell(row=chips_zelle_row, column=12).value, "Zelle")
        chips_total = matching.cell(row=chips_zelle_row, column=15).value
        self.assertIsInstance(chips_total, str)
        self.assertIn("$C$2:$C$2", chips_total)
        self.assertIn("ABS($C$2:$C$2)", chips_total)
        self.assertIn('($F$2:$F$2="GTO Zelle")', chips_total)
        self.assertEqual(matching.auto_filter.ref, "A1:J2")
        self.assertEqual(list(matching.tables), [])
        header_fill = matching.cell(row=1, column=1).fill.fgColor.rgb
        self.assertTrue(str(header_fill).endswith("306A54"))
        self.assertEqual(matching.cell(row=1, column=1).font.name, _MATCHING_HEADER_FONT.name)
        self.assertEqual(matching.cell(row=1, column=1).font.size, _MATCHING_HEADER_FONT.size)
        self.assertEqual(matching.cell(row=2, column=5).font.name, _MATCHING_BODY_FONT.name)
        self.assertEqual(matching.cell(row=2, column=5).font.size, _MATCHING_BODY_FONT.size)
        # First data row is not banded.
        self.assertNotEqual(
            (matching.cell(row=2, column=1).fill.fgColor or None)
            and matching.cell(row=2, column=1).fill.fgColor.rgb,
            _MATCHING_BAND_FILL.fgColor.rgb,
        )
        self.assertEqual(matching.row_dimensions[1].height, _MATCHING_ROW_HEIGHT)
        self.assertEqual(matching.row_dimensions[2].height, _MATCHING_ROW_HEIGHT)
        self.assertEqual(matching.column_dimensions["A"].width, MATCHING_WIDTHS[0])
        self.assertEqual(matching.column_dimensions["H"].width, MATCHING_WIDTHS[7])
        self.assertGreaterEqual(MATCHING_WIDTHS[0], 18)
        self.assertGreaterEqual(MATCHING_WIDTHS[7], 18)
        validations = list(matching.data_validations.dataValidation)
        self.assertEqual(len(validations), 2)
        source_dv = next(
            dv
            for dv in validations
            if "INDIRECT" not in (dv.formula1 or "")
        )
        self.assertIn("F2:F2", source_dv.sqref)
        self.assertFalse((source_dv.formula1 or "").startswith("="))
        self.assertIn("ClubGTO", source_dv.formula1)
        self.assertIn("$AD$1:", source_dv.formula1)
        self.assertNotIn("Source lists", wb.sheetnames)
        variant_dv = next(dv for dv in validations if "INDIRECT" in (dv.formula1 or ""))
        self.assertIn("J2:J2", variant_dv.sqref)
        self.assertIn("MATCH(", variant_dv.formula1)
        self.assertIn("ADDRESS(", variant_dv.formula1)
        source_list = [
            matching.cell(row=r, column=30).value for r in range(1, 50)
        ]
        self.assertEqual(source_list[0], "Stripe")
        self.assertIn("Union Zelle", source_list)
        self.assertIn("Large cashout Zelle", source_list)
        self.assertIn("Union Venmo", source_list)
        self.assertIn("Chip Transfer (Player)", source_list)
        self.assertIn("Free Play", source_list)
        self.assertIn("Back to Club", source_list)
        self.assertIn("GTO INC", source_list)
        self.assertNotIn("Chip Transfer (RT↔AT)", source_list)
        self.assertNotIn("Chip Transfer (AT↔CC)", source_list)
        self.assertTrue(matching.column_dimensions["AD"].hidden)
        self.assertTrue(matching.column_dimensions["AE"].hidden)
        hidden_headers = [
            matching.cell(row=1, column=col).value for col in range(30, 120)
        ]
        self.assertIn("Cashout Venmo", hidden_headers)
        self.assertIn("Vaughn Cashout Venmo", hidden_headers)
        self.assertIn("Vaughn Cashout Cash App", hidden_headers)
        self.assertIn("Chip Transfer (Player)", hidden_headers)
        self.assertNotIn("Chip Transfer (RT↔AT)", hidden_headers)
        self.assertNotIn("Chip Transfer (AT↔CC)", hidden_headers)
        cf_formulas: list[str] = []
        for cf_range in matching.conditional_formatting._cf_rules:
            self.assertIn("F2", str(cf_range))
            for rule in matching.conditional_formatting._cf_rules[cf_range]:
                cf_formulas.extend(rule.formula or [])
        self.assertTrue(any("Cashout Venmo" in f for f in cf_formulas))
        self.assertTrue(any("Vaughn Cashout Venmo" in f for f in cf_formulas))
        self.assertTrue(any("Stripe" in f for f in cf_formulas))
        self.assertTrue(any("Mateos Venmo" in f for f in cf_formulas))
        self.assertTrue(any("Chip Transfer (Player)" in f for f in cf_formulas))
        self.assertTrue(any("Chip Transfer (RT↔AT)" in f for f in cf_formulas))
        self.assertTrue(any("Chip Transfer (AT↔CC)" in f for f in cf_formulas))
        self.assertTrue(any("Free Play" in f for f in cf_formulas))
        self.assertTrue(any("Back to Club" in f for f in cf_formulas))
        self.assertTrue(any("GTO INC" in f for f in cf_formulas))
        self.assertEqual(
            len(cf_formulas),
            len(_MATCHING_SOURCE_FILL_HEX),
        )

    def test_unresolved_sheet_lists_unmatched_ledger(self):
        occurred = datetime(2026, 7, 3, 15, 30, tzinfo=timezone.utc)
        report = AuditReconcileReport(
            audit_date=date(2026, 7, 3),
            club_slug="clubgto",
            club_name="ClubGTO",
            status="pass",
            players=[],
            trade_lines=[
                TradeLineForMatch(
                    line_id=1,
                    occurred_at=occurred,
                    amount=Decimal("-50"),
                    member_gg_player_id="1111-2222",
                    member_nickname="P1",
                    sheet_row=1,
                    manager_nickname="Mgr",
                    trade_club_slug="clubgto",
                ),
            ],
            ledger_lines=[
                LedgerLine(
                    gg_player_id="1111-2222",
                    member_nickname="P1",
                    source="deposit_zelle",
                    source_label="Zelle",
                    amount_signed=Decimal("-50"),
                    occurred_at_utc=occurred,
                    external_id="deposit_zelle:1",
                    display_name="Matched Payer",
                    variant="2133729202",
                    detail="GTO / 1111-2222 / P1",
                ),
                LedgerLine(
                    gg_player_id="3333-4444",
                    member_nickname="LeftOver",
                    source="deposit_zelle",
                    source_label="Zelle",
                    amount_signed=Decimal("-22"),
                    occurred_at_utc=occurred,
                    external_id="deposit_zelle:9",
                    display_name="Charlie Kim",
                    variant="rt-zelle-inbox@example.com",
                    detail="GTO / 3333-4444 / Charlie Kim",
                    method_owner="round-table",
                ),
                LedgerLine(
                    gg_player_id="5555-6666",
                    member_nickname="GtoOnly",
                    source="deposit_stripe",
                    source_label="Stripe",
                    amount_signed=Decimal("-15"),
                    occurred_at_utc=occurred,
                    external_id="deposit_stripe:2",
                    display_name="Gto Only",
                    detail="GTO / 5555-6666 / Gto Only",
                ),
            ],
        )
        wb = load_workbook(
            io.BytesIO(
                build_all_clubs_matching_workbook(_all_clubs_reports(clubgto=report))
            )
        )
        unresolved = wb["Unresolved"]
        self.assertEqual(
            [unresolved.cell(row=1, column=c).value for c in range(1, 8)],
            UNRESOLVED_HEADERS,
        )
        # Unmatched ledger order follows match leftover order (stripe after zelle leftover
        # depends on match algorithm — assert by Source+Variant content).
        rows = [
            (
                unresolved.cell(row=r, column=1).value,
                unresolved.cell(row=r, column=2).value,
                unresolved.cell(row=r, column=3).value,
                unresolved.cell(row=r, column=4).value,
                unresolved.cell(row=r, column=6).value,
            )
            for r in range(2, 4)
        ]
        self.assertEqual(
            set(rows),
            {
                ("RT Zelle", "rt-zelle-inbox@example.com", 22.0, "Charlie Kim", "ClubGTO"),
                ("Stripe", None, 15.0, "Gto Only", "ClubGTO"),
            },
        )
        self.assertTrue(
            str(unresolved.cell(row=2, column=7).value).endswith("AM")
            or str(unresolved.cell(row=2, column=7).value).endswith("PM")
        )
        self.assertIsNone(unresolved.cell(row=4, column=1).value)
        self.assertIn("Unresolved_all", unresolved.tables)

    def test_cashout_method_label_matching_and_unresolved(self):
        self.assertFalse(CLUBGTO_CASHOUT_LABELING_ENABLED)
        report = _clubgto_cashout_label_report()
        wb = load_workbook(
            io.BytesIO(
                build_all_clubs_matching_workbook(_all_clubs_reports(clubgto=report))
            )
        )
        matching = wb["ClubGTO"]
        self.assertEqual(matching.cell(row=2, column=2).value, "Mgr")
        self.assertEqual(matching.cell(row=2, column=5).value, "P1")
        self.assertIn(matching.cell(row=2, column=6).value, ("", None))
        self.assertIn(matching.cell(row=2, column=7).value, ("", None))
        self.assertIn(matching.cell(row=2, column=8).value, ("", None))
        self.assertIsNone(matching.cell(row=2, column=9).value)
        self.assertIn(matching.cell(row=2, column=10).value, ("", None))
        unresolved = wb["Unresolved"]
        self.assertEqual(unresolved.cell(row=2, column=1).value, "Cashout Venmo")
        self.assertEqual(unresolved.cell(row=3, column=1).value, "Cashout Zelle")
        self.assertEqual(unresolved.cell(row=2, column=4).value, "P1")
        self.assertEqual(unresolved.cell(row=3, column=4).value, "Charlie Kim")

    @patch("api.audit_reconcile_export.CLUBGTO_CASHOUT_LABELING_ENABLED", True)
    def test_clubgto_cashout_labeling_when_enabled(self):
        report = _clubgto_cashout_label_report()
        wb = load_workbook(
            io.BytesIO(
                build_all_clubs_matching_workbook(_all_clubs_reports(clubgto=report))
            )
        )
        matching = wb["ClubGTO"]
        self.assertEqual(matching.cell(row=2, column=6).value, "Cashout Venmo")
        self.assertIsNone(matching.cell(row=2, column=10).value)
        unresolved = wb["Unresolved"]
        self.assertEqual(unresolved.cell(row=2, column=1).value, "Cashout Zelle")

    def test_round_table_source_dropdown_includes_free_play_and_back_to_club(self):
        occurred = datetime(2026, 7, 3, 15, 30, tzinfo=timezone.utc)
        report = AuditReconcileReport(
            audit_date=date(2026, 7, 3),
            club_slug="round-table",
            club_name="Round Table",
            status="pass",
            players=[],
            trade_lines=[
                TradeLineForMatch(
                    line_id=1,
                    occurred_at=occurred,
                    amount=Decimal("-50"),
                    member_gg_player_id="1111-2222",
                    member_nickname="P1",
                    sheet_row=1,
                    trade_club_slug="round-table",
                ),
            ],
            ledger_lines=[],
        )
        wb = load_workbook(
            io.BytesIO(
                build_all_clubs_matching_workbook(_all_clubs_reports(round_table=report))
            )
        )
        matching = wb["Round Table"]
        source_list = [
            matching.cell(row=r, column=30).value for r in range(1, 80)
        ]
        source_list = [value for value in source_list if value]
        self.assertIn("Free Play", source_list)
        self.assertIn("Back to Club", source_list)

    def test_creator_club_mateos_pivot(self):
        occurred = datetime(2026, 7, 3, 15, 30, tzinfo=timezone.utc)
        report = AuditReconcileReport(
            audit_date=date(2026, 7, 3),
            club_slug="creator-club",
            club_name="Creator Club",
            status="pass",
            players=[],
            trade_lines=[
                TradeLineForMatch(
                    line_id=1,
                    occurred_at=occurred,
                    amount=Decimal("-25"),
                    member_gg_player_id="1111-2222",
                    member_nickname="P1",
                    sheet_row=1,
                    trade_club_slug="creator-club",
                ),
            ],
            ledger_lines=[
                LedgerLine(
                    gg_player_id="1111-2222",
                    member_nickname="P1",
                    source="deposit_venmo",
                    source_label="Venmo",
                    amount_signed=Decimal("-25"),
                    occurred_at_utc=occurred,
                    external_id="deposit_venmo:1",
                    display_name="Payer",
                    variant="@mateos-handle",
                    method_owner="mateos",
                ),
            ],
        )
        wb = load_workbook(
            io.BytesIO(
                build_all_clubs_matching_workbook(_all_clubs_reports(creator_club=report))
            )
        )
        matching = wb["Creator Club"]
        self.assertEqual(matching.cell(row=1, column=12).value, "Mateos methods")
        self.assertEqual(matching.cell(row=3, column=12).value, "Venmo")
        self.assertEqual(matching.cell(row=3, column=13).value, "@mateos-handle")
        self.assertEqual(matching.cell(row=2, column=6).value, "Mateos Venmo")

    def test_matching_vaughn_tally_only_clubgto(self):
        report = _empty_report(club_slug="round-table", club_name="Round Table")
        report.ledger_lines = [
            LedgerLine(
                gg_player_id="1",
                member_nickname="P",
                source="deposit_zelle",
                source_label="Zelle",
                amount_signed=Decimal("-10"),
                occurred_at_utc=None,
                external_id="deposit_zelle:1",
                variant="2133729202",
            ),
        ]
        wb = load_workbook(
            io.BytesIO(
                build_all_clubs_matching_workbook(_all_clubs_reports(round_table=report))
            )
        )
        matching = wb["Round Table"]
        self.assertIsNone(matching.cell(row=1, column=12).value)
        self.assertEqual(matching.auto_filter.ref, "A1:J1")
        self.assertEqual(list(matching.tables), [])

    def test_all_clubs_matching_workbook_sheet_order(self):
        reports = {
            "round-table": _empty_report(
                club_slug="round-table", club_name="Round Table"
            ),
            "clubgto": _empty_report(club_slug="clubgto", club_name="ClubGTO"),
            "creator-club": _empty_report(
                club_slug="creator-club", club_name="Creator Club"
            ),
        }
        wb = load_workbook(io.BytesIO(build_all_clubs_matching_workbook(reports)))
        self.assertEqual(
            wb.sheetnames,
            [
                "Round Table",
                "Aces Table",
                "ClubGTO",
                "Creator Club",
                "Unresolved",
            ],
        )
        for name in ("Round Table", "Aces Table", "ClubGTO", "Creator Club"):
            self.assertEqual(wb[name].cell(row=1, column=1).value, "Trade Time")
            self.assertEqual(wb[name].cell(row=1, column=2).value, "Manager")
        self.assertEqual(wb["ClubGTO"].cell(row=1, column=12).value, "Vaughn methods")
        self.assertIsNone(wb["Round Table"].cell(row=1, column=12).value)
        self.assertIsNone(wb["Aces Table"].cell(row=1, column=12).value)
        self.assertEqual(wb["Creator Club"].cell(row=1, column=12).value, "Mateos methods")
        self.assertEqual(wb["Round Table"].auto_filter.ref, "A1:J1")
        self.assertEqual(list(wb["Round Table"].tables), [])
        self.assertEqual(list(wb["Aces Table"].tables), [])
        self.assertEqual(list(wb["ClubGTO"].tables), [])
        self.assertEqual(list(wb["Creator Club"].tables), [])
        unresolved = wb["Unresolved"]
        self.assertEqual(
            [unresolved.cell(row=1, column=c).value for c in range(1, 8)],
            UNRESOLVED_HEADERS,
        )
        self.assertIn("Unresolved_all", unresolved.tables)

    def test_all_clubs_splits_composite_match_by_trade_club_slug(self):
        """Option C: one composite match, rows land on RT vs Aces by trade upload."""
        occurred = datetime(2026, 7, 3, 15, 0, tzinfo=timezone.utc)
        rt_report = _empty_report(club_slug="round-table", club_name="Round Table")
        rt_report.trade_lines = [
            TradeLineForMatch(
                line_id=1,
                occurred_at=occurred,
                amount=Decimal("-50"),
                member_gg_player_id="1111-1111",
                member_nickname="RtPlayer",
                sheet_row=1,
                trade_club_slug="round-table",
            ),
            TradeLineForMatch(
                line_id=2,
                occurred_at=occurred,
                amount=Decimal("-75"),
                member_gg_player_id="2222-2222",
                member_nickname="AtPlayer",
                sheet_row=2,
                trade_club_slug="aces-table",
            ),
        ]
        rt_report.ledger_lines = [
            LedgerLine(
                gg_player_id="1111-1111",
                member_nickname="RtPlayer",
                source="deposit_stripe",
                source_label="Stripe",
                amount_signed=Decimal("-50"),
                occurred_at_utc=occurred,
                external_id="deposit_stripe:1",
                display_name="RtPlayer",
                club_slug="round-table",
            ),
            LedgerLine(
                gg_player_id="2222-2222",
                member_nickname="AtPlayer",
                source="deposit_zelle",
                source_label="Zelle",
                amount_signed=Decimal("-75"),
                occurred_at_utc=occurred,
                external_id="deposit_zelle:1",
                display_name="AtPlayer",
                club_slug="aces-table",
            ),
            LedgerLine(
                gg_player_id=None,
                member_nickname=None,
                source="deposit_venmo",
                source_label="Venmo",
                amount_signed=Decimal("-10"),
                occurred_at_utc=occurred,
                external_id="deposit_venmo:orphan",
                display_name="Orphan AT",
                club_slug="aces-table",
            ),
        ]
        reports = {
            "round-table": rt_report,
            "clubgto": _empty_report(club_slug="clubgto", club_name="ClubGTO"),
            "creator-club": _empty_report(
                club_slug="creator-club", club_name="Creator Club"
            ),
        }
        wb = load_workbook(io.BytesIO(build_all_clubs_matching_workbook(reports)))
        rt_sheet = wb["Round Table"]
        at_sheet = wb["Aces Table"]
        self.assertEqual(rt_sheet.cell(row=2, column=4).value, "1111-1111")
        self.assertEqual(rt_sheet.cell(row=2, column=5).value, "RtPlayer")
        self.assertIsNone(rt_sheet.cell(row=3, column=4).value)
        self.assertEqual(at_sheet.cell(row=2, column=4).value, "2222-2222")
        self.assertEqual(at_sheet.cell(row=2, column=5).value, "AtPlayer")
        self.assertIsNone(at_sheet.cell(row=3, column=4).value)

        unresolved = wb["Unresolved"]
        # Only the unmatched AT Venmo orphan (matched deposits consumed).
        self.assertEqual(unresolved.cell(row=2, column=4).value, "Orphan AT")
        self.assertEqual(unresolved.cell(row=2, column=6).value, "Aces Table")
        self.assertIsNone(unresolved.cell(row=3, column=4).value)

    def test_all_clubs_unresolved_sorts_mixed_naive_aware_times(self):
        """Regression: export-all 500'd on naive vs aware occurred_at compare."""
        aware = datetime(2026, 7, 22, 4, 0, tzinfo=timezone.utc)
        naive = datetime(2026, 7, 22, 5, 0)  # naive
        reports = {
            "round-table": _empty_report(
                club_slug="round-table", club_name="Round Table"
            ),
            "clubgto": _empty_report(club_slug="clubgto", club_name="ClubGTO"),
            "creator-club": _empty_report(
                club_slug="creator-club", club_name="Creator Club"
            ),
        }
        reports["clubgto"].ledger_lines = [
            LedgerLine(
                gg_player_id="1",
                member_nickname="A",
                source="deposit_stripe",
                source_label="Stripe",
                amount_signed=Decimal("-10"),
                occurred_at_utc=aware,
                external_id="deposit_stripe:1",
                display_name="Aware",
            ),
        ]
        reports["round-table"].ledger_lines = [
            LedgerLine(
                gg_player_id="2",
                member_nickname="B",
                source="deposit_venmo",
                source_label="Venmo",
                amount_signed=Decimal("-20"),
                occurred_at_utc=naive,
                external_id="deposit_venmo:1",
                display_name="Naive",
                club_slug="round-table",
            ),
        ]
        # Must not raise TypeError on sort.
        wb = load_workbook(io.BytesIO(build_all_clubs_matching_workbook(reports)))
        unresolved = wb["Unresolved"]
        self.assertEqual(unresolved.cell(row=2, column=4).value, "Aware")
        self.assertEqual(unresolved.cell(row=3, column=4).value, "Naive")
        self.assertEqual(unresolved.cell(row=3, column=6).value, "Round Table")

    def test_matching_chip_transfer_player_and_unresolved_ledger_only(self):
        occurred = datetime(2026, 7, 3, 15, 30, tzinfo=timezone.utc)
        report = AuditReconcileReport(
            audit_date=date(2026, 7, 3),
            club_slug="creator-club",
            club_name="Creator Club",
            status="pass",
            players=[],
            trade_lines=[
                TradeLineForMatch(
                    line_id=1,
                    occurred_at=occurred,
                    amount=Decimal("-40"),
                    member_gg_player_id="1111-1111",
                    member_nickname="Alice",
                    sheet_row=1,
                    trade_club_slug="creator-club",
                ),
                TradeLineForMatch(
                    line_id=2,
                    occurred_at=occurred,
                    amount=Decimal("40"),
                    member_gg_player_id="2222-2222",
                    member_nickname="Bob",
                    sheet_row=2,
                    trade_club_slug="creator-club",
                ),
            ],
            ledger_lines=[
                LedgerLine(
                    gg_player_id="9999-9999",
                    member_nickname="Orphan",
                    source="deposit_stripe",
                    source_label="Stripe",
                    amount_signed=Decimal("-12"),
                    occurred_at_utc=occurred,
                    external_id="deposit_stripe:orphan",
                    display_name="Orphan",
                ),
            ],
        )
        wb = load_workbook(
            io.BytesIO(
                build_all_clubs_matching_workbook(_all_clubs_reports(creator_club=report))
            )
        )
        matching = wb["Creator Club"]
        self.assertEqual(matching.cell(row=2, column=6).value, "Chip Transfer (Player)")
        self.assertEqual(matching.cell(row=2, column=7).value, "Bob")
        self.assertEqual(matching.cell(row=2, column=9).value, 40.0)
        self.assertEqual(matching.cell(row=3, column=6).value, "Chip Transfer (Player)")
        self.assertEqual(matching.cell(row=3, column=7).value, "Alice")
        self.assertFalse(matching.cell(row=3, column=10).value)
        hidden_headers = [
            matching.cell(row=1, column=col).value for col in range(30, 120)
        ]
        self.assertIn("Chip Transfer (Player)", hidden_headers)
        self.assertNotIn("Chip Transfer (RT↔AT)", hidden_headers)
        self.assertIn("Chip Transfer (AT↔CC)", hidden_headers)
        unresolved = wb["Unresolved"]
        self.assertEqual(unresolved.cell(row=2, column=4).value, "Orphan")
        self.assertIsNone(unresolved.cell(row=3, column=1).value)

    def test_all_clubs_chip_transfer_rt_at_split_sheets(self):
        occurred = datetime(2026, 7, 3, 15, 0, tzinfo=timezone.utc)
        rt_report = _empty_report(club_slug="round-table", club_name="Round Table")
        rt_report.trade_lines = [
            TradeLineForMatch(
                line_id=1,
                occurred_at=occurred,
                amount=Decimal("-90"),
                member_gg_player_id="1111-1111",
                member_nickname="Alice",
                sheet_row=1,
                trade_club_slug="round-table",
            ),
            TradeLineForMatch(
                line_id=2,
                occurred_at=occurred,
                amount=Decimal("90"),
                member_gg_player_id="1111-1111",
                member_nickname="Alice",
                sheet_row=2,
                trade_club_slug="aces-table",
            ),
        ]
        reports = {
            "round-table": rt_report,
            "clubgto": _empty_report(club_slug="clubgto", club_name="ClubGTO"),
            "creator-club": _empty_report(
                club_slug="creator-club", club_name="Creator Club"
            ),
        }
        wb = load_workbook(io.BytesIO(build_all_clubs_matching_workbook(reports)))
        rt_sheet = wb["Round Table"]
        at_sheet = wb["Aces Table"]
        self.assertEqual(rt_sheet.cell(row=2, column=6).value, "Chip Transfer (RT↔AT)")
        self.assertEqual(rt_sheet.cell(row=2, column=7).value, "Aces Table")
        self.assertEqual(at_sheet.cell(row=2, column=6).value, "Chip Transfer (RT↔AT)")
        self.assertEqual(at_sheet.cell(row=2, column=7).value, "Round Table")
        rt_hidden = [rt_sheet.cell(row=1, column=col).value for col in range(30, 120)]
        at_hidden = [at_sheet.cell(row=1, column=col).value for col in range(30, 120)]
        self.assertIn("Chip Transfer (RT↔AT)", rt_hidden)
        self.assertNotIn("Chip Transfer (AT↔CC)", rt_hidden)
        self.assertIn("Chip Transfer (RT↔AT)", at_hidden)
        self.assertIn("Chip Transfer (AT↔CC)", at_hidden)
        unresolved = wb["Unresolved"]
        self.assertIsNone(unresolved.cell(row=2, column=1).value)

    def test_all_clubs_chip_transfer_at_cc_split_sheets(self):
        occurred = datetime(2026, 7, 3, 15, 0, tzinfo=timezone.utc)
        rt_report = _empty_report(club_slug="round-table", club_name="Round Table")
        rt_report.trade_lines = [
            TradeLineForMatch(
                line_id=1,
                occurred_at=occurred,
                amount=Decimal("-75"),
                member_gg_player_id="1111-1111",
                member_nickname="Alice",
                sheet_row=1,
                trade_club_slug="aces-table",
            ),
        ]
        cc_report = _empty_report(
            club_slug="creator-club", club_name="Creator Club"
        )
        cc_report.trade_lines = [
            TradeLineForMatch(
                line_id=2,
                occurred_at=occurred,
                amount=Decimal("75"),
                member_gg_player_id="1111-1111",
                member_nickname="Alice",
                sheet_row=1,
                trade_club_slug="creator-club",
            ),
        ]
        reports = {
            "round-table": rt_report,
            "clubgto": _empty_report(club_slug="clubgto", club_name="ClubGTO"),
            "creator-club": cc_report,
        }
        wb = load_workbook(io.BytesIO(build_all_clubs_matching_workbook(reports)))
        at_sheet = wb["Aces Table"]
        cc_sheet = wb["Creator Club"]
        self.assertEqual(at_sheet.cell(row=2, column=6).value, "Chip Transfer (AT↔CC)")
        self.assertEqual(at_sheet.cell(row=2, column=7).value, "Creator Club")
        self.assertEqual(cc_sheet.cell(row=2, column=6).value, "Chip Transfer (AT↔CC)")
        self.assertEqual(cc_sheet.cell(row=2, column=7).value, "Aces Table")
        at_hidden = [at_sheet.cell(row=1, column=col).value for col in range(30, 120)]
        cc_hidden = [cc_sheet.cell(row=1, column=col).value for col in range(30, 120)]
        self.assertIn("Chip Transfer (AT↔CC)", at_hidden)
        self.assertIn("Chip Transfer (AT↔CC)", cc_hidden)
        self.assertNotIn("Chip Transfer (RT↔AT)", cc_hidden)

    def test_all_clubs_cc_at_aces_fallback_matches_creator_club_payment(self):
        occurred = datetime(2026, 7, 3, 15, 0, tzinfo=timezone.utc)
        rt_report = _empty_report(club_slug="round-table", club_name="Round Table")
        rt_report.trade_lines = [
            TradeLineForMatch(
                line_id=1,
                occurred_at=occurred,
                amount=Decimal("-50"),
                member_gg_player_id="8879-5560",
                member_nickname="V",
                sheet_row=1,
                trade_club_slug="aces-table",
            ),
        ]
        cc_report = _empty_report(
            club_slug="creator-club", club_name="Creator Club"
        )
        cc_report.ledger_lines = [
            LedgerLine(
                gg_player_id="8879-5560",
                member_nickname="V",
                source="deposit_stripe",
                source_label="Stripe",
                amount_signed=Decimal("-50"),
                occurred_at_utc=occurred,
                external_id="deposit_stripe:1",
                display_name="V",
                detail="CC AT / 8879-5560 / V",
                club_slug="creator-club",
            ),
        ]
        reports = {
            "round-table": rt_report,
            "clubgto": _empty_report(club_slug="clubgto", club_name="ClubGTO"),
            "creator-club": cc_report,
        }
        wb = load_workbook(io.BytesIO(build_all_clubs_matching_workbook(reports)))
        at_sheet = wb["Aces Table"]
        self.assertEqual(at_sheet.cell(row=2, column=6).value, "Stripe")
        self.assertEqual(at_sheet.cell(row=2, column=7).value, "V")
        self.assertEqual(at_sheet.cell(row=2, column=9).value, 50.0)
        unresolved = wb["Unresolved"]
        self.assertIsNone(unresolved.cell(row=2, column=1).value)

    def test_cc_at_duplicate_in_composite_not_on_unresolved_when_matched(self):
        """CC AT payments can appear in both RT composite and CC ledgers."""
        occurred = datetime(2026, 7, 3, 15, 0, tzinfo=timezone.utc)
        cc_line = LedgerLine(
            gg_player_id="8879-5560",
            member_nickname="V",
            source="deposit_stripe",
            source_label="Stripe",
            amount_signed=Decimal("-50"),
            occurred_at_utc=occurred,
            external_id="deposit_stripe:1",
            display_name="V",
            detail="CC AT / 8879-5560 / V",
            club_slug="creator-club",
        )
        rt_report = _empty_report(club_slug="round-table", club_name="Round Table")
        rt_report.trade_lines = [
            TradeLineForMatch(
                line_id=1,
                occurred_at=occurred,
                amount=Decimal("-50"),
                member_gg_player_id="8879-5560",
                member_nickname="V",
                sheet_row=1,
                trade_club_slug="aces-table",
            ),
        ]
        rt_report.ledger_lines = [replace(cc_line, club_slug="aces-table")]
        cc_report = _empty_report(
            club_slug="creator-club", club_name="Creator Club"
        )
        cc_report.ledger_lines = [cc_line]
        reports = {
            "round-table": rt_report,
            "clubgto": _empty_report(club_slug="clubgto", club_name="ClubGTO"),
            "creator-club": cc_report,
        }
        wb = load_workbook(io.BytesIO(build_all_clubs_matching_workbook(reports)))
        at_sheet = wb["Aces Table"]
        self.assertEqual(at_sheet.cell(row=2, column=6).value, "Stripe")
        unresolved = wb["Unresolved"]
        self.assertIsNone(unresolved.cell(row=2, column=1).value)

    def test_rt_at_duplicate_ledger_not_on_unresolved_when_matched(self):
        """Partner double-fetch can leave a second copy in unmatched_ledger."""
        occurred = datetime(2026, 7, 3, 15, 0, tzinfo=timezone.utc)
        rt_line = LedgerLine(
            gg_player_id="1111-1111",
            member_nickname="RtPlayer",
            source="deposit_zelle",
            source_label="RT Zelle",
            amount_signed=Decimal("-50"),
            occurred_at_utc=occurred,
            external_id="deposit_zelle:1",
            display_name="RtPlayer",
            detail="RT / 1111-1111 / RtPlayer",
            club_slug="round-table",
        )
        rt_report = _empty_report(club_slug="round-table", club_name="Round Table")
        rt_report.trade_lines = [
            TradeLineForMatch(
                line_id=1,
                occurred_at=occurred,
                amount=Decimal("-50"),
                member_gg_player_id="1111-1111",
                member_nickname="RtPlayer",
                sheet_row=1,
                trade_club_slug="round-table",
            ),
        ]
        rt_report.ledger_lines = [
            rt_line,
            replace(rt_line, club_slug="aces-table"),
        ]
        reports = {
            "round-table": rt_report,
            "clubgto": _empty_report(club_slug="clubgto", club_name="ClubGTO"),
            "creator-club": _empty_report(
                club_slug="creator-club", club_name="Creator Club"
            ),
        }
        wb = load_workbook(io.BytesIO(build_all_clubs_matching_workbook(reports)))
        self.assertEqual(wb["Round Table"].cell(row=2, column=6).value, "RT Zelle")
        unresolved = wb["Unresolved"]
        self.assertIsNone(unresolved.cell(row=2, column=1).value)


if __name__ == "__main__":
    unittest.main()
