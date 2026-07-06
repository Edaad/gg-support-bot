"""Tests for trade record XLSX parser."""

from __future__ import annotations

import unittest
from datetime import date

from api.trade_record_parser import (
    TradeRecordParseError,
    TradeRecordValidationError,
    extract_audit_date_from_metadata,
    parse_trade_record_workbook,
    resolve_club_slug_from_metadata,
    validate_trade_upload_pair,
)
from tests.fixtures.trade_record_xlsx import build_sample_trade_record_xlsx


class TradeRecordParserTestCase(unittest.TestCase):
    def test_parse_extracts_identities_and_transactions(self):
        raw = build_sample_trade_record_xlsx(audit_date=date(2026, 6, 21))
        parsed = parse_trade_record_workbook(raw)

        self.assertEqual(parsed.audit_date, date(2026, 6, 21))
        self.assertEqual(parsed.club_slug, "aces-table")
        self.assertEqual(len(parsed.transactions), 2)
        self.assertEqual(parsed.transactions[0].member_gg_player_id, "3011-9668")
        self.assertEqual(parsed.transactions[0].amount, 100)
        self.assertEqual(parsed.transactions[1].amount, -50)

        ids = {(i.role, i.gg_player_id) for i in parsed.identities}
        self.assertIn(("member", "3011-9668"), ids)
        self.assertIn(("agent", "2000-2001"), ids)
        self.assertIn(("superAgent", "1000-1001"), ids)
        self.assertEqual(parsed.transactions[0].member_nickname, "MemberOne")

    def test_extracts_club_and_date_from_metadata(self):
        raw = build_sample_trade_record_xlsx(
            club_label="Aces Table",
            audit_date=date(2026, 6, 21),
        )
        parsed = parse_trade_record_workbook(raw)

        self.assertEqual(parsed.metadata.club_text, "Aces Table")
        self.assertEqual(parsed.metadata.club_id_text, "983183")
        self.assertIn("2026-06-21", parsed.metadata.date_text)
        self.assertEqual(
            extract_audit_date_from_metadata(parsed.metadata),
            date(2026, 6, 21),
        )
        self.assertEqual(resolve_club_slug_from_metadata(parsed.metadata), "aces-table")

    def test_rejects_unknown_club(self):
        raw = build_sample_trade_record_xlsx(club_label="Unknown Club")
        with self.assertRaises(TradeRecordValidationError):
            parse_trade_record_workbook(raw)

    def test_extract_audit_date_accepts_period_range(self):
        raw = build_sample_trade_record_xlsx(
            club_label="Aces Table",
            audit_date=date(2024, 6, 21),
        )
        parsed = parse_trade_record_workbook(raw)
        self.assertEqual(parsed.audit_date, date(2024, 6, 21))
        self.assertIn("2024-06-21", parsed.metadata.date_text)

    def test_rejects_multi_day_period(self):
        from openpyxl import Workbook
        import io

        from api.trade_record_parser import SHEET_NAME

        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.cell(row=1, column=1, value="Club Name")
        ws.cell(row=1, column=2, value="Aces Table")
        ws.cell(row=3, column=1, value="Period")
        ws.cell(row=3, column=2, value="2024-06-20 ~ 2024-06-21 (UTC-5:00)")
        buf = io.BytesIO()
        wb.save(buf)
        with self.assertRaises(TradeRecordValidationError):
            parse_trade_record_workbook(buf.getvalue())

    def test_occurred_at_stored_as_utc_for_fixed_offset_club(self):
        raw = build_sample_trade_record_xlsx(
            club_label="Aces Table",
            audit_date=date(2026, 6, 21),
        )
        parsed = parse_trade_record_workbook(raw)
        occurred = parsed.transactions[0].occurred_at
        self.assertIsNotNone(occurred)
        assert occurred is not None
        from datetime import timezone

        self.assertEqual(occurred.tzinfo, timezone.utc)
        self.assertEqual(occurred.hour, 19)
        self.assertEqual(occurred.minute, 30)

    def test_missing_sheet_raises(self):
        from openpyxl import Workbook
        import io

        wb = Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        with self.assertRaises(TradeRecordParseError):
            parse_trade_record_workbook(buf.getvalue())

    def test_validate_trade_upload_pair_matching_dates(self):
        rt_raw = build_sample_trade_record_xlsx(
            club_label="Round Table",
            audit_date=date(2026, 6, 21),
            period_tz="UTC-4:00",
        )
        at_raw = build_sample_trade_record_xlsx(
            club_label="Aces Table",
            audit_date=date(2026, 6, 21),
        )
        rt = parse_trade_record_workbook(rt_raw)
        at = parse_trade_record_workbook(at_raw)
        validate_trade_upload_pair(rt, at)

    def test_validate_trade_upload_pair_rejects_mismatched_dates(self):
        rt_raw = build_sample_trade_record_xlsx(
            club_label="Round Table",
            audit_date=date(2026, 6, 21),
            period_tz="UTC-4:00",
        )
        at_raw = build_sample_trade_record_xlsx(
            club_label="Aces Table",
            audit_date=date(2026, 6, 22),
        )
        rt = parse_trade_record_workbook(rt_raw)
        at = parse_trade_record_workbook(at_raw)
        with self.assertRaises(TradeRecordValidationError):
            validate_trade_upload_pair(rt, at)


if __name__ == "__main__":
    unittest.main()
