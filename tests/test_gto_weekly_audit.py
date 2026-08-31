"""Tests for GTO weekly audit export."""

from __future__ import annotations

import io
import os
import unittest
import zipfile
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from unittest.mock import MagicMock, patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from api.audit_reconcile_export import MATCHING_HEADERS
from api.auth import create_token, get_current_admin
from api.gto_weekly_audit import (
    BonusRailRow,
    GtoWeeklyAuditError,
    PaymentRailRow,
    _fetch_vaughn_payments_for_day,
    build_gto_weekly_audit_workbook,
    date_from_filename,
    expected_week_dates,
    fetch_clubgto_bonus_rails,
    output_filename,
    parse_clubgto_rows,
    rail_bucket,
    validate_upload_set,
    _zelle_variant,
)
from api.routes.audit import router
from db.connection import get_db_dependency
from db.models import ZellePayment

TOKEN = create_token()
MONDAY = date(2026, 8, 10)


def _matching_xlsx(
    rows: list[tuple],
    *,
    sheet_name: str = "ClubGTO",
    headers: list[str] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    use_headers = headers if headers is not None else list(MATCHING_HEADERS)
    for col, h in enumerate(use_headers, start=1):
        ws.cell(1, col, h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    for title in ("Round Table", "Aces Table", "Creator Club"):
        if title not in wb.sheetnames:
            wb.create_sheet(title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _week_files(
    monday: date = MONDAY,
    *,
    row_factory=None,
) -> list[tuple[str, bytes]]:
    files: list[tuple[str, bytes]] = []
    for i in range(7):
        day = monday + timedelta(days=i)
        if row_factory is None:
            rows = [
                (
                    datetime(day.year, day.month, day.day, 10, 0, 0),
                    "TrafficLight7",
                    -50.0,
                    "1111-0001",
                    "PlayerA",
                    "GTO Zelle",
                    "Jane Doe",
                    datetime(day.year, day.month, day.day, 9, 30, 0),
                    50.0,
                    "ClubGTO",
                )
            ]
        else:
            rows = row_factory(day, i)
        name = f"reconcile-all-clubs-{day.isoformat()}.xlsx"
        files.append((name, _matching_xlsx(rows)))
    return files


def _mock_session() -> MagicMock:
    return MagicMock()


def _empty_payment_rails() -> dict[str, list[PaymentRailRow]]:
    return {"zelle": [], "venmo": [], "crypto": []}


def _processed_row_factory(day: date, i: int) -> list[tuple]:
    base = datetime(day.year, day.month, day.day, 12, 0, 0)
    if i == 0:
        return [
            (
                base,
                "Admin1",
                -40.0,
                "1000-0001",
                "Alice",
                "GTO Zelle",
                "Alice Payer",
                base.replace(hour=11),
                40.0,
                "ClubGTO",
            ),
            (
                base.replace(minute=1),
                "Admin1",
                -30.0,
                "1000-0002",
                "Bob",
                "RT Zelle",
                "Bob Payer",
                base.replace(hour=10),
                30.0,
                "Other",
            ),
            (
                base.replace(minute=2),
                "Admin1",
                25.0,
                "1000-0003",
                "Cara",
                "Cashout Zelle",
                "Cara",
                base.replace(hour=9),
                25.0,
                "ClubGTO",
            ),
            (
                base.replace(minute=3),
                "Admin1",
                -5.0,
                "1000-0004",
                "Dan",
                "Bonus",
                "Dan Nick",
                base.replace(hour=8),
                5.0,
                "",
            ),
            (
                base.replace(minute=4),
                "Admin1",
                -20.0,
                "1000-0005",
                "Eve",
                "Vaughn venmo",
                "Eve Payer",
                None,
                None,
                "@Janseashells",
            ),
        ]
    if i == 1:
        return [
            (
                base,
                "Admin2",
                -100.0,
                "2000-0001",
                "Frank",
                "GTO Crypto",
                "bc1qabc…",
                base.replace(hour=7),
                100.0,
                "BTC",
            )
        ]
    return [
        (
            base,
            "AdminX",
            -10.0,
            f"3000-000{i}",
            f"P{i}",
            "GTO Stripe",
            "",
            None,
            None,
            "",
        )
    ]


class GtoWeeklyAuditUnitTestCase(unittest.TestCase):
    def test_expected_week_dates_rejects_non_monday(self):
        with self.assertRaises(GtoWeeklyAuditError) as ctx:
            expected_week_dates(date(2026, 8, 11))
        self.assertIn("Monday", str(ctx.exception))

    def test_validate_upload_set_ok(self):
        names = [
            f"reconcile-all-clubs-{(MONDAY + timedelta(days=i)).isoformat()}.xlsx"
            for i in range(7)
        ]
        self.assertEqual(validate_upload_set(MONDAY, names), expected_week_dates(MONDAY))

    def test_validate_wrong_count(self):
        with self.assertRaises(GtoWeeklyAuditError) as ctx:
            validate_upload_set(MONDAY, ["reconcile-all-clubs-2026-08-10.xlsx"])
        self.assertIn("exactly 7", str(ctx.exception))

    def test_validate_bad_filename(self):
        with self.assertRaises(GtoWeeklyAuditError):
            date_from_filename("reconcile-all-clubs-2026-08-10 (1).xlsx")

    def test_validate_wrong_dates(self):
        names = [
            f"reconcile-all-clubs-{(date(2026, 8, 11) + timedelta(days=i)).isoformat()}.xlsx"
            for i in range(7)
        ]
        with self.assertRaises(GtoWeeklyAuditError) as ctx:
            validate_upload_set(MONDAY, names)
        self.assertIn("Missing", str(ctx.exception))

    def test_rail_bucket_rules(self):
        self.assertEqual(rail_bucket("GTO Zelle"), "zelle")
        self.assertEqual(rail_bucket("Vaughn venmo"), "venmo")
        self.assertEqual(rail_bucket("GTO Crypto"), "crypto")
        self.assertEqual(rail_bucket("Bonus"), "bonuses")
        self.assertEqual(rail_bucket("Admin Bonus Offer"), "bonuses")
        self.assertIsNone(rail_bucket("RT Zelle"))
        self.assertIsNone(rail_bucket("Cashout Zelle"))
        self.assertIsNone(rail_bucket("GTO Stripe"))
        self.assertIsNone(rail_bucket("rb"))

    def test_output_filename(self):
        self.assertEqual(output_filename(MONDAY), "GTO Audit Aug10_16-2026.xlsx")

    def test_parse_missing_sheet(self):
        wb = Workbook()
        wb.active.title = "Round Table"
        with self.assertRaises(GtoWeeklyAuditError) as ctx:
            parse_clubgto_rows(wb, filename="reconcile-all-clubs-2026-08-10.xlsx")
        self.assertIn("ClubGTO", str(ctx.exception))

    def test_parse_zero_rows(self):
        raw = _matching_xlsx([])
        wb = load_workbook(io.BytesIO(raw))
        with self.assertRaises(GtoWeeklyAuditError) as ctx:
            parse_clubgto_rows(wb, filename="reconcile-all-clubs-2026-08-10.xlsx")
        self.assertIn("no data rows", str(ctx.exception))

    def test_parse_missing_headers(self):
        raw = _matching_xlsx(
            [("x",)],
            headers=["Trade Time"],
        )
        wb = load_workbook(io.BytesIO(raw))
        with self.assertRaises(GtoWeeklyAuditError) as ctx:
            parse_clubgto_rows(wb, filename="reconcile-all-clubs-2026-08-10.xlsx")
        self.assertIn("missing required headers", str(ctx.exception))

    def test_zelle_variant_canonicalizes_bank_label(self):
        self.assertEqual(
            _zelle_variant({"zelle_recipient": "Citizens V"}),
            "starship5vllc@gmail.com",
        )

    @patch("api.gto_weekly_audit.fetch_clubgto_bonus_rails", return_value=[])
    @patch(
        "api.gto_weekly_audit.fetch_vaughn_payment_rails",
        return_value=_empty_payment_rails(),
    )
    def test_build_workbook_processed(self, _mock_payments, _mock_bonuses):
        files = _week_files(MONDAY, row_factory=_processed_row_factory)
        content = build_gto_weekly_audit_workbook(
            MONDAY, files, session=_mock_session()
        )
        wb = load_workbook(io.BytesIO(content))
        self.assertEqual(
            wb.sheetnames[:5],
            ["Processed", "Zelle", "Venmo", "Crypto", "Bonuses"],
        )

        processed = wb["Processed"]
        data_rows = []
        for r in range(2, processed.max_row + 1):
            if processed.cell(r, 6).value and processed.cell(r, 6).value != "Missing data":
                data_rows.append(
                    (
                        processed.cell(r, 2).value,
                        processed.cell(r, 4).value,
                        processed.cell(r, 6).value,
                        processed.cell(r, 7).value,
                    )
                )
        self.assertEqual(len(data_rows), 11)
        self.assertTrue(all(row[3] for row in data_rows))
        self.assertIn("ProcessedData", processed.tables)

        times = [
            processed.cell(r, 1).value
            for r in range(2, processed.max_row + 1)
            if isinstance(processed.cell(r, 1).value, datetime)
        ]
        self.assertEqual(times, sorted(times))

        stripe_processed = [
            r
            for r in range(2, processed.max_row + 1)
            if processed.cell(r, 6).value == "GTO Stripe"
        ]
        self.assertGreaterEqual(len(stripe_processed), 1)

        with zipfile.ZipFile(io.BytesIO(content)) as z:
            pivot_parts = [n for n in z.namelist() if "pivotTables/" in n]
            self.assertTrue(pivot_parts, "expected pivot table part in output")

    @patch("api.gto_weekly_audit.fetch_clubgto_bonus_rails")
    @patch("api.gto_weekly_audit.fetch_vaughn_payment_rails")
    def test_build_workbook_rails_from_db(self, mock_payments, mock_bonuses):
        day1 = MONDAY + timedelta(days=1)
        mock_payments.return_value = {
            "zelle": [
                PaymentRailRow(
                    audit_date=MONDAY,
                    occurred_at=datetime(2026, 8, 10, 9, 0),
                    name="Alice Payer",
                    variant="2133729202",
                    amount_usd=40.0,
                ),
                PaymentRailRow(
                    audit_date=day1,
                    occurred_at=datetime(2026, 8, 11, 8, 0),
                    name="Later Payer",
                    variant="3105670961",
                    amount_usd=15.0,
                ),
            ],
            "venmo": [
                PaymentRailRow(
                    audit_date=MONDAY,
                    occurred_at=datetime(2026, 8, 10, 12, 4),
                    name="Eve Payer",
                    variant="@janseashells",
                    amount_usd=20.0,
                ),
            ],
            "crypto": [
                PaymentRailRow(
                    audit_date=day1,
                    occurred_at=datetime(2026, 8, 11, 7, 0),
                    name="bc1qabc…",
                    variant="BTC",
                    amount_usd=100.0,
                ),
            ],
        }
        mock_bonuses.return_value = [
            BonusRailRow(
                audit_date=MONDAY,
                occurred_at=datetime(2026, 8, 10, 8, 0),
                player="danplayer",
                amount_usd=5.0,
            ),
        ]

        files = _week_files(MONDAY)
        content = build_gto_weekly_audit_workbook(
            MONDAY, files, session=_mock_session()
        )
        wb = load_workbook(io.BytesIO(content))

        zelle = wb["Zelle"]
        self.assertEqual(zelle.cell(2, 2).value, "Alice Payer")
        self.assertEqual(zelle.cell(2, 5).value, MONDAY.isoformat())
        self.assertEqual(zelle.cell(3, 2).value, "Later Payer")
        self.assertEqual(zelle.cell(3, 5).value, day1.isoformat())
        self.assertEqual(zelle.cell(4, 3).value, 55.0)

        venmo = wb["Venmo"]
        self.assertEqual(venmo.cell(2, 2).value, "Eve Payer")
        self.assertEqual(venmo.cell(2, 4).value, "@janseashells")
        self.assertEqual(venmo.cell(3, 3).value, 20.0)

        crypto = wb["Crypto"]
        self.assertEqual(crypto.cell(2, 2).value, "bc1qabc…")
        self.assertEqual(crypto.cell(2, 4).value, "BTC")
        self.assertEqual(crypto.cell(2, 5).value, day1.isoformat())

        bonuses = wb["Bonuses"]
        self.assertEqual(bonuses.cell(2, 2).value, "danplayer")
        self.assertEqual(bonuses.cell(2, 4).value, MONDAY.isoformat())
        self.assertEqual(bonuses.cell(3, 3).value, 5.0)


class GtoWeeklyAuditFetchTestCase(unittest.TestCase):
    @patch("api.gto_weekly_audit.payment_in_audit_day_for_club", return_value=True)
    @patch("api.gto_weekly_audit._apply_audit_manual_filters")
    def test_fetch_vaughn_zelle_included(self, mock_filters, _mock_audit_day):
        payment = MagicMock()
        query = MagicMock()
        query.filter.return_value = query
        query.order_by.return_value = query
        query.all.return_value = [payment]
        session = MagicMock()
        session.query.return_value = query
        mock_filters.return_value = query
        mock_build = MagicMock(
            return_value={
                "payer_name": "Jane",
                "zelle_recipient": "2133729202",
                "amount_usd": "50.00",
                "created_at": datetime(2026, 8, 10, 14, 0, tzinfo=timezone.utc),
                "club_id": 2,
            }
        )

        rows = _fetch_vaughn_payments_for_day(
            session,
            payment_cls=ZellePayment,
            build_read=mock_build,
            audit_date=MONDAY,
            name_fn=lambda d: (d.get("payer_name") or "").strip(),
            variant_fn=_zelle_variant,
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].name, "Jane")
        self.assertEqual(rows[0].variant, "2133729202")
        self.assertEqual(rows[0].amount_usd, 50.0)
        query.filter.assert_called()

    @patch("api.gto_weekly_audit.payment_in_audit_day_for_club", return_value=False)
    @patch("api.gto_weekly_audit._apply_audit_manual_filters")
    def test_fetch_vaughn_zelle_excluded_when_not_clubgto_audit_day(
        self, mock_filters, _mock_audit_day
    ):
        payment = MagicMock()
        query = MagicMock()
        query.filter.return_value = query
        query.order_by.return_value = query
        query.all.return_value = [payment]
        session = MagicMock()
        session.query.return_value = query
        mock_filters.return_value = query
        mock_build = MagicMock(
            return_value={
                "payer_name": "Jane",
                "zelle_recipient": "2133729202",
                "amount_usd": "50.00",
                "created_at": datetime(2026, 8, 10, 14, 0, tzinfo=timezone.utc),
                "club_id": 1,
            }
        )

        rows = _fetch_vaughn_payments_for_day(
            session,
            payment_cls=ZellePayment,
            build_read=mock_build,
            audit_date=MONDAY,
            name_fn=lambda d: (d.get("payer_name") or "").strip(),
            variant_fn=_zelle_variant,
        )
        self.assertEqual(rows, [])

    @patch("api.gto_weekly_audit.payment_in_audit_day_for_club", return_value=True)
    @patch("api.gto_weekly_audit.resolve_club_id", return_value=2)
    def test_fetch_clubgto_bonus_rails(self, _mock_club_id, _mock_audit_day):
        record = MagicMock()
        record.club_id = 2
        record.player_username = "bonus_player"
        record.amount = Decimal("12.50")
        record.created_at = datetime(2026, 8, 10, 16, 0, tzinfo=timezone.utc)

        query = MagicMock()
        query.filter.return_value = query
        query.order_by.return_value = query
        query.all.return_value = [record]
        session = MagicMock()
        session.query.return_value = query

        rows = fetch_clubgto_bonus_rails(session, [MONDAY])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].player, "bonus_player")
        self.assertEqual(rows[0].amount_usd, 12.5)
        self.assertEqual(rows[0].audit_date, MONDAY)


def _creator_week_files(monday: date = MONDAY) -> list[tuple[str, bytes]]:
    files: list[tuple[str, bytes]] = []
    for i in range(7):
        day = monday + timedelta(days=i)
        rows = [
            (
                datetime(day.year, day.month, day.day, 10, 0, 0),
                "TrafficLight7",
                -50.0,
                "1111-0001",
                "PlayerA",
                "Mateos Zelle",
                "Jane Doe",
                datetime(day.year, day.month, day.day, 9, 30, 0),
                50.0,
                "@mateos-handle",
            )
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Creator Club"
        for col, h in enumerate(MATCHING_HEADERS, start=1):
            ws.cell(1, col, h)
        for c_idx, value in enumerate(rows[0], start=1):
            ws.cell(2, c_idx, value)
        for title in ("Round Table", "Aces Table", "ClubGTO"):
            wb.create_sheet(title)
        buf = io.BytesIO()
        wb.save(buf)
        files.append((f"reconcile-all-clubs-{day.isoformat()}.xlsx", buf.getvalue()))
    return files


class PartnerWeeklyAuditApiTestCase(unittest.TestCase):
    def setUp(self):
        self.env_patch = patch.dict(
            os.environ, {"DASHBOARD_PASSWORD": "changeme"}, clear=False
        )
        self.env_patch.start()
        app = FastAPI()
        app.include_router(router)
        app.dependency_overrides[get_current_admin] = lambda: "admin"

        def override_db():
            yield MagicMock()

        app.dependency_overrides[get_db_dependency] = override_db
        self.client = TestClient(app)

    def tearDown(self):
        self.env_patch.stop()

    def _upload_files(self, files: list[tuple[str, bytes]]) -> list:
        return [
            (
                "files",
                (name, raw, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            )
            for name, raw in files
        ]

    @patch("api.gto_weekly_audit.fetch_clubgto_bonus_rails", return_value=[])
    @patch(
        "api.gto_weekly_audit.fetch_vaughn_payment_rails",
        return_value=_empty_payment_rails(),
    )
    def test_export_gto_ok(self, _mock_payments, _mock_bonuses):
        files = _week_files(MONDAY)
        res = self.client.post(
            "/api/audit/partner-weekly-audit/export",
            data={"monday": MONDAY.isoformat(), "club": "clubgto"},
            files=self._upload_files(files),
            headers={"Authorization": f"Bearer {TOKEN}"},
        )
        self.assertEqual(res.status_code, 200, res.text)
        self.assertIn(
            "GTO Audit Aug10_16-2026.xlsx",
            res.headers.get("content-disposition", ""),
        )
        wb = load_workbook(io.BytesIO(res.content))
        self.assertEqual(wb.sheetnames[0], "Processed")

    @patch("api.creator_weekly_audit.fetch_creator_club_bonus_rails", return_value=[])
    @patch(
        "api.creator_weekly_audit.fetch_mateos_payment_rails",
        return_value=_empty_payment_rails(),
    )
    def test_export_creator_club_ok(self, _mock_payments, _mock_bonuses):
        files = _creator_week_files(MONDAY)
        res = self.client.post(
            "/api/audit/partner-weekly-audit/export",
            data={"monday": MONDAY.isoformat(), "club": "creator-club"},
            files=self._upload_files(files),
            headers={"Authorization": f"Bearer {TOKEN}"},
        )
        self.assertEqual(res.status_code, 200, res.text)
        self.assertIn(
            "Creator Audit Aug10_16-2026.xlsx",
            res.headers.get("content-disposition", ""),
        )
        wb = load_workbook(io.BytesIO(res.content))
        self.assertEqual(wb.sheetnames[0], "Processed")

    def test_export_invalid_club(self):
        files = _week_files(MONDAY)
        res = self.client.post(
            "/api/audit/partner-weekly-audit/export",
            data={"monday": MONDAY.isoformat(), "club": "round-table"},
            files=self._upload_files(files),
            headers={"Authorization": f"Bearer {TOKEN}"},
        )
        self.assertEqual(res.status_code, 400)
        self.assertIn("club must be one of", res.json()["detail"])

    def test_export_validation_error(self):
        files = _week_files(MONDAY)[:3]
        res = self.client.post(
            "/api/audit/partner-weekly-audit/export",
            data={"monday": MONDAY.isoformat(), "club": "clubgto"},
            files=self._upload_files(files),
            headers={"Authorization": f"Bearer {TOKEN}"},
        )
        self.assertEqual(res.status_code, 400)
        self.assertIn("exactly 7", res.json()["detail"])


if __name__ == "__main__":
    unittest.main()
