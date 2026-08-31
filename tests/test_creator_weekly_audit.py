"""Tests for Creator Club weekly audit export."""

from __future__ import annotations

import io
import unittest
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook

from api.audit_reconcile_export import MATCHING_HEADERS
from api.creator_weekly_audit import (
    BonusRailRow,
    CreatorWeeklyAuditError,
    PaymentRailRow,
    _fetch_mateos_payments_for_day,
    build_creator_weekly_audit_workbook,
    expected_week_dates,
    fetch_creator_club_bonus_rails,
    output_filename,
    parse_creator_club_rows,
    validate_upload_set,
    _zelle_variant,
)
from db.models import ZellePayment

MONDAY = date(2026, 8, 10)


def _matching_xlsx(
    rows: list[tuple],
    *,
    sheet_name: str = "Creator Club",
    headers: list[str] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    use_headers = headers if headers is not None else list(MATCHING_HEADERS)
    for col, h in enumerate(use_headers, start=1):
        ws.cell(1, col, h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, value)
    for title in ("Round Table", "Aces Table", "ClubGTO"):
        if title not in wb.sheetnames:
            wb.create_sheet(title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _creator_week_files(
    monday: date = MONDAY,
    *,
    row_factory=None,
) -> list[tuple[str, bytes]]:
    files: list[tuple[str, bytes]] = []
    for i in range(7):
        day = monday + timedelta(days=i)
        if row_factory is None:
            rows = [
                (
                    datetime(day.year, day.month, day.day, 10, 0, 0),
                    "TrafficLight7",
                    -50.0,
                    "1111-0001",
                    "PlayerA",
                    "Mateos Zelle",
                    "Jane Doe",
                    datetime(day.year, day.month, day.day, 9, 30, 0),
                    50.0,
                    "@mateos-handle",
                )
            ]
        else:
            rows = row_factory(day, i)
        name = f"reconcile-all-clubs-{day.isoformat()}.xlsx"
        files.append((name, _matching_xlsx(rows)))
    return files


def _mock_session() -> MagicMock:
    return MagicMock()


def _empty_payment_rails() -> dict[str, list[PaymentRailRow]]:
    return {"zelle": [], "venmo": [], "crypto": []}


class CreatorWeeklyAuditUnitTestCase(unittest.TestCase):
    def test_expected_week_dates_rejects_non_monday(self):
        with self.assertRaises(CreatorWeeklyAuditError) as ctx:
            expected_week_dates(date(2026, 8, 11))
        self.assertIn("Monday", str(ctx.exception))

    def test_validate_upload_set_ok(self):
        names = [
            f"reconcile-all-clubs-{(MONDAY + timedelta(days=i)).isoformat()}.xlsx"
            for i in range(7)
        ]
        self.assertEqual(validate_upload_set(MONDAY, names), expected_week_dates(MONDAY))

    def test_output_filename(self):
        self.assertEqual(output_filename(MONDAY), "Creator Audit Aug10_16-2026.xlsx")

    def test_parse_missing_sheet(self):
        wb = Workbook()
        wb.active.title = "Round Table"
        with self.assertRaises(CreatorWeeklyAuditError) as ctx:
            parse_creator_club_rows(wb, filename="reconcile-all-clubs-2026-08-10.xlsx")
        self.assertIn("Creator Club", str(ctx.exception))

    def test_parse_zero_rows(self):
        raw = _matching_xlsx([])
        wb = load_workbook(io.BytesIO(raw))
        with self.assertRaises(CreatorWeeklyAuditError) as ctx:
            parse_creator_club_rows(wb, filename="reconcile-all-clubs-2026-08-10.xlsx")
        self.assertIn("no data rows", str(ctx.exception))

    def test_parse_missing_headers(self):
        raw = _matching_xlsx(
            [("x",)],
            headers=["Trade Time"],
        )
        wb = load_workbook(io.BytesIO(raw))
        with self.assertRaises(CreatorWeeklyAuditError) as ctx:
            parse_creator_club_rows(wb, filename="reconcile-all-clubs-2026-08-10.xlsx")
        self.assertIn("missing required headers", str(ctx.exception))

    @patch("api.creator_weekly_audit.fetch_creator_club_bonus_rails", return_value=[])
    @patch(
        "api.creator_weekly_audit.fetch_mateos_payment_rails",
        return_value=_empty_payment_rails(),
    )
    def test_build_workbook_processed(self, _mock_payments, _mock_bonuses):
        files = _creator_week_files(MONDAY)
        content = build_creator_weekly_audit_workbook(
            MONDAY, files, session=_mock_session()
        )
        wb = load_workbook(io.BytesIO(content))
        self.assertEqual(
            wb.sheetnames[:5],
            ["Processed", "Zelle", "Venmo", "Crypto", "Bonuses"],
        )
        processed = wb["Processed"]
        self.assertIn("ProcessedData", processed.tables)
        self.assertEqual(processed.cell(2, 6).value, "Mateos Zelle")

    @patch("api.creator_weekly_audit.fetch_creator_club_bonus_rails")
    @patch("api.creator_weekly_audit.fetch_mateos_payment_rails")
    def test_build_workbook_rails_from_db(self, mock_payments, mock_bonuses):
        day1 = MONDAY + timedelta(days=1)
        mock_payments.return_value = {
            "zelle": [
                PaymentRailRow(
                    audit_date=MONDAY,
                    occurred_at=datetime(2026, 8, 10, 9, 0),
                    name="Alice Payer",
                    variant="2133729202",
                    amount_usd=40.0,
                ),
            ],
            "venmo": [
                PaymentRailRow(
                    audit_date=day1,
                    occurred_at=datetime(2026, 8, 11, 12, 4),
                    name="Eve Payer",
                    variant="@mateos-handle",
                    amount_usd=20.0,
                ),
            ],
            "crypto": [],
        }
        mock_bonuses.return_value = [
            BonusRailRow(
                audit_date=MONDAY,
                occurred_at=datetime(2026, 8, 10, 8, 0),
                player="cc_player",
                amount_usd=5.0,
            ),
        ]

        files = _creator_week_files(MONDAY)
        content = build_creator_weekly_audit_workbook(
            MONDAY, files, session=_mock_session()
        )
        wb = load_workbook(io.BytesIO(content))

        self.assertEqual(wb["Zelle"].cell(2, 2).value, "Alice Payer")
        self.assertEqual(wb["Venmo"].cell(2, 4).value, "@mateos-handle")
        self.assertEqual(wb["Bonuses"].cell(2, 2).value, "cc_player")


class CreatorWeeklyAuditFetchTestCase(unittest.TestCase):
    @patch("api.creator_weekly_audit.payment_in_audit_day_for_club", return_value=True)
    @patch("api.creator_weekly_audit._apply_audit_manual_filters")
    def test_fetch_mateos_zelle_included(self, mock_filters, _mock_audit_day):
        payment = MagicMock()
        query = MagicMock()
        query.filter.return_value = query
        query.order_by.return_value = query
        query.all.return_value = [payment]
        session = MagicMock()
        session.query.return_value = query
        mock_filters.return_value = query
        mock_build = MagicMock(
            return_value={
                "payer_name": "Jane",
                "zelle_recipient": "2133729202",
                "amount_usd": "50.00",
                "created_at": datetime(2026, 8, 10, 14, 0, tzinfo=timezone.utc),
                "club_id": 3,
            }
        )

        rows = _fetch_mateos_payments_for_day(
            session,
            payment_cls=ZellePayment,
            build_read=mock_build,
            audit_date=MONDAY,
            name_fn=lambda d: (d.get("payer_name") or "").strip(),
            variant_fn=_zelle_variant,
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].name, "Jane")
        self.assertEqual(rows[0].amount_usd, 50.0)
        query.filter.assert_called()

    @patch("api.creator_weekly_audit.payment_in_audit_day_for_club", return_value=False)
    @patch("api.creator_weekly_audit._apply_audit_manual_filters")
    def test_fetch_mateos_zelle_excluded_when_not_creator_audit_day(
        self, mock_filters, _mock_audit_day
    ):
        payment = MagicMock()
        query = MagicMock()
        query.filter.return_value = query
        query.order_by.return_value = query
        query.all.return_value = [payment]
        session = MagicMock()
        session.query.return_value = query
        mock_filters.return_value = query
        mock_build = MagicMock(
            return_value={
                "payer_name": "Jane",
                "zelle_recipient": "2133729202",
                "amount_usd": "50.00",
                "created_at": datetime(2026, 8, 10, 14, 0, tzinfo=timezone.utc),
                "club_id": 1,
            }
        )

        rows = _fetch_mateos_payments_for_day(
            session,
            payment_cls=ZellePayment,
            build_read=mock_build,
            audit_date=MONDAY,
            name_fn=lambda d: (d.get("payer_name") or "").strip(),
            variant_fn=_zelle_variant,
        )
        self.assertEqual(rows, [])

    @patch("api.creator_weekly_audit.payment_in_audit_day_for_club", return_value=True)
    @patch("api.creator_weekly_audit.resolve_club_id", return_value=3)
    def test_fetch_creator_club_bonus_rails(self, _mock_club_id, _mock_audit_day):
        record = MagicMock()
        record.club_id = 3
        record.player_username = "bonus_player"
        record.amount = Decimal("12.50")
        record.created_at = datetime(2026, 8, 10, 16, 0, tzinfo=timezone.utc)

        query = MagicMock()
        query.filter.return_value = query
        query.order_by.return_value = query
        query.all.return_value = [record]
        session = MagicMock()
        session.query.return_value = query

        rows = fetch_creator_club_bonus_rails(session, [MONDAY])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].player, "bonus_player")
        self.assertEqual(rows[0].amount_usd, 12.5)
        self.assertEqual(rows[0].audit_date, MONDAY)


if __name__ == "__main__":
    unittest.main()
