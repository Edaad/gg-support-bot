"""Tests for early-rakeback sync from aon-beta into Postgres."""

from __future__ import annotations

import os
import unittest
from datetime import date, datetime, timezone
from decimal import Decimal
from unittest.mock import MagicMock, patch

from api.early_rakeback_sync import (
    _flatten_archive_entries,
    backfill_early_rakeback_from_archives,
    sync_early_rakeback_for_date,
)
from db.models import Club, EarlyRakebackLine, EarlyRakebackSnapshot


def _sample_entries() -> list[dict]:
    return [
        {
            "_id": "entry1",
            "memberNickname": "Alice",
            "memberType": "player",
            "gg_player_id": "3011-9668",
            "records": [
                {
                    "_id": "rec1",
                    "rake": 100,
                    "pl": -50,
                    "rakebackPercentage": 0.5,
                    "calculatedAmount": 25.5,
                    "timestamp": "2026-06-19T12:00:00.000Z",
                }
            ],
        },
        {
            "_id": "entry2",
            "memberNickname": "Unknown",
            "memberType": "player",
            "gg_player_id": None,
            "records": [
                {
                    "_id": "rec2",
                    "rake": 10,
                    "pl": 0,
                    "rakebackPercentage": 0.5,
                    "calculatedAmount": 5,
                    "timestamp": "2026-06-19T13:00:00.000Z",
                }
            ],
        },
    ]


class EarlyRakebackSyncTestCase(unittest.TestCase):
    def setUp(self):
        self.club = Club(id=2, name="Round Table", telegram_user_id=1)
        self.snapshots: list[EarlyRakebackSnapshot] = []
        self.lines: list[EarlyRakebackLine] = []
        self._snapshot_id = 0
        self._line_id = 0
        self.mock_db = MagicMock()

        def query_model(model):
            q = MagicMock()

            if model is Club:
                q.filter.return_value.params.return_value.first.return_value = self.club
                q.filter.return_value.first.return_value = self.club
                return q

            if model is EarlyRakebackSnapshot:
                def filter_by(**kwargs):
                    inner = MagicMock()
                    matches = [
                        s
                        for s in self.snapshots
                        if all(getattr(s, k) == v for k, v in kwargs.items())
                    ]
                    inner.first.return_value = matches[0] if matches else None
                    return inner

                q.filter_by.side_effect = filter_by
                return q

            if model is EarlyRakebackLine:
                q.filter_by.return_value.delete.return_value = None
                return q

            return q

        self.mock_db.query.side_effect = query_model

        def add(obj):
            if isinstance(obj, EarlyRakebackSnapshot):
                self._snapshot_id += 1
                obj.id = self._snapshot_id
                self.snapshots.append(obj)
            elif isinstance(obj, EarlyRakebackLine):
                self._line_id += 1
                obj.id = self._line_id
                self.lines.append(obj)

        self.mock_db.add.side_effect = add
        self.mock_db.flush = MagicMock()

    @patch.dict(
        os.environ,
        {
            "AON_BETA_BASE_URL": "https://api.example.com/api",
            "AON_BETA_INTERNAL_API_KEY": "test-key",
        },
        clear=False,
    )
    @patch("api.early_rakeback_sync.fetch_early_rakeback_archives", return_value=[])
    @patch("api.early_rakeback_sync.fetch_early_rakeback_entries")
    @patch("api.early_rakeback_sync.audit_day_window_utc")
    def test_sync_flattens_and_skips_unmapped(
        self,
        mock_window,
        mock_fetch,
        _mock_archives,
    ):
        mock_window.return_value = (
            datetime(2026, 6, 19, 4, 0, tzinfo=timezone.utc),
            datetime(2026, 6, 20, 4, 59, 59, 999999, tzinfo=timezone.utc),
        )
        mock_fetch.return_value = _sample_entries()

        report = sync_early_rakeback_for_date(
            self.mock_db,
            date(2026, 6, 19),
            club_slugs=["round-table"],
        )

        self.assertEqual(report.clubs_synced, 1)
        self.assertEqual(report.total_lines_stored, 1)
        self.assertEqual(report.total_lines_skipped_unmapped, 1)
        self.assertEqual(len(self.lines), 1)
        line = self.lines[0]
        self.assertEqual(line.gg_player_id, "3011-9668")
        self.assertEqual(line.amount_usd, Decimal("25.5"))
        self.assertEqual(line.member_nickname, "Alice")
        self.assertEqual(len(self.snapshots), 1)
        self.assertEqual(self.snapshots[0].lines_stored, 1)
        self.assertEqual(self.snapshots[0].lines_skipped_unmapped, 1)

    @patch.dict(
        os.environ,
        {
            "AON_BETA_BASE_URL": "https://api.example.com/api",
            "AON_BETA_INTERNAL_API_KEY": "test-key",
        },
        clear=False,
    )
    @patch("api.early_rakeback_sync.fetch_early_rakeback_archives", return_value=[])
    @patch("api.early_rakeback_sync.fetch_early_rakeback_entries")
    @patch("api.early_rakeback_sync.audit_day_window_utc")
    def test_sync_replaces_existing_snapshot(
        self,
        mock_window,
        mock_fetch,
        _mock_archives,
    ):
        mock_window.return_value = (
            datetime(2026, 6, 19, 4, 0, tzinfo=timezone.utc),
            datetime(2026, 6, 20, 4, 59, 59, 999999, tzinfo=timezone.utc),
        )
        mock_fetch.return_value = _sample_entries()

        sync_early_rakeback_for_date(
            self.mock_db,
            date(2026, 6, 19),
            club_slugs=["round-table"],
        )
        first_snapshot_id = self.snapshots[0].id
        self.lines.clear()

        mock_fetch.return_value = [
            {
                "_id": "entry3",
                "memberNickname": "Bob",
                "memberType": "player",
                "gg_player_id": "3011-9999",
                "records": [
                    {
                        "_id": "rec3",
                        "calculatedAmount": 10,
                        "timestamp": "2026-06-19T14:00:00.000Z",
                    }
                ],
            }
        ]

        report = sync_early_rakeback_for_date(
            self.mock_db,
            date(2026, 6, 19),
            club_slugs=["round-table"],
        )

        self.assertEqual(report.total_lines_stored, 1)
        self.assertEqual(len(self.snapshots), 1)
        self.assertEqual(self.snapshots[0].id, first_snapshot_id)
        self.assertEqual(len(self.lines), 1)
        self.assertEqual(self.lines[0].gg_player_id, "3011-9999")

    def test_flatten_archive_entries_filters_by_window(self):
        from_utc = datetime(2026, 6, 19, 4, 0, tzinfo=timezone.utc)
        to_utc = datetime(2026, 6, 20, 4, 59, 59, 999999, tzinfo=timezone.utc)
        archives = [
            {
                "_id": "arch1",
                "entries": [
                    {
                        "memberNickname": "Carol",
                        "memberType": "player",
                        "gg_player_id": "3011-1111",
                        "records": [
                            {
                                "calculatedAmount": 15,
                                "timestamp": "2026-06-19T18:00:00.000Z",
                            },
                            {
                                "calculatedAmount": 99,
                                "timestamp": "2026-06-18T18:00:00.000Z",
                            },
                        ],
                    }
                ],
            }
        ]
        lines, skipped, nicknames = _flatten_archive_entries(archives, from_utc, to_utc)
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0]["gg_player_id"], "3011-1111")
        self.assertEqual(lines[0]["source_entry_id"], "archive:arch1:0")
        self.assertEqual(lines[0]["amount_usd"], Decimal("15"))
        self.assertEqual(skipped, 0)
        self.assertEqual(nicknames, [])

    @patch.dict(
        os.environ,
        {
            "AON_BETA_BASE_URL": "https://api.example.com/api",
            "AON_BETA_INTERNAL_API_KEY": "test-key",
        },
        clear=False,
    )
    @patch("api.early_rakeback_sync.fetch_early_rakeback_archives")
    @patch("api.early_rakeback_sync.fetch_early_rakeback_entries", return_value=[])
    @patch("api.early_rakeback_sync.audit_day_window_utc")
    def test_sync_merges_archive_lines(
        self,
        mock_window,
        _mock_fetch,
        mock_archives,
    ):
        mock_window.return_value = (
            datetime(2026, 6, 19, 4, 0, tzinfo=timezone.utc),
            datetime(2026, 6, 20, 4, 59, 59, 999999, tzinfo=timezone.utc),
        )
        mock_archives.return_value = [
            {
                "_id": "arch1",
                "entries": [
                    {
                        "memberNickname": "Carol",
                        "memberType": "player",
                        "gg_player_id": "3011-1111",
                        "records": [
                            {
                                "calculatedAmount": 15,
                                "timestamp": "2026-06-19T18:00:00.000Z",
                            }
                        ],
                    }
                ],
            }
        ]

        report = sync_early_rakeback_for_date(
            self.mock_db,
            date(2026, 6, 19),
            club_slugs=["round-table"],
        )

        self.assertEqual(report.total_lines_stored, 1)
        self.assertEqual(self.lines[0].source_entry_id, "archive:arch1:0")

    @patch.dict(
        os.environ,
        {
            "AON_BETA_BASE_URL": "https://api.example.com/api",
            "AON_BETA_INTERNAL_API_KEY": "test-key",
        },
        clear=False,
    )
    @patch("api.early_rakeback_sync.sync_early_rakeback_for_date")
    @patch("api.early_rakeback_sync.fetch_early_rakeback_archives")
    def test_backfill_calls_sync_for_archive_dates(
        self,
        mock_archives,
        mock_sync,
    ):
        mock_archives.return_value = [
            {
                "_id": "arch1",
                "entries": [
                    {
                        "memberNickname": "Carol",
                        "memberType": "player",
                        "gg_player_id": "3011-1111",
                        "records": [
                            {
                                "calculatedAmount": 15,
                                "timestamp": "2026-06-19T18:00:00.000Z",
                            }
                        ],
                    }
                ],
            }
        ]
        mock_sync.return_value = MagicMock()

        reports = backfill_early_rakeback_from_archives(
            self.mock_db,
            club_slugs=["round-table"],
        )

        self.assertEqual(len(reports), 1)
        mock_sync.assert_called_once_with(
            self.mock_db,
            date(2026, 6, 19),
            club_slugs=["round-table"],
        )


class EarlyRakebackExportTestCase(unittest.TestCase):
    @patch("api.audit_export._fetch_stripe_rows", return_value=[])
    @patch("api.audit_export._fetch_tagged_manual_rows", return_value=[])
    @patch("api.audit_export._club_name_map", return_value={})
    def test_export_reads_early_rakeback_from_postgres(
        self,
        _club_map,
        _tagged,
        _stripe,
    ):
        from api.audit_export import build_audit_workbook
        import io
        from openpyxl import load_workbook

        snapshot = EarlyRakebackSnapshot(
            id=1,
            club_id=2,
            club_slug="round-table",
            audit_date=date(2026, 6, 19),
            fetch_from_utc=datetime(2026, 6, 19, 4, 0, tzinfo=timezone.utc),
            fetch_to_utc=datetime(2026, 6, 20, 4, 59, 59, tzinfo=timezone.utc),
            lines_fetched=1,
            lines_stored=1,
            lines_skipped_unmapped=0,
        )
        line = EarlyRakebackLine(
            id=1,
            snapshot_id=1,
            source_entry_id="e1",
            source_record_id="r1",
            gg_player_id="3011-9668",
            member_nickname="Alice",
            member_type="player",
            amount_usd=Decimal("25.50"),
            occurred_at=datetime(2026, 6, 19, 16, 0, tzinfo=timezone.utc),
        )

        session = MagicMock()

        def query_model(model):
            q = MagicMock()
            if model is EarlyRakebackSnapshot:
                q.filter.return_value.all.return_value = [snapshot]
                return q
            if model is EarlyRakebackLine:
                q.filter.return_value.order_by.return_value.all.return_value = [line]
                return q
            return q

        session.query.side_effect = query_model

        content = build_audit_workbook(session, "2026-06-19")
        wb = load_workbook(io.BytesIO(content))
        ws = wb["Early Rakeback"]
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws["C2"].value, "Early RB")
        self.assertEqual(ws["A2"].value, 25.5)
        self.assertIn("3011-9668", str(ws["B2"].value))


if __name__ == "__main__":
    unittest.main()
