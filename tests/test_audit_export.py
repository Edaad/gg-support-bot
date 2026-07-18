"""Tests for cross-club audit XLSX export."""

from __future__ import annotations

import io
import os
import unittest
from datetime import datetime, timezone
from decimal import Decimal
from unittest.mock import MagicMock, patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.auth import create_token, get_current_admin
from api.audit_export import (
    SHEET_SPECS,
    ManualAuditRow,
    StripeAuditRow,
    TaggedManualAuditRow,
    _bonus_group_cell,
    _bonus_payer_display,
    _fmt_manual_audit_time,
    _fmt_stripe_audit_time,
    _manual_club_name,
    _manual_group_cell,
    _manual_row,
    _stripe_player_cell,
    _tagged_manual_row,
    audit_day_window_utc,
    build_audit_workbook,
    eastern_audit_end_utc,
    eastern_day_bounds_utc,
)
from api.payments_helpers import build_crypto_payment_read
from api.routes.payments import router
from bot.services.crypto_payments import ALERT_SCOPE_LABELS, ALERT_SCOPE_CLUBGTO
from db.connection import get_db_dependency
from db.models import BonusRecord, BonusType, CryptoPayment

TOKEN = create_token()


def _make_app() -> FastAPI:
    app = FastAPI()
    app.include_router(router)

    def override_admin():
        return "admin"

    def override_db():
        yield MagicMock()

    app.dependency_overrides[get_current_admin] = override_admin
    app.dependency_overrides[get_db_dependency] = override_db
    return app


class AuditExportFormattingTestCase(unittest.TestCase):
    def test_audit_day_window_utc_edt(self):
        start, end = audit_day_window_utc("2026-06-19")
        self.assertEqual(start, datetime(2026, 6, 19, 4, 0, tzinfo=timezone.utc))
        self.assertEqual(end, datetime(2026, 6, 20, 4, 59, 59, 999999, tzinfo=timezone.utc))

    def test_eastern_day_bounds_utc_edt(self):
        start, end = eastern_day_bounds_utc("2026-06-19")
        self.assertEqual(start, datetime(2026, 6, 19, 4, 0, tzinfo=timezone.utc))
        self.assertEqual(end, datetime(2026, 6, 20, 3, 59, 59, 999999, tzinfo=timezone.utc))

    def test_eastern_day_bounds_utc_est(self):
        start, end = eastern_day_bounds_utc("2026-01-15")
        self.assertEqual(start, datetime(2026, 1, 15, 4, 0, tzinfo=timezone.utc))
        self.assertEqual(end, datetime(2026, 1, 16, 3, 59, 59, 999999, tzinfo=timezone.utc))

    def test_eastern_day_bounds_utc_accepts_iso_prefix(self):
        start, end = eastern_day_bounds_utc("2026-06-19T00:00:00Z")
        self.assertEqual(start, datetime(2026, 6, 19, 4, 0, tzinfo=timezone.utc))
        self.assertEqual(end, datetime(2026, 6, 20, 3, 59, 59, 999999, tzinfo=timezone.utc))

    def test_eastern_audit_end_utc_edt(self):
        end = eastern_audit_end_utc("2026-06-21")
        self.assertEqual(end, datetime(2026, 6, 22, 4, 59, 59, 999999, tzinfo=timezone.utc))

    def test_eastern_audit_end_utc_est(self):
        end = eastern_audit_end_utc("2026-01-15")
        self.assertEqual(end, datetime(2026, 1, 16, 4, 59, 59, 999999, tzinfo=timezone.utc))

    def test_fmt_stripe_audit_time_uses_ordinal_eastern(self):
        dt = datetime(2026, 6, 19, 4, 58, tzinfo=timezone.utc)
        self.assertEqual(_fmt_stripe_audit_time(dt), "Jun 19th 2026, 12:58 AM")

    def test_fmt_manual_audit_time_uses_full_month_and_at(self):
        dt = datetime(2026, 6, 19, 4, 34, tzinfo=timezone.utc)
        self.assertEqual(_fmt_manual_audit_time(dt), "June 19, 2026 at 12:34 AM")

    def test_fmt_manual_audit_time_clubgto_shows_new_york_time(self):
        """Display is always America/New_York; club policy affects bucketing only."""
        dt = datetime(2026, 7, 4, 2, 30, tzinfo=timezone.utc)
        self.assertEqual(
            _fmt_manual_audit_time(dt, club_slug="clubgto"),
            "July 3, 2026 at 10:30 PM",
        )
        self.assertEqual(
            _fmt_manual_audit_time(dt, club_slug="round-table"),
            "July 3, 2026 at 10:30 PM",
        )

    def test_fmt_manual_audit_time_uses_new_york_dst_in_winter(self):
        """Payment display follows America/New_York (EST in January)."""
        dt = datetime(2026, 1, 15, 17, 0, tzinfo=timezone.utc)
        self.assertEqual(
            _fmt_manual_audit_time(dt, club_slug="round-table"),
            "January 15, 2026 at 12:00 PM",
        )

    def test_stripe_player_cell_uses_group_title_when_present(self):
        self.assertEqual(
            _stripe_player_cell(
                group_title="GTO / 3011-9668 / Pvtenis",
                club_name="ClubGTO",
                gg_player_id="3011-9668",
                gg_nickname="Pvtenis",
            ),
            "GTO / 3011-9668 / Pvtenis",
        )

    def test_stripe_player_cell_builds_fallback(self):
        self.assertEqual(
            _stripe_player_cell(
                group_title=None,
                club_name="ClubGTO",
                gg_player_id="3011-9668",
                gg_nickname="Pvtenis",
            ),
            "GTO / 3011-9668 / Pvtenis",
        )

    def test_manual_group_cell_returns_bound_group_title(self):
        self.assertEqual(
            _manual_group_cell({"group_title": "GTO / 3011-9668 / Pvtenis"}),
            "GTO / 3011-9668 / Pvtenis",
        )

    def test_bonus_group_cell_uses_type_name(self):
        record = BonusRecord(
            player_username="@player",
            amount=Decimal("50.00"),
            bonus_type=BonusType(name="Referral"),
        )
        self.assertEqual(_bonus_group_cell(record), "Referral")

    def test_bonus_payer_display_prefers_group_title(self):
        record = BonusRecord(
            player_username="legacy",
            amount=Decimal("50.00"),
            group_title="GTO / 3011-9668 / Pvtenis",
        )
        self.assertEqual(
            _bonus_payer_display(record),
            "GTO / 3011-9668 / Pvtenis",
        )

    def test_bonus_payer_display_falls_back_to_player_username(self):
        record = BonusRecord(
            player_username="CC / 1111-2222 / Old",
            amount=Decimal("50.00"),
        )
        self.assertEqual(_bonus_payer_display(record), "CC / 1111-2222 / Old")

    def test_bonus_group_cell_uses_custom_description_for_other(self):
        record = BonusRecord(
            player_username="@player",
            amount=Decimal("25.00"),
            custom_description="Birthday promo",
        )
        self.assertEqual(_bonus_group_cell(record), "Birthday promo")

    def test_bonus_group_cell_combines_type_and_description(self):
        record = BonusRecord(
            player_username="@player",
            amount=Decimal("10.00"),
            bonus_type=BonusType(name="VIP"),
            custom_description="Extra comp",
        )
        self.assertEqual(_bonus_group_cell(record), "VIP — Extra comp")

    def test_manual_group_cell_empty_when_unbound(self):
        self.assertEqual(_manual_group_cell({}), "")
        self.assertEqual(_manual_group_cell({"group_title": None}), "")

    def test_manual_club_name_from_club_id(self):
        club_names = {1: "ClubGTO"}
        self.assertEqual(
            _manual_club_name(
                {
                    "club_id": 1,
                    "group_title": "GTO / 3011-9668 / Pvtenis",
                },
                club_names,
            ),
            "ClubGTO",
        )

    def test_manual_club_name_falls_back_to_title_parsing(self):
        self.assertEqual(
            _manual_club_name(
                {"group_title": "GTO / 8190-5287 / ThePirate343"},
                {},
            ),
            "ClubGTO",
        )

    def test_manual_club_name_empty_when_unbound(self):
        self.assertEqual(
            _manual_club_name({"zelle_recipient": "clubgto1234@gmail.com"}, {}),
            "",
        )

    def test_manual_row_includes_group_and_club(self):
        created = datetime(2026, 6, 22, 1, 51, tzinfo=timezone.utc)
        row = _manual_row(
            MagicMock(),
            {
                "amount_usd": Decimal("100.00"),
                "payer_name": "Jackson Taylor",
                "group_title": "RT / 6485-8168 / Angus Mcgoon",
                "club_id": 2,
                "created_at": created,
            },
            {2: "Round Table"},
        )
        self.assertEqual(row.amount_usd, 100.0)
        self.assertEqual(row.payer_name, "Jackson Taylor")
        self.assertEqual(row.group_title, "RT / 6485-8168 / Angus Mcgoon")
        self.assertEqual(row.club_label, "Round Table")
        self.assertEqual(row.time_label, _fmt_manual_audit_time(created))

    def test_tagged_manual_row_includes_account_tag(self):
        created = datetime(2026, 6, 22, 1, 51, tzinfo=timezone.utc)
        row = _tagged_manual_row(
            MagicMock(),
            {
                "amount_usd": Decimal("50.00"),
                "payer_name": "Jane Doe",
                "venmo_handle": "@godfather4444",
                "group_title": "GTO / 3011-9668 / Pvtenis",
                "club_id": 1,
                "created_at": created,
            },
            {1: "ClubGTO"},
            tag_field="venmo_handle",
        )
        self.assertEqual(row.account_tag, "@godfather4444")
        self.assertEqual(row.payer_name, "Jane Doe")
        self.assertEqual(row.club_label, "ClubGTO")

    def test_tagged_manual_row_reads_zelle_recipient(self):
        created = datetime(2026, 6, 22, 1, 51, tzinfo=timezone.utc)
        row = _tagged_manual_row(
            MagicMock(),
            {
                "amount_usd": Decimal("100.00"),
                "payer_name": "MR ROHIT KOTHLAPURAM",
                "zelle_recipient": "clubgto1234@gmail.com",
                "group_title": "GTO / 3011-9668 / Pvtenis",
                "club_id": 1,
                "created_at": created,
            },
            {1: "ClubGTO"},
            tag_field="zelle_recipient",
        )
        self.assertEqual(row.account_tag, "clubgto1234@gmail.com")

    def test_tagged_manual_row_uses_crypto_from_label_and_token(self):
        created = datetime(2026, 6, 21, 3, 57, tzinfo=timezone.utc)
        row = _tagged_manual_row(
            MagicMock(),
            {
                "amount_usd": Decimal("122.00"),
                "from_label": "Binance (0x8894…D4E3)",
                "token_symbol": "USDC",
                "group_title": "GTO / 3011-9668 / Pvtenis",
                "club_id": 1,
                "created_at": created,
            },
            {1: "ClubGTO"},
            tag_field="token_symbol",
        )
        self.assertEqual(row.payer_name, "Binance (0x8894…D4E3)")
        self.assertEqual(row.account_tag, "USDC")
        self.assertEqual(row.group_title, "GTO / 3011-9668 / Pvtenis")
        self.assertEqual(row.club_label, "ClubGTO")


class AuditExportWorkbookTestCase(unittest.TestCase):
    def test_alert_scope_labels_importable(self):
        self.assertIn(ALERT_SCOPE_CLUBGTO, ALERT_SCOPE_LABELS)

    def test_sheet_specs_use_title_case_tab_names(self):
        self.assertEqual(
            [spec.title for spec in SHEET_SPECS],
            [
                "Stripe",
                "Zelle",
                "Venmo",
                "Cash App",
                "PayPal",
                "Crypto",
                "Bonus",
                "Early Rakeback",
            ],
        )

    def test_build_crypto_payment_read_does_not_raise(self):
        payment = CryptoPayment(
            id=1,
            amount_cents=10000,
            token_symbol="USDC",
            chain="ethereum",
            from_address="0xfrom1234567890abcdef",
            to_address="0xto",
            transaction_hash="0xhash",
            alert_scope=ALERT_SCOPE_CLUBGTO,
            is_test=False,
            auto_bound=False,
            created_at=datetime(2026, 6, 17, tzinfo=timezone.utc),
        )
        data = build_crypto_payment_read(MagicMock(), payment)
        self.assertEqual(data["alert_scope_label"], ALERT_SCOPE_LABELS[ALERT_SCOPE_CLUBGTO])
        self.assertIn("0xfrom", data["from_label"])

    @patch("api.audit_export._fetch_early_rakeback_rows", return_value=[])
    @patch("api.audit_export._fetch_stripe_rows", return_value=[])
    @patch("api.audit_export._fetch_tagged_manual_rows", return_value=[])
    @patch("api.audit_export._fetch_bonus_rows", return_value=[])
    @patch("api.audit_export._club_name_map", return_value={})
    def test_build_audit_workbook_has_eight_sheets_with_headers(
        self,
        _club_map,
        _bonus,
        _tagged,
        _stripe,
        _early_rb,
    ):
        session = MagicMock()
        content = build_audit_workbook(session, "2026-01-31")
        wb = load_workbook(io.BytesIO(content))
        self.assertEqual(wb.sheetnames, [spec.title for spec in SHEET_SPECS])
        for spec in SHEET_SPECS:
            ws = wb[spec.title]
            self.assertEqual([cell.value for cell in ws[1]], spec.headers)

    @patch("api.audit_export._fetch_early_rakeback_rows", return_value=[])
    @patch("api.audit_export._fetch_tagged_manual_rows", return_value=[])
    @patch("api.audit_export._fetch_bonus_rows", return_value=[])
    @patch("api.audit_export._club_name_map", return_value={})
    def test_build_audit_workbook_styles_stripe_sheet(self, _club_map, _bonus, _tagged, _early_rb):
        stripe_rows = [
            StripeAuditRow(
                amount_usd=42.0,
                player="GTO / 3011-9668 / Pvtenis",
                method_label="Manual (/stripe)",
                group_title="GTO / 3011-9668 / Pvtenis",
                club_label="ClubGTO",
                time_label="Jun 19th 2026, 12:58 AM",
                stripe_fee_usd=Decimal("1.52"),
            )
        ]

        with patch("api.audit_export._fetch_stripe_rows", return_value=stripe_rows):
            content = build_audit_workbook(MagicMock(), "2026-01-31")

        wb = load_workbook(io.BytesIO(content))
        ws = wb["Stripe"]
        header = ws["A1"]
        self.assertEqual(header.value, "Amount")
        self.assertEqual(header.fill.start_color.rgb, "0038761D")
        self.assertTrue(header.font.bold)
        self.assertEqual(header.font.color.rgb, "00FFFFFF")

        amount_cell = ws["A2"]
        self.assertEqual(amount_cell.value, 42.0)
        self.assertEqual(amount_cell.number_format, "$#,##0.00")
        self.assertIsNotNone(amount_cell.comment)
        self.assertIn("Stripe fee", amount_cell.comment.text)
        self.assertEqual(
            [cell.value for cell in ws[1]],
            ["Amount", "Player", "Method", "Group", "Club", "Time"],
        )
        self.assertEqual(ws["B2"].value, "GTO / 3011-9668 / Pvtenis")
        self.assertEqual(ws["C2"].value, "Manual (/stripe)")
        self.assertEqual(ws["D2"].value, "GTO / 3011-9668 / Pvtenis")
        self.assertEqual(ws["E2"].value, "ClubGTO")

        zelle_ws = wb["Zelle"]
        self.assertEqual(zelle_ws.max_row, 1)

    @patch("api.audit_export._fetch_early_rakeback_rows", return_value=[])
    @patch("api.audit_export._fetch_stripe_rows", return_value=[])
    @patch("api.audit_export._fetch_tagged_manual_rows", return_value=[])
    @patch("api.audit_export._club_name_map", return_value={})
    def test_build_audit_workbook_writes_tagged_manual_columns(
        self,
        _club_map,
        _tagged,
        _stripe,
        _early_rb,
    ):
        tagged_rows = [
            TaggedManualAuditRow(
                amount_usd=199.99,
                payer_name="MR ROHIT KOTHLAPURAM",
                account_tag="clubgto1234@gmail.com",
                group_title="GTO / 3011-9668 / Pvtenis",
                club_label="ClubGTO",
                time_label="June 21, 2026 at 11:57 PM",
            )
        ]

        def tagged_side_effect(
            session, payment_cls, build_read, club_names, from_dt, to_dt, *, audit_date, tag_field
        ):
            if payment_cls.__name__ == "ZellePayment":
                return tagged_rows
            return []

        with patch(
            "api.audit_export._fetch_tagged_manual_rows",
            side_effect=tagged_side_effect,
        ):
            content = build_audit_workbook(
                MagicMock(),
                "2026-06-21",
            )

        wb = load_workbook(io.BytesIO(content))
        zelle_ws = wb["Zelle"]
        self.assertEqual(
            [cell.value for cell in zelle_ws[1]],
            ["Amount", "Name", "Tag", "Group", "Club", "Time"],
        )
        self.assertEqual(zelle_ws["A2"].value, 199.99)
        self.assertEqual(zelle_ws["B2"].value, "MR ROHIT KOTHLAPURAM")
        self.assertEqual(zelle_ws["C2"].value, "clubgto1234@gmail.com")
        self.assertEqual(zelle_ws["D2"].value, "GTO / 3011-9668 / Pvtenis")
        self.assertEqual(zelle_ws["E2"].value, "ClubGTO")
        self.assertEqual(zelle_ws["F2"].value, "June 21, 2026 at 11:57 PM")

    @patch("api.audit_export._fetch_stripe_rows", return_value=[])
    @patch("api.audit_export._fetch_bonus_rows", return_value=[])
    @patch("api.audit_export._club_name_map", return_value={})
    def test_build_audit_workbook_writes_venmo_tag_column(self, _club_map, _bonus, _stripe):
        venmo_rows = [
            TaggedManualAuditRow(
                amount_usd=100.0,
                payer_name="Jane Doe",
                account_tag="@godfather4444",
                group_title="GTO / 3011-9668 / Pvtenis",
                club_label="ClubGTO",
                time_label="June 21, 2026 at 11:57 PM",
            )
        ]

        def tagged_side_effect(
            session, payment_cls, build_read, club_names, from_dt, to_dt, *, audit_date, tag_field
        ):
            if payment_cls.__name__ == "VenmoPayment":
                return venmo_rows
            return []

        with patch(
            "api.audit_export._fetch_tagged_manual_rows",
            side_effect=tagged_side_effect,
        ):
            content = build_audit_workbook(
                MagicMock(),
                "2026-06-21",
            )

        wb = load_workbook(io.BytesIO(content))
        venmo_ws = wb["Venmo"]
        self.assertEqual(
            [cell.value for cell in venmo_ws[1]],
            ["Amount", "Name", "Tag", "Group", "Club", "Time"],
        )
        self.assertEqual(venmo_ws["C2"].value, "@godfather4444")
        self.assertEqual(venmo_ws["D2"].value, "GTO / 3011-9668 / Pvtenis")
        self.assertEqual(venmo_ws["E2"].value, "ClubGTO")

    @patch("api.audit_export._fetch_early_rakeback_rows", return_value=[])
    @patch("api.audit_export._fetch_stripe_rows", return_value=[])
    @patch("api.audit_export._fetch_bonus_rows", return_value=[])
    @patch("api.audit_export._club_name_map", return_value={})
    def test_build_audit_workbook_writes_crypto_sheet(self, _club_map, _bonus, _stripe, _early_rb):
        crypto_rows = [
            TaggedManualAuditRow(
                amount_usd=122.0,
                payer_name="Binance (0x8894…D4E3)",
                account_tag="USDC",
                group_title="GTO / 3011-9668 / Pvtenis",
                club_label="ClubGTO",
                time_label="June 21, 2026 at 11:57 PM",
            )
        ]

        def tagged_side_effect(
            session, payment_cls, build_read, club_names, from_dt, to_dt, *, audit_date, tag_field
        ):
            if payment_cls.__name__ == "CryptoPayment":
                return crypto_rows
            return []

        with patch(
            "api.audit_export._fetch_tagged_manual_rows",
            side_effect=tagged_side_effect,
        ):
            content = build_audit_workbook(
                MagicMock(),
                "2026-06-21",
            )

        wb = load_workbook(io.BytesIO(content))
        crypto_ws = wb["Crypto"]
        self.assertEqual(
            [cell.value for cell in crypto_ws[1]],
            ["Amount", "Name", "Tag", "Group", "Club", "Time"],
        )
        self.assertEqual(crypto_ws["A2"].value, 122.0)
        self.assertEqual(crypto_ws["B2"].value, "Binance (0x8894…D4E3)")
        self.assertEqual(crypto_ws["C2"].value, "USDC")
        self.assertEqual(crypto_ws["D2"].value, "GTO / 3011-9668 / Pvtenis")
        self.assertEqual(crypto_ws["E2"].value, "ClubGTO")

    @patch("api.audit_export._fetch_early_rakeback_rows", return_value=[])
    @patch("api.audit_export._fetch_stripe_rows", return_value=[])
    @patch("api.audit_export._fetch_tagged_manual_rows", return_value=[])
    @patch("api.audit_export._club_name_map", return_value={})
    def test_build_audit_workbook_writes_bonus_sheet(self, _club_map, _tagged, _stripe, _early_rb):
        bonus_rows = [
            ManualAuditRow(
                amount_usd=75.0,
                payer_name="@luckyplayer",
                group_title="Referral",
                club_label="ClubGTO",
                time_label="June 21, 2026 at 11:57 PM",
            )
        ]

        with patch("api.audit_export._fetch_bonus_rows", return_value=bonus_rows):
            content = build_audit_workbook(MagicMock(), "2026-06-21")

        wb = load_workbook(io.BytesIO(content))
        bonus_ws = wb["Bonus"]
        self.assertEqual(
            [cell.value for cell in bonus_ws[1]],
            ["Amount", "Name", "Group", "Club", "Time"],
        )
        self.assertEqual(bonus_ws["A2"].value, 75.0)
        self.assertEqual(bonus_ws["B2"].value, "@luckyplayer")
        self.assertEqual(bonus_ws["C2"].value, "Referral")
        self.assertEqual(bonus_ws["D2"].value, "ClubGTO")
        self.assertEqual(bonus_ws["E2"].value, "June 21, 2026 at 11:57 PM")


class AuditExportApiTestCase(unittest.TestCase):
    def setUp(self):
        self.env_patch = patch.dict(os.environ, {"DASHBOARD_PASSWORD": "changeme"}, clear=False)
        self.env_patch.start()
        self.client = TestClient(_make_app())

    def tearDown(self):
        self.env_patch.stop()

    def test_audit_export_requires_auth(self):
        app = FastAPI()
        app.include_router(router)
        client = TestClient(app)
        response = client.get(
            "/api/payments/audit-export",
            params={"date": "2026-01-31"},
        )
        self.assertIn(response.status_code, (401, 403))

    def test_audit_export_requires_date(self):
        response = self.client.get(
            "/api/payments/audit-export",
            headers={"Authorization": f"Bearer {TOKEN}"},
        )
        self.assertEqual(response.status_code, 422)

    @patch("api.routes.payments.build_audit_workbook")
    def test_audit_export_success(self, mock_build):
        mock_build.return_value = b"fake-xlsx"
        response = self.client.get(
            "/api/payments/audit-export",
            params={"date": "2026-01-31"},
            headers={"Authorization": f"Bearer {TOKEN}"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("audit-export-2026-01-31.xlsx", response.headers["content-disposition"])
        self.assertEqual(response.content, b"fake-xlsx")
        mock_build.assert_called_once()

    @patch("api.routes.payments.build_audit_workbook")
    def test_audit_export_uses_audit_date(self, mock_build):
        mock_build.return_value = b"fake-xlsx"
        response = self.client.get(
            "/api/payments/audit-export",
            params={"date": "2026-06-19"},
            headers={"Authorization": f"Bearer {TOKEN}"},
        )
        self.assertEqual(response.status_code, 200)
        mock_build.assert_called_once()
        _, audit_date = mock_build.call_args[0]
        self.assertEqual(audit_date, "2026-06-19")


if __name__ == "__main__":
    unittest.main()
