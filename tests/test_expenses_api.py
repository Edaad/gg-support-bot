"""Admin expenses API: CRUD, filters, role gate, XLSX export."""

from __future__ import annotations

import os
import unittest
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from api.auth import (
    ROLE_ACCOUNT_MANAGER,
    ROLE_ADMIN,
    create_token,
    require_admin,
)
from api.routes.expenses import router
from db.connection import get_db_dependency
from db.models import Base, Club, Expense


def _make_app(session_factory):
    app = FastAPI()
    app.include_router(router)

    def override_db():
        session = session_factory()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    app.dependency_overrides[get_db_dependency] = override_db
    return app


class ExpensesApiTestCase(unittest.TestCase):
    def setUp(self) -> None:
        import api.auth as auth_mod

        auth_mod._SECRET = None
        self.env_patch = patch.dict(
            os.environ,
            {"DASHBOARD_PASSWORD": "admin-secret", "DASHBOARD_AM_PASSWORD": "am-secret"},
            clear=False,
        )
        self.env_patch.start()
        auth_mod._SECRET = None

        self.engine = create_engine(
            "sqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine, tables=[Club.__table__, Expense.__table__])
        self.Session = sessionmaker(bind=self.engine)
        session = self.Session()
        session.add(Club(id=1, name="Round Table", telegram_user_id=1001, is_active=True))
        session.add(Club(id=2, name="ClubGTO", telegram_user_id=1002, is_active=True))
        session.commit()
        session.close()

        self.app = _make_app(self.Session)
        self.client = TestClient(self.app)
        self.admin_headers = {"Authorization": f"Bearer {create_token(ROLE_ADMIN)}"}
        self.am_headers = {"Authorization": f"Bearer {create_token(ROLE_ACCOUNT_MANAGER)}"}

    def tearDown(self) -> None:
        import api.auth as auth_mod

        auth_mod._SECRET = None
        self.env_patch.stop()
        self.engine.dispose()

    def test_am_forbidden(self) -> None:
        r = self.client.get("/api/expenses", headers=self.am_headers)
        self.assertEqual(r.status_code, 403)

    def test_crud_and_filters(self) -> None:
        create = self.client.post(
            "/api/expenses",
            headers=self.admin_headers,
            json={
                "amount": "12.50",
                "expense_type": "Software",
                "description": "Notion",
                "club_id": 1,
                "expense_date": "2026-08-01",
                "pending": True,
            },
        )
        self.assertEqual(create.status_code, 201, create.text)
        body = create.json()
        self.assertEqual(body["club_name"], "Round Table")
        self.assertEqual(Decimal(str(body["amount"])), Decimal("12.50"))
        self.assertTrue(body["pending"])
        expense_id = body["id"]

        self.client.post(
            "/api/expenses",
            headers=self.admin_headers,
            json={
                "amount": "5.00",
                "expense_type": "Ads",
                "description": "FB",
                "club_id": 2,
                "expense_date": "2026-08-10",
                "pending": False,
            },
        )

        listed = self.client.get(
            "/api/expenses",
            headers=self.admin_headers,
            params={"club_id": 1, "pending": "true", "from": "2026-08-01", "to": "2026-08-31"},
        )
        self.assertEqual(listed.status_code, 200)
        rows = listed.json()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["expense_type"], "Software")

        searched = self.client.get(
            "/api/expenses",
            headers=self.admin_headers,
            params={"q": "notion", "from": "2026-08-01", "to": "2026-08-31"},
        )
        self.assertEqual(len(searched.json()), 1)

        patched = self.client.patch(
            f"/api/expenses/{expense_id}",
            headers=self.admin_headers,
            json={"pending": False, "amount": "13.00"},
        )
        self.assertEqual(patched.status_code, 200)
        self.assertFalse(patched.json()["pending"])
        self.assertEqual(Decimal(str(patched.json()["amount"])), Decimal("13.00"))

        deleted = self.client.delete(
            f"/api/expenses/{expense_id}",
            headers=self.admin_headers,
        )
        self.assertEqual(deleted.status_code, 204)
        after = self.client.get(
            "/api/expenses",
            headers=self.admin_headers,
            params={"from": "2026-08-01", "to": "2026-08-31"},
        )
        self.assertEqual(len(after.json()), 1)

    def test_amount_must_be_positive(self) -> None:
        r = self.client.post(
            "/api/expenses",
            headers=self.admin_headers,
            json={
                "amount": "0",
                "expense_type": "X",
                "club_id": 1,
                "expense_date": "2026-08-01",
            },
        )
        self.assertEqual(r.status_code, 400)

    def test_export_xlsx_respects_filters(self) -> None:
        self.client.post(
            "/api/expenses",
            headers=self.admin_headers,
            json={
                "amount": "10",
                "expense_type": "Keep",
                "club_id": 1,
                "expense_date": "2026-08-05",
                "pending": True,
            },
        )
        self.client.post(
            "/api/expenses",
            headers=self.admin_headers,
            json={
                "amount": "20",
                "expense_type": "Skip",
                "club_id": 2,
                "expense_date": "2026-08-05",
                "pending": True,
            },
        )
        r = self.client.get(
            "/api/expenses/export",
            headers=self.admin_headers,
            params={"club_id": 1, "from": "2026-08-01", "to": "2026-08-31"},
        )
        self.assertEqual(r.status_code, 200)
        self.assertIn(
            "spreadsheetml",
            r.headers.get("content-type", ""),
        )
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        # header + 1 data row
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws.cell(2, 2).value, "Keep")

    def test_require_admin_dependency(self) -> None:
        # Direct dependency behavior via JWT roles already covered; sanity-check helpers
        self.assertEqual(require_admin(ROLE_ADMIN), ROLE_ADMIN)
        with self.assertRaises(Exception):
            # FastAPI wraps HTTPException; call dependency function directly
            from fastapi import HTTPException

            try:
                require_admin(ROLE_ACCOUNT_MANAGER)
            except HTTPException as e:
                self.assertEqual(e.status_code, 403)
                raise


if __name__ == "__main__":
    unittest.main()
