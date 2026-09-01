"""Creator Club weekly audit workbook from 7 days of all-clubs Matching exports."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from api.audit_ledger import _apply_audit_manual_filters, payment_in_audit_day_for_club
from api.audit_reconcile_export import MATCHING_HEADERS
from api.club_audit_timezone import partner_audit_day_window_utc, zone_for_slug
from api.club_slug import resolve_club_id
from api.method_owner import METHOD_OWNER_MATEOS
from api.payments_helpers import (
    build_crypto_payment_read,
    build_venmo_payment_read,
    build_zelle_payment_read,
)
from api.vaughn_methods import normalize_venmo_handle
from bot.services.payment_method_binding import canonicalize_zelle_recipient
from db.models import BonusRecord, CryptoPayment, VenmoPayment, ZellePayment

FILENAME_RE = re.compile(
    r"^reconcile-all-clubs-(\d{4}-\d{2}-\d{2})\.xlsx$",
    re.IGNORECASE,
)

CREATOR_CLUB_SHEET = "Creator Club"
_CREATOR_CLUB_SLUG = "creator-club"
TEMPLATE_PATH = (
    Path(__file__).resolve().parent / "templates" / "creator_weekly_audit_base.xlsx"
)

PROCESSED_HEADERS = [
    "Date / Time",
    "Admin Username",
    "Amount",
    "Player ID",
    "Player Username",
    "Category",
    "Source date",
]

ZELLE_VENMO_HEADERS = ["Time", "Name", "Amount", "Variant", "Source date"]
CRYPTO_HEADERS = ["Time", "From", "USD", "Token", "Source date"]
BONUSES_HEADERS = ["Time", "Player", "Amount", "Source date"]

MISSING_DATA = "Missing data"

_HEADER_FILL = PatternFill("solid", fgColor="38761D")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_CURRENCY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'
_PROCESSED_TABLE = "ProcessedData"
_PROCESSED_COL_COUNT = len(PROCESSED_HEADERS)


class CreatorWeeklyAuditError(ValueError):
    """User-facing validation / parse error for Creator Club weekly audit export."""


@dataclass(frozen=True)
class MatchingRow:
    trade_time: datetime | None
    manager: str
    amount: Decimal | float | int | None
    player_id: str
    nickname: str
    source: str
    name: str
    match_time: datetime | None
    match_amount: Decimal | float | int | None
    variant: str
    audit_date: date


@dataclass(frozen=True)
class PaymentRailRow:
    audit_date: date
    occurred_at: datetime | None
    name: str
    variant: str
    amount_usd: float


@dataclass(frozen=True)
class BonusRailRow:
    audit_date: date
    occurred_at: datetime | None
    player: str
    amount_usd: float


def _display_cell(value: object) -> object:
    """Blank → Missing data; keep datetimes/numbers/non-empty strings."""
    if value is None:
        return MISSING_DATA
    if isinstance(value, str) and not value.strip():
        return MISSING_DATA
    return value


def expected_week_dates(monday: date) -> list[date]:
    if monday.weekday() != 0:
        raise CreatorWeeklyAuditError(
            f"Week start must be a Monday; got {monday.isoformat()} "
            f"({monday.strftime('%A')})."
        )
    return [monday + timedelta(days=i) for i in range(7)]


def parse_monday(value: str) -> date:
    text = (value or "").strip()
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise CreatorWeeklyAuditError(
            f"monday must be YYYY-MM-DD; got {value!r}."
        ) from exc


def date_from_filename(filename: str) -> date:
    base = Path(filename).name
    match = FILENAME_RE.match(base)
    if not match:
        raise CreatorWeeklyAuditError(
            f"Filename must be reconcile-all-clubs-YYYY-MM-DD.xlsx; got {base!r}."
        )
    try:
        return date.fromisoformat(match.group(1))
    except ValueError as exc:
        raise CreatorWeeklyAuditError(
            f"Filename date is not valid YYYY-MM-DD: {base!r}."
        ) from exc


def validate_upload_set(monday: date, filenames: list[str]) -> list[date]:
    expected = expected_week_dates(monday)
    if len(filenames) != 7:
        raise CreatorWeeklyAuditError(
            f"Expected exactly 7 files for {expected[0].isoformat()}–"
            f"{expected[-1].isoformat()}; got {len(filenames)}."
        )

    dates: list[date] = []
    seen: set[date] = set()
    for name in filenames:
        d = date_from_filename(name)
        if d in seen:
            raise CreatorWeeklyAuditError(
                f"Duplicate file for {d.isoformat()}: {Path(name).name!r}."
            )
        seen.add(d)
        dates.append(d)

    expected_set = set(expected)
    if seen != expected_set:
        missing = sorted(expected_set - seen)
        extra = sorted(seen - expected_set)
        parts: list[str] = [
            f"Files must cover Mon–Sun {expected[0].isoformat()} through "
            f"{expected[-1].isoformat()}."
        ]
        if missing:
            parts.append("Missing: " + ", ".join(d.isoformat() for d in missing) + ".")
        if extra:
            parts.append("Unexpected: " + ", ".join(d.isoformat() for d in extra) + ".")
        raise CreatorWeeklyAuditError(" ".join(parts))

    return expected


def output_filename(monday: date) -> str:
    sunday = monday + timedelta(days=6)
    mon_part = f"{monday.strftime('%b')}{monday.day}"
    sun_part = str(sunday.day)
    return f"Creator Audit {mon_part}_{sun_part}-{monday.year}.xlsx"


def _cell_datetime(value: object) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    return None


def _cell_number(value: object) -> Decimal | float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        text = value.strip().replace("$", "").replace(",", "")
        if not text:
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None
    return None


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_map(ws: Worksheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(1, col).value
        if raw is None:
            continue
        name = str(raw).strip()
        if name:
            mapping[name] = col
    return mapping


def _row_is_empty(ws: Worksheet, row: int, cols: list[int]) -> bool:
    for col in cols:
        value = ws.cell(row, col).value
        if value is not None and str(value).strip() != "":
            return False
    return True


def parse_creator_club_rows(
    workbook: Workbook,
    *,
    filename: str = "",
    audit_date: date | None = None,
) -> list[MatchingRow]:
    label = Path(filename).name if filename else "workbook"
    day = audit_date
    if day is None:
        if not filename:
            raise CreatorWeeklyAuditError("audit_date is required when filename is empty.")
        day = date_from_filename(filename)
    if CREATOR_CLUB_SHEET not in workbook.sheetnames:
        raise CreatorWeeklyAuditError(
            f"{label}: missing sheet {CREATOR_CLUB_SHEET!r}."
        )
    ws = workbook[CREATOR_CLUB_SHEET]
    headers = _header_map(ws)
    missing = [h for h in MATCHING_HEADERS if h not in headers]
    if missing:
        raise CreatorWeeklyAuditError(
            f"{label}: Creator Club sheet missing required headers: "
            + ", ".join(missing)
            + "."
        )

    cols = [headers[h] for h in MATCHING_HEADERS]
    rows: list[MatchingRow] = []
    for row_idx in range(2, ws.max_row + 1):
        if _row_is_empty(ws, row_idx, cols):
            continue
        rows.append(
            MatchingRow(
                trade_time=_cell_datetime(ws.cell(row_idx, headers["Trade Time"]).value),
                manager=_cell_str(ws.cell(row_idx, headers["Manager"]).value),
                amount=_cell_number(ws.cell(row_idx, headers["Amount"]).value),
                player_id=_cell_str(ws.cell(row_idx, headers["Player ID"]).value),
                nickname=_cell_str(ws.cell(row_idx, headers["Nickname"]).value),
                source=_cell_str(ws.cell(row_idx, headers["Source"]).value),
                name=_cell_str(ws.cell(row_idx, headers["Name"]).value),
                match_time=_cell_datetime(ws.cell(row_idx, headers["Match Time"]).value),
                match_amount=_cell_number(ws.cell(row_idx, headers["$"]).value),
                variant=_cell_str(ws.cell(row_idx, headers["Variant"]).value),
                audit_date=day,
            )
        )

    if not rows:
        raise CreatorWeeklyAuditError(
            f"{label}: Creator Club sheet has no data rows "
            f"(need at least one non-empty Matching row)."
        )
    return rows


def _sort_key_time(value: datetime | None) -> tuple[int, datetime]:
    if value is None:
        return (1, datetime.max)
    return (0, value.replace(tzinfo=None) if value.tzinfo else value)


def _rail_sort_key(row: PaymentRailRow | BonusRailRow) -> tuple[date, datetime]:
    return (row.audit_date, row.occurred_at or datetime.max)


def _creator_club_excel_time(occurred_at: datetime | None) -> datetime | None:
    """Payment/bonus time as naive datetime in Creator Club audit timezone (UTC-4)."""
    if occurred_at is None:
        return None
    dt = occurred_at
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    local = dt.astimezone(zone_for_slug(_CREATOR_CLUB_SLUG))
    return local.replace(tzinfo=None)


def _crypto_from_label(data: dict) -> str:
    entity = (data.get("from_entity_name") or "").strip()
    if entity:
        return entity
    return (data.get("from_address") or "").strip()


def _venmo_variant(data: dict) -> str:
    handle = normalize_venmo_handle(data.get("venmo_handle") or "")
    return f"@{handle}" if handle else ""


def _zelle_variant(data: dict) -> str:
    return canonicalize_zelle_recipient(data.get("zelle_recipient") or "") or ""


def _fetch_mateos_payments_for_day(
    session: Session,
    *,
    payment_cls,
    build_read: Callable,
    audit_date: date,
    name_fn: Callable[[dict], str],
    variant_fn: Callable[[dict], str],
) -> list[PaymentRailRow]:
    from_dt, to_dt = partner_audit_day_window_utc(_CREATOR_CLUB_SLUG, audit_date)
    query = _apply_audit_manual_filters(
        session,
        session.query(payment_cls),
        payment_cls,
        from_dt=from_dt,
        to_dt=to_dt,
    )
    query = query.filter(payment_cls.method_owner == METHOD_OWNER_MATEOS)
    query = query.order_by(payment_cls.created_at.asc(), payment_cls.id.asc())
    out: list[PaymentRailRow] = []
    for row in query.all():
        data = build_read(session, row)
        if not payment_in_audit_day_for_club(
            session,
            club_slug=_CREATOR_CLUB_SLUG,
            audit_date=audit_date,
            club_id=data.get("club_id"),
            occurred_at=data.get("created_at"),
            data=data,
        ):
            continue
        amount = data.get("amount_usd")
        amount_f = abs(float(amount)) if amount is not None else 0.0
        out.append(
            PaymentRailRow(
                audit_date=audit_date,
                occurred_at=_creator_club_excel_time(data.get("created_at")),
                name=name_fn(data),
                variant=variant_fn(data),
                amount_usd=amount_f,
            )
        )
    return out


def fetch_mateos_payment_rails(
    session: Session,
    audit_dates: list[date],
) -> dict[str, list[PaymentRailRow]]:
    """Mateos-owned Zelle/Venmo/Crypto deposits for Creator Club audit days."""
    buckets: dict[str, list[PaymentRailRow]] = {
        "zelle": [],
        "venmo": [],
        "crypto": [],
    }
    configs: list[tuple[str, type, Callable, Callable[[dict], str], Callable[[dict], str]]] = [
        (
            "zelle",
            ZellePayment,
            build_zelle_payment_read,
            lambda d: (d.get("payer_name") or "").strip(),
            _zelle_variant,
        ),
        (
            "venmo",
            VenmoPayment,
            build_venmo_payment_read,
            lambda d: (d.get("payer_name") or "").strip(),
            _venmo_variant,
        ),
        (
            "crypto",
            CryptoPayment,
            build_crypto_payment_read,
            _crypto_from_label,
            lambda d: (d.get("token_symbol") or "").strip(),
        ),
    ]
    for audit_date in audit_dates:
        for key, payment_cls, build_read, name_fn, variant_fn in configs:
            buckets[key].extend(
                _fetch_mateos_payments_for_day(
                    session,
                    payment_cls=payment_cls,
                    build_read=build_read,
                    audit_date=audit_date,
                    name_fn=name_fn,
                    variant_fn=variant_fn,
                )
            )
    for rows in buckets.values():
        rows.sort(key=_rail_sort_key)
    return buckets


def fetch_creator_club_bonus_rails(
    session: Session,
    audit_dates: list[date],
) -> list[BonusRailRow]:
    """All Creator Club bonus records for the given audit days."""
    club_id = resolve_club_id(session, _CREATOR_CLUB_SLUG)
    out: list[BonusRailRow] = []
    for audit_date in audit_dates:
        from_dt, to_dt = partner_audit_day_window_utc(_CREATOR_CLUB_SLUG, audit_date)
        records = (
            session.query(BonusRecord)
            .filter(
                BonusRecord.club_id == club_id,
                BonusRecord.created_at >= from_dt,
                BonusRecord.created_at <= to_dt,
            )
            .order_by(BonusRecord.created_at.asc(), BonusRecord.id.asc())
            .all()
        )
        for record in records:
            if not payment_in_audit_day_for_club(
                session,
                club_slug=_CREATOR_CLUB_SLUG,
                audit_date=audit_date,
                club_id=record.club_id,
                occurred_at=record.created_at,
            ):
                continue
            out.append(
                BonusRailRow(
                    audit_date=audit_date,
                    occurred_at=_creator_club_excel_time(record.created_at),
                    player=str(record.player_username).strip(),
                    amount_usd=float(Decimal(str(record.amount))),
                )
            )
    out.sort(key=_rail_sort_key)
    return out


def _style_headers(ws: Worksheet, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _clear_sheet_body(ws: Worksheet) -> None:
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)


def _resize_processed_table(ws: Worksheet, last_row: int) -> None:
    end_col = get_column_letter(_PROCESSED_COL_COUNT)
    ref = f"A1:{end_col}{max(last_row, 2)}"
    if _PROCESSED_TABLE not in ws.tables:
        tab = Table(
            displayName=_PROCESSED_TABLE,
            ref=ref,
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        ws.add_table(tab)
        return
    ws.tables[_PROCESSED_TABLE].ref = ref


def _write_processed(ws: Worksheet, rows: list[MatchingRow]) -> None:
    _clear_sheet_body(ws)
    for col, header in enumerate(PROCESSED_HEADERS, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    if ws["H1"].value is None:
        ws["H1"] = "Pivot Table"
        ws["H1"].font = Font(bold=True)

    for offset, row in enumerate(rows):
        r = offset + 2
        values = [
            _display_cell(row.trade_time),
            _display_cell(row.manager),
            _display_cell(_as_float(row.amount) if row.amount is not None else None),
            _display_cell(row.player_id),
            _display_cell(row.nickname),
            _display_cell(row.source),
            row.audit_date.isoformat(),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(r, col, value)
            if col == 3 and isinstance(value, (int, float)):
                cell.number_format = _CURRENCY_FORMAT

    last_row = 1 + len(rows) if rows else 2
    if not rows:
        for col in range(1, _PROCESSED_COL_COUNT + 1):
            ws.cell(2, col, MISSING_DATA if col < _PROCESSED_COL_COUNT else None)
        last_row = 2
    _resize_processed_table(ws, last_row)

    for col in range(1, _PROCESSED_COL_COUNT + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14


def _as_float(value: Decimal | float | int | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _payment_rail_tuple(row: PaymentRailRow) -> tuple[object, ...]:
    return (
        _display_cell(row.occurred_at),
        _display_cell(row.name),
        _display_cell(row.amount_usd),
        _display_cell(row.variant),
        row.audit_date.isoformat(),
    )


def _bonus_rail_tuple(row: BonusRailRow) -> tuple[object, ...]:
    return (
        _display_cell(row.occurred_at),
        _display_cell(row.player),
        _display_cell(row.amount_usd),
        row.audit_date.isoformat(),
    )


def _write_rail_sheet(
    ws: Worksheet,
    *,
    headers: list[str],
    rows: list[tuple[object, ...]],
    amount_col: int,
) -> None:
    ws.delete_rows(1, ws.max_row)
    _style_headers(ws, headers)
    for offset, values in enumerate(rows):
        r = offset + 2
        for col, value in enumerate(values, start=1):
            cell = ws.cell(r, col, value)
            if col == amount_col and isinstance(value, (int, float)):
                cell.number_format = _CURRENCY_FORMAT

    total_row = 2 + len(rows)
    total = 0.0
    has_any = False
    for values in rows:
        amount = values[amount_col - 1]
        if isinstance(amount, (int, float)):
            total += float(amount)
            has_any = True
    if has_any or rows:
        cell = ws.cell(total_row, amount_col, total if has_any else 0)
        cell.number_format = _CURRENCY_FORMAT
        cell.font = Font(bold=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def build_creator_weekly_audit_workbook(
    monday: date,
    files: list[tuple[str, bytes]],
    *,
    session: Session,
) -> bytes:
    """Build weekly audit XLSX from Matching exports (Processed) + DB payment rails."""
    expected = validate_upload_set(monday, [name for name, _ in files])
    by_date = {date_from_filename(name): (name, raw) for name, raw in files}

    all_rows: list[MatchingRow] = []
    for day in expected:
        name, raw = by_date[day]
        if not raw:
            raise CreatorWeeklyAuditError(f"{Path(name).name}: file is empty.")
        try:
            wb = load_workbook(io.BytesIO(raw), data_only=False)
        except Exception as exc:  # noqa: BLE001 — surface as clear 400
            raise CreatorWeeklyAuditError(
                f"{Path(name).name}: could not read workbook ({exc})."
            ) from exc
        all_rows.extend(parse_creator_club_rows(wb, filename=name))

    all_rows.sort(key=lambda r: _sort_key_time(r.trade_time))

    payment_rails = fetch_mateos_payment_rails(session, expected)
    bonus_rails = fetch_creator_club_bonus_rails(session, expected)

    if TEMPLATE_PATH.is_file():
        out_wb = load_workbook(TEMPLATE_PATH)
    else:
        out_wb = Workbook()
        out_wb.active.title = "Processed"
        for title in ("Zelle", "Venmo", "Crypto", "Bonuses"):
            out_wb.create_sheet(title)

    for title in ("Processed", "Zelle", "Venmo", "Crypto", "Bonuses"):
        if title not in out_wb.sheetnames:
            out_wb.create_sheet(title)

    _write_processed(out_wb["Processed"], all_rows)

    _write_rail_sheet(
        out_wb["Zelle"],
        headers=ZELLE_VENMO_HEADERS,
        rows=[_payment_rail_tuple(r) for r in payment_rails["zelle"]],
        amount_col=3,
    )

    _write_rail_sheet(
        out_wb["Venmo"],
        headers=ZELLE_VENMO_HEADERS,
        rows=[_payment_rail_tuple(r) for r in payment_rails["venmo"]],
        amount_col=3,
    )

    _write_rail_sheet(
        out_wb["Crypto"],
        headers=CRYPTO_HEADERS,
        rows=[_payment_rail_tuple(r) for r in payment_rails["crypto"]],
        amount_col=3,
    )

    _write_rail_sheet(
        out_wb["Bonuses"],
        headers=BONUSES_HEADERS,
        rows=[_bonus_rail_tuple(r) for r in bonus_rails],
        amount_col=3,
    )

    desired = ["Processed", "Zelle", "Venmo", "Crypto", "Bonuses"]
    for idx, title in enumerate(desired):
        out_wb.move_sheet(title, offset=idx - out_wb.sheetnames.index(title))

    buf = io.BytesIO()
    out_wb.save(buf)
    return buf.getvalue()


def build_creator_weekly_audit_from_uploads(
    *,
    monday: str | date,
    uploads: list[tuple[str, BinaryIO | bytes]],
    session: Session,
) -> tuple[bytes, str]:
    """Validate monday + uploads; return (xlsx_bytes, download_filename)."""
    monday_date = monday if isinstance(monday, date) else parse_monday(monday)
    files: list[tuple[str, bytes]] = []
    for name, payload in uploads:
        if hasattr(payload, "read"):
            raw = payload.read()
        else:
            raw = payload
        if not isinstance(raw, (bytes, bytearray)):
            raise CreatorWeeklyAuditError(f"{Path(name).name}: could not read file bytes.")
        files.append((name, bytes(raw)))
    content = build_creator_weekly_audit_workbook(monday_date, files, session=session)
    return content, output_filename(monday_date)
