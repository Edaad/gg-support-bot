"""XLSX export for audit reconcile runs."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from api.audit_ledger import (
    CASHOUT_SOURCE_LABELS,
    LEDGER_SOURCE_LABELS,
    LedgerLine,
    UNION_MATCHING_SOURCE_OPTIONS,
)
from api.audit_reconcile import AuditReconcileReport
from api.audit_reconcile_matching import (
    BACK_TO_CLUB_LABEL,
    CHIP_TRANSFER_AT_CC_LABEL,
    CHIP_TRANSFER_PLAYER_LABEL,
    CHIP_TRANSFER_RT_AT_LABEL,
    FREE_PLAY_LABEL,
    GTO_INC_LABEL,
    MatchedTradeRow,
    TradeLedgerMatchResult,
    _sort_key_occurred_at,
    apply_cc_at_aces_ledger_fallback,
    apply_chip_transfer_matches,
    apply_trade_record_source_overrides,
    match_trade_lines_to_ledger,
)
from api.club_audit_timezone import zone_for_payment_display
from api.vaughn_methods import (
    VAUGHN_VENMO_HANDLES,
    VAUGHN_ZELLE_RECIPIENTS,
    clubgto_matching_source_options,
    matching_source_label,
)

_HEADER_FILL = PatternFill("solid", fgColor="38761D")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SECTION_FONT = Font(bold=True, size=12)
_CURRENCY_FORMAT = '$#,##0.00;[Red]-$#,##0.00'

# Matching sheet look (aligned to native Sheets "Table1" reference).
_MATCHING_HEADER_FILL = PatternFill("solid", fgColor="306A54")
_MATCHING_BAND_FILL = PatternFill("solid", fgColor="F6F6F9")
_MATCHING_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_MATCHING_BODY_FONT = Font(name="Arial", size=11)
_MATCHING_ROW_HEIGHT = 18

MATCHING_HEADERS = [
    "Trade Time",
    "Manager",
    "Amount",
    "Player ID",
    "Nickname",
    "Source",
    "Name",
    "Match Time",
    "$",
    "Variant",
]

UNRESOLVED_HEADERS = [
    "Source",
    "Variant",
    "Amount",
    "Name / player",
    "Group",
    "Club",
    "Time",
]

VAUGHN_TALLY_HEADERS = [
    "Method",
    "Tag",
    "Count",
    "Total",
]

# Excel Time number format (underlying value is full datetime).
_EXCEL_TIME_FORMAT = "h:mm:ss AM/PM"

# Trade Time / Match Time widened so headers are not truncated.
MATCHING_WIDTHS = [18, 18, 12, 16, 22, 14, 22, 18, 10, 22, 3, 12, 18, 10, 14]
UNRESOLVED_WIDTHS = [18, 22, 12, 28, 40, 16, 24]

# Matching sheet: left table cols 1–10; spacer 11; Vaughn tally starts at 12.
_MATCHING_TALLY_START_COL = 12
_MATCHING_TABLE_COLS = len(MATCHING_HEADERS)
_MATCHING_AMOUNT_COL = 3  # Trade Amount (chips)
_MATCHING_SOURCE_COL = 6  # Source (after Nickname)
_MATCHING_DOLLAR_COL = 9  # $
_MATCHING_VARIANT_COL = 10  # Variant

MATCHING_SOURCE_OPTIONS: tuple[str, ...] = tuple(
    label for src, label in LEDGER_SOURCE_LABELS.items() if src != "cashout"
) + CASHOUT_SOURCE_LABELS + UNION_MATCHING_SOURCE_OPTIONS + (
    CHIP_TRANSFER_PLAYER_LABEL,
    FREE_PLAY_LABEL,
    BACK_TO_CLUB_LABEL,
)

# Matching Source fills: one family color; GTO / Vaughn uses the darker shade.
_MATCHING_SOURCE_FILL_HEX: dict[str, str] = {
    "Stripe": "E6D5F5",
    "GTO Stripe": "C4A3E0",
    "Zelle": "F5D0F0",
    "Union Zelle": "F5D0F0",
    "RT Zelle": "F5D0F0",
    "Cashout Zelle": "F5D0F0",
    "GTO Zelle": "D98BCF",
    "Vaughn Cashout Zelle": "D98BCF",
    "Venmo": "D4E6F7",
    "RT Venmo": "D4E6F7",
    "Cashout Venmo": "D4E6F7",
    "GTO Venmo": "7FB3D5",
    "Vaughn Cashout Venmo": "7FB3D5",
    "Cash App": "D5F5E3",
    "Union Cash App": "D5F5E3",
    "RT Cash App": "D5F5E3",
    "Cashout Cash App": "D5F5E3",
    "Vaughn Cashout Cash App": "7DCEA0",
    "PayPal": "D6F5FB",
    "Union Apple Pay": "E8DAEF",
    "Union Venmo": "D4E6F7",
    "RT PayPal": "D6F5FB",
    "Cashout PayPal": "D6F5FB",
    "Crypto": "FDEBD0",
    "Cashout Crypto": "FDEBD0",
    "GTO Crypto": "F0B27A",
    "Vaughn Cashout Crypto": "F0B27A",
    "Cashout Revolut": "D5F5F5",
    "Early RB": "FCF3CF",
    "Bonus": "FADBD8",
    "RB settlement (Monday)": "D5D8DC",
    "Cashout": "E5E7E9",
    FREE_PLAY_LABEL: "FADBD8",
    BACK_TO_CLUB_LABEL: "E5E7E9",
    GTO_INC_LABEL: "C4A3E0",
    CHIP_TRANSFER_PLAYER_LABEL: "C5CAE9",
    CHIP_TRANSFER_RT_AT_LABEL: "9FA8DA",
    CHIP_TRANSFER_AT_CC_LABEL: "7986CB",
}


def _variant_options_by_source(
    ledger_lines: list[LedgerLine],
    *,
    club_slug: str,
) -> dict[str, list[str]]:
    buckets: dict[str, set[str]] = {}
    for line in ledger_lines:
        label = matching_source_label(
            source=line.source,
            variant=line.variant,
            club_slug=club_slug,
            source_label=line.source_label,
            memo=line.memo,
        )
        if not label:
            continue
        buckets.setdefault(label, set())
        tag = (line.variant or "").strip()
        if tag:
            buckets[label].add(tag)
    return {label: sorted(tags) for label, tags in buckets.items()}


def _matching_source_dropdown_options(club_slug: str) -> tuple[str, ...]:
    key = club_slug.strip().lower()
    if key == "clubgto":
        return clubgto_matching_source_options() + (CHIP_TRANSFER_PLAYER_LABEL,)
    extra: tuple[str, ...] = ()
    if key in {"round-table", "aces-table"}:
        extra += (CHIP_TRANSFER_RT_AT_LABEL,)
    if key in {"aces-table", "creator-club"}:
        extra += (CHIP_TRANSFER_AT_CC_LABEL,)
    return MATCHING_SOURCE_OPTIONS + extra


# Far-right lookup cols on Matching (hidden). Source list is vertical so
# data validation can use a same-sheet range (Sheets drops extra-tab lists).
_SOURCE_LIST_COL = 30
_VARIANT_LISTS_START_COL = 31


def _style_matching_block(
    ws: Worksheet,
    *,
    header_row: int,
    last_row: int,
    num_cols: int,
) -> None:
    """Matching look + AutoFilter (not an Excel Table — Sheets drops DV on tables)."""
    end_col = get_column_letter(num_cols)
    end_row = max(last_row, header_row)
    ws.auto_filter.ref = f"A{header_row}:{end_col}{end_row}"

    ws.row_dimensions[header_row].height = _MATCHING_ROW_HEIGHT
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = _MATCHING_HEADER_FILL
        cell.font = _MATCHING_HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in range(header_row + 1, end_row + 1):
        ws.row_dimensions[row].height = _MATCHING_ROW_HEIGHT
        band = (row - header_row) % 2 == 0
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _MATCHING_BODY_FONT
            if band:
                cell.fill = _MATCHING_BAND_FILL


def _add_excel_table(
    ws: Worksheet,
    *,
    display_name: str,
    header_row: int,
    last_row: int,
    num_cols: int,
) -> None:
    """Register an Excel Table over header+data (filters + Matching look)."""
    end_col = get_column_letter(num_cols)
    end_row = max(last_row, header_row)
    tab = Table(
        displayName=display_name,
        ref=f"A{header_row}:{end_col}{end_row}",
    )
    # Stripes off — we paint soft #F6F6F9 bands (TableStyleLight1 grey is too dark).
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    ws.row_dimensions[header_row].height = _MATCHING_ROW_HEIGHT
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = _MATCHING_HEADER_FILL
        cell.font = _MATCHING_HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in range(header_row + 1, end_row + 1):
        ws.row_dimensions[row].height = _MATCHING_ROW_HEIGHT
        band = (row - header_row) % 2 == 0
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _MATCHING_BODY_FONT
            if band:
                cell.fill = _MATCHING_BAND_FILL


def _unresolved_player_name(line: LedgerLine) -> str:
    for candidate in (
        line.display_name,
        line.member_nickname,
        line.gg_player_id,
    ):
        text = (candidate or "").strip()
        if text:
            return text
    return ""


def _format_unresolved_time(club_slug: str, occurred_at: datetime | None) -> str:
    """Match audit unresolved sheet: 'Jul 21st 2026, 1:40 AM' (America/New_York)."""
    local = _local_datetime(club_slug, occurred_at)
    if local is None:
        return ""
    day = local.day
    if 10 <= day % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    month = local.strftime("%b")
    clock = local.strftime("%I:%M %p").lstrip("0")
    return f"{month} {day}{suffix} {local.year}, {clock}"


def _unresolved_table_display_name(suffix: str = "all") -> str:
    safe = "".join(ch if ch.isalnum() else "_" for ch in suffix.strip().lower())
    return f"Unresolved_{safe or 'all'}"


def _write_unresolved_sheet(
    ws: Worksheet,
    rows: list[tuple[LedgerLine, str, str]],
    *,
    table_suffix: str = "all",
) -> None:
    """rows: (ledger_line, club_slug, club_name).

    Source uses Matching-style mapping (e.g. RT Zelle / GTO Zelle); Variant is
    the payment tag (recipient / handle) when present.
    """
    header_row = 1
    _style_header_row(ws, header_row, UNRESOLVED_HEADERS)
    row_idx = header_row + 1
    amount_col = 3
    for line, club_slug, club_name in rows:
        label_slug = (line.club_slug or club_slug or "").strip() or club_slug
        ws.cell(
            row=row_idx,
            column=1,
            value=matching_source_label(
                source=line.source,
                variant=line.variant,
                club_slug=label_slug,
                source_label=line.source_label,
                memo=line.memo,
            ),
        )
        ws.cell(
            row=row_idx,
            column=2,
            value=(line.variant or "").strip() or None,
        )
        amount_cell = ws.cell(
            row=row_idx,
            column=amount_col,
            value=_decimal_cell(abs(line.amount_signed)),
        )
        amount_cell.number_format = "0.##"
        amount_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=4, value=_unresolved_player_name(line))
        ws.cell(row=row_idx, column=5, value=(line.detail or "").strip())
        ws.cell(row=row_idx, column=6, value=(club_name or "").strip())
        ws.cell(
            row=row_idx,
            column=7,
            value=_format_unresolved_time(club_slug, line.occurred_at_utc),
        )
        row_idx += 1

    last_data_row = row_idx - 1
    _add_excel_table(
        ws,
        display_name=_unresolved_table_display_name(table_suffix),
        header_row=header_row,
        last_row=last_data_row,
        num_cols=len(UNRESOLVED_HEADERS),
    )
    for row in range(header_row + 1, last_data_row + 1):
        ws.cell(row=row, column=amount_col).alignment = Alignment(horizontal="right")


def _add_matching_source_variant_dropdowns(
    ws: Worksheet,
    report: AuditReconcileReport,
    *,
    first_row: int,
    last_row: int,
    dropdown_club_slug: str | None = None,
) -> None:
    """Source list + Variant list dependent on Source (Excel / Sheets DV)."""
    if last_row < first_row:
        return

    by_source = _variant_options_by_source(
        report.ledger_lines,
        club_slug=report.club_slug,
    )
    list_slug = (dropdown_club_slug or report.club_slug).strip().lower()
    source_options = list(_matching_source_dropdown_options(list_slug))
    source_columns = list(source_options)
    for label in by_source:
        if label not in source_columns:
            source_columns.append(label)

    source_list_letter = get_column_letter(_SOURCE_LIST_COL)
    n_sources = max(len(source_columns), 1)
    for row_i, source in enumerate(source_columns, start=1):
        ws.cell(row=row_i, column=_SOURCE_LIST_COL, value=source)
    ws.column_dimensions[source_list_letter].hidden = True
    # No leading "=" — Sheets drops list DV that starts with "=" or another tab.
    source_list_formula = (
        f"{quote_sheetname(ws.title)}!${source_list_letter}$1:"
        f"${source_list_letter}${n_sources}"
    )

    lists_start_col = _VARIANT_LISTS_START_COL

    max_option_rows = 1
    for offset, source in enumerate(source_columns):
        col_idx = lists_start_col + offset
        ws.cell(row=1, column=col_idx, value=source)
        options = by_source.get(source) or []
        if options:
            for row_i, opt in enumerate(options, start=2):
                ws.cell(row=row_i, column=col_idx, value=opt)
            max_option_rows = max(max_option_rows, len(options))
        else:
            ws.cell(row=2, column=col_idx, value="—")
            max_option_rows = max(max_option_rows, 1)
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    lists_end_col = lists_start_col + len(source_columns) - 1
    header_range = (
        f"${get_column_letter(lists_start_col)}$1:"
        f"${get_column_letter(lists_end_col)}$1"
    )
    list_end_row = max(2, 1 + max_option_rows)

    source_letter = get_column_letter(_MATCHING_SOURCE_COL)
    variant_letter = get_column_letter(_MATCHING_VARIANT_COL)
    source_sqref = f"{source_letter}{first_row}:{source_letter}{last_row}"
    variant_sqref = f"{variant_letter}{first_row}:{variant_letter}{last_row}"

    # showDropDown=False is the openpyxl quirk that *shows* the in-cell dropdown.
    dv_source = DataValidation(
        type="list",
        formula1=source_list_formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv_source)
    dv_source.add(source_sqref)

    # MATCH finds the Source header among hidden list columns; ADDRESS builds
    # that column's variant range. $F locks column; row stays relative per DV cell.
    first_source = f"${source_letter}{first_row}"
    col_expr = f"MATCH({first_source},{header_range},0)+{lists_start_col - 1}"
    variant_formula = (
        f"=INDIRECT(ADDRESS(2,{col_expr})&\":\"&ADDRESS({list_end_row},{col_expr}))"
    )
    dv_variant = DataValidation(
        type="list",
        formula1=variant_formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv_variant)
    dv_variant.add(variant_sqref)


def _add_matching_source_colors(
    ws: Worksheet,
    *,
    first_row: int,
    last_row: int,
) -> None:
    """Color Matching Source cells by label; updates when the dropdown changes."""
    if last_row < first_row:
        return
    source_letter = get_column_letter(_MATCHING_SOURCE_COL)
    sqref = f"{source_letter}{first_row}:{source_letter}{last_row}"
    for label, hex_color in _MATCHING_SOURCE_FILL_HEX.items():
        escaped = label.replace('"', '""')
        ws.conditional_formatting.add(
            sqref,
            CellIsRule(
                operator="equal",
                formula=[f'"{escaped}"'],
                fill=PatternFill("solid", fgColor=hex_color),
            ),
        )


ALL_CLUBS_MATCHING_SHEET_ORDER: tuple[tuple[str, str], ...] = (
    ("round-table", "Round Table"),
    ("aces-table", "Aces Table"),
    ("clubgto", "ClubGTO"),
    ("creator-club", "Creator Club"),
)

# Display labels for matching tabs / Unresolved (aces ≠ DB CLUB_SLUG_TO_NAME).
MATCHING_CLUB_DISPLAY: dict[str, str] = dict(ALL_CLUBS_MATCHING_SHEET_ORDER)

_ROUND_TABLE_COMPOSITE_SHEET_SLUGS: frozenset[str] = frozenset(
    {"round-table", "aces-table"}
)


def _decimal_cell(value: Decimal) -> float:
    return float(value)


def _style_header_row(
    ws: Worksheet,
    row: int,
    headers: list[str],
    *,
    start_col: int = 1,
) -> None:
    for offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + offset, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_section_title(
    ws: Worksheet,
    row: int,
    title: str,
    *,
    col: int = 1,
) -> None:
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = _SECTION_FONT


def _set_column_widths(ws: Worksheet, widths: list[int]) -> None:
    for col_idx, width in enumerate(widths, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = width


def _local_datetime(club_slug: str, occurred_at: datetime | None) -> datetime | None:
    """America/New_York naive datetime suitable for Excel cells (all clubs)."""
    del club_slug  # kept for call-site compatibility; display is always Eastern
    if occurred_at is None:
        return None
    dt = occurred_at
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    local = dt.astimezone(zone_for_payment_display())
    return local.replace(tzinfo=None)


def _write_excel_time_cell(
    ws: Worksheet,
    row: int,
    col: int,
    club_slug: str,
    occurred_at: datetime | None,
) -> None:
    cell = ws.cell(row=row, column=col)
    local = _local_datetime(club_slug, occurred_at)
    if local is None:
        cell.value = None
        return
    cell.value = local
    cell.number_format = _EXCEL_TIME_FORMAT


def _vaughn_tag_scoped(tag: str) -> bool:
    return bool(tag) and not tag.startswith("(")


def _vaughn_countifs_formula(*, source: str, tag: str) -> str:
    """Count Matching rows for a GTO method; Zelle/Venmo also match Variant."""
    src_col = f"${get_column_letter(_MATCHING_SOURCE_COL)}:${get_column_letter(_MATCHING_SOURCE_COL)}"
    if not _vaughn_tag_scoped(tag):
        return f'=COUNTIF({src_col},"{source}")'
    tag_col = f"${get_column_letter(_MATCHING_VARIANT_COL)}:${get_column_letter(_MATCHING_VARIANT_COL)}"
    return f'=COUNTIFS({src_col},"{source}",{tag_col},"{tag}")'


def _vaughn_sum_dollar_formula(*, source: str, tag: str) -> str:
    """Sum $ column for a GTO method; Zelle/Venmo also match Variant."""
    src_col = f"${get_column_letter(_MATCHING_SOURCE_COL)}:${get_column_letter(_MATCHING_SOURCE_COL)}"
    dollar_col = f"${get_column_letter(_MATCHING_DOLLAR_COL)}:${get_column_letter(_MATCHING_DOLLAR_COL)}"
    if not _vaughn_tag_scoped(tag):
        return f'=SUMIF({src_col},"{source}",{dollar_col})'
    tag_col = f"${get_column_letter(_MATCHING_VARIANT_COL)}:${get_column_letter(_MATCHING_VARIANT_COL)}"
    return f'=SUMIFS({dollar_col},{src_col},"{source}",{tag_col},"{tag}")'


def _matching_data_col_range(col_idx: int, first_row: int, last_row: int) -> str:
    letter = get_column_letter(col_idx)
    return f"${letter}${first_row}:${letter}${last_row}"


def _vaughn_sum_chips_formula(
    *,
    source: str,
    tag: str,
    first_row: int,
    last_row: int,
) -> str:
    """Sum abs(Amount) chips; data rows only so Sheets ABS skips the header."""
    src_col = _matching_data_col_range(_MATCHING_SOURCE_COL, first_row, last_row)
    amount_col = _matching_data_col_range(_MATCHING_AMOUNT_COL, first_row, last_row)
    if not _vaughn_tag_scoped(tag):
        return f'=SUMPRODUCT(({src_col}="{source}")*ABS({amount_col}))'
    tag_col = _matching_data_col_range(_MATCHING_VARIANT_COL, first_row, last_row)
    return (
        f'=SUMPRODUCT(({src_col}="{source}")*({tag_col}="{tag}")*ABS({amount_col}))'
    )


def _vaughn_method_rows() -> list[tuple[str, str, str]]:
    """(method_label, tag, matching Source label) for Vaughn tally tables."""
    method_rows: list[tuple[str, str, str]] = []
    for digits in sorted(VAUGHN_ZELLE_RECIPIENTS):
        method_rows.append(("Zelle", digits, "GTO Zelle"))
    for handle in sorted(VAUGHN_VENMO_HANDLES):
        method_rows.append(("Venmo", f"@{handle}", "GTO Venmo"))
    method_rows.append(("Crypto", "(all ClubGTO)", "GTO Crypto"))
    method_rows.append(("Stripe", "(all ClubGTO)", "GTO Stripe"))
    return method_rows


def _write_vaughn_tally_table(
    ws: Worksheet,
    *,
    title: str,
    section_row: int,
    start_col: int,
    total_formula,
) -> int:
    """Write one Vaughn tally; returns the Total row index."""
    _style_section_title(ws, section_row, title, col=start_col)
    header_row = section_row + 1
    _style_header_row(ws, header_row, VAUGHN_TALLY_HEADERS, start_col=start_col)

    first_data = header_row + 1
    row_idx = first_data
    for method_label, tag, source_label in _vaughn_method_rows():
        ws.cell(row=row_idx, column=start_col, value=method_label)
        ws.cell(row=row_idx, column=start_col + 1, value=tag)
        ws.cell(
            row=row_idx,
            column=start_col + 2,
            value=_vaughn_countifs_formula(source=source_label, tag=tag),
        )
        total_cell = ws.cell(
            row=row_idx,
            column=start_col + 3,
            value=total_formula(source=source_label, tag=tag),
        )
        total_cell.number_format = _CURRENCY_FORMAT
        row_idx += 1

    last_data = row_idx - 1
    count_letter = get_column_letter(start_col + 2)
    total_letter = get_column_letter(start_col + 3)

    total_label = ws.cell(row=row_idx, column=start_col, value="Total")
    total_label.font = Font(bold=True)
    count_total = ws.cell(
        row=row_idx,
        column=start_col + 2,
        value=f"=SUM({count_letter}{first_data}:{count_letter}{last_data})",
    )
    count_total.font = Font(bold=True)
    total_cell = ws.cell(
        row=row_idx,
        column=start_col + 3,
        value=f"=SUM({total_letter}{first_data}:{total_letter}{last_data})",
    )
    total_cell.number_format = _CURRENCY_FORMAT
    total_cell.font = Font(bold=True)
    return row_idx


def _write_vaughn_tally(
    ws: Worksheet,
    *,
    section_row: int,
    start_col: int,
    matching_first_row: int,
    matching_last_row: int,
) -> None:
    """Vaughn receipt ($) tally, then chips (abs Amount) tally underneath."""
    receipt_total_row = _write_vaughn_tally_table(
        ws,
        title="Vaughn methods",
        section_row=section_row,
        start_col=start_col,
        total_formula=_vaughn_sum_dollar_formula,
    )
    chips_section_row = receipt_total_row + 2  # blank spacer row

    def chips_formula(*, source: str, tag: str) -> str:
        return _vaughn_sum_chips_formula(
            source=source,
            tag=tag,
            first_row=matching_first_row,
            last_row=matching_last_row,
        )

    _write_vaughn_tally_table(
        ws,
        title="Vaughn methods (chips)",
        section_row=chips_section_row,
        start_col=start_col,
        total_formula=chips_formula,
    )


def _write_matching_rows(
    ws: Worksheet,
    rows: list[MatchedTradeRow],
    *,
    time_club_slug: str,
    table_slug: str,
    report_for_dropdowns: AuditReconcileReport,
    sheet_title: str | None = None,
    include_vaughn_tally: bool = False,
) -> None:
    if sheet_title:
        ws.title = sheet_title

    header_row = 1
    _style_header_row(ws, header_row, MATCHING_HEADERS)

    row_idx = header_row + 1
    for matched in rows:
        trade = matched.trade
        _write_excel_time_cell(
            ws, row_idx, 1, time_club_slug, trade.occurred_at
        )
        ws.cell(row=row_idx, column=2, value=trade.manager_nickname or "")
        cell = ws.cell(
            row=row_idx,
            column=3,
            value=_decimal_cell(trade.amount),
        )
        cell.number_format = _CURRENCY_FORMAT
        ws.cell(row=row_idx, column=4, value=trade.member_gg_player_id or "")
        ws.cell(row=row_idx, column=5, value=trade.member_nickname or "")
        ws.cell(row=row_idx, column=6, value=matched.match_source)
        ws.cell(row=row_idx, column=7, value=matched.match_name)
        _write_excel_time_cell(
            ws, row_idx, 8, time_club_slug, matched.match_occurred_at
        )
        if matched.match_amount is not None:
            dollar_cell = ws.cell(
                row=row_idx,
                column=9,
                value=_decimal_cell(matched.match_amount),
            )
            dollar_cell.number_format = _CURRENCY_FORMAT
        else:
            ws.cell(row=row_idx, column=9, value=None)
        ws.cell(row=row_idx, column=10, value=matched.variant)
        row_idx += 1

    last_data_row = row_idx - 1
    _style_matching_block(
        ws,
        header_row=header_row,
        last_row=last_data_row,
        num_cols=_MATCHING_TABLE_COLS,
    )
    _add_matching_source_variant_dropdowns(
        ws,
        report_for_dropdowns,
        first_row=header_row + 1,
        last_row=last_data_row,
        dropdown_club_slug=table_slug,
    )
    _add_matching_source_colors(
        ws,
        first_row=header_row + 1,
        last_row=last_data_row,
    )

    if include_vaughn_tally:
        _write_vaughn_tally(
            ws,
            section_row=header_row,
            start_col=_MATCHING_TALLY_START_COL,
            matching_first_row=header_row + 1,
            matching_last_row=max(last_data_row, header_row + 1),
        )


def _partition_matching_rows(
    rows: list[MatchedTradeRow],
) -> dict[str, list[MatchedTradeRow]]:
    by_slug: dict[str, list[MatchedTradeRow]] = {
        slug: [] for slug, _ in ALL_CLUBS_MATCHING_SHEET_ORDER
    }
    for matched in rows:
        slug = (matched.trade.trade_club_slug or "round-table").strip().lower()
        if slug not in by_slug:
            slug = "round-table"
        by_slug[slug].append(matched)
    return by_slug


def build_all_clubs_matching_workbook(
    reports_by_slug: dict[str, AuditReconcileReport],
) -> bytes:
    """Matching sheets per club + shared Unresolved tab.

    Round Table + Aces share one composite ledger match. Unmatched Aces trades
    may then match Creator Club ledger lines whose group title is CC AT. Then
    leftover trades are paired together (player, RT↔AT, AT↔CC) and split back
    onto sheets by trade upload slug.
    """
    wb = Workbook()
    unresolved_rows: list[tuple[LedgerLine, str, str]] = []

    rt_report = reports_by_slug["round-table"]
    rt_match = match_trade_lines_to_ledger(
        rt_report.trade_lines,
        rt_report.ledger_lines,
        club_slug=rt_report.club_slug,
    )
    all_rows: list[MatchedTradeRow] = list(rt_match.rows)
    other_matches: dict[str, TradeLedgerMatchResult] = {}
    for slug in ("clubgto", "creator-club"):
        report = reports_by_slug[slug]
        result = match_trade_lines_to_ledger(
            report.trade_lines,
            report.ledger_lines,
            club_slug=report.club_slug,
        )
        other_matches[slug] = result
        all_rows.extend(result.rows)

    all_rows, cc_unmatched_ledger = apply_cc_at_aces_ledger_fallback(
        all_rows,
        other_matches["creator-club"].unmatched_ledger,
    )
    partitioned = _partition_matching_rows(
        apply_trade_record_source_overrides(apply_chip_transfer_matches(all_rows))
    )

    first = True
    for slug, title in ALL_CLUBS_MATCHING_SHEET_ORDER:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()

        if slug in _ROUND_TABLE_COMPOSITE_SHEET_SLUGS:
            dropdown_report = rt_report
        else:
            dropdown_report = reports_by_slug[slug]
        _write_matching_rows(
            ws,
            partitioned[slug],
            time_club_slug=slug,
            table_slug=slug,
            report_for_dropdowns=dropdown_report,
            sheet_title=title,
            include_vaughn_tally=slug == "clubgto",
        )
        _set_column_widths(ws, MATCHING_WIDTHS)

    for line in rt_match.unmatched_ledger:
        line_slug = (line.club_slug or "round-table").strip().lower()
        if line_slug not in MATCHING_CLUB_DISPLAY:
            line_slug = "round-table"
        unresolved_rows.append(
            (line, line_slug, MATCHING_CLUB_DISPLAY[line_slug])
        )
    for line in other_matches["clubgto"].unmatched_ledger:
        report = reports_by_slug["clubgto"]
        unresolved_rows.append((line, report.club_slug, report.club_name))
    cc_report = reports_by_slug["creator-club"]
    for line in cc_unmatched_ledger:
        unresolved_rows.append((line, cc_report.club_slug, cc_report.club_name))

    unresolved_rows.sort(
        key=lambda item: (
            _sort_key_occurred_at(item[0].occurred_at_utc),
            item[2],
            item[0].external_id,
        )
    )
    unresolved = wb.create_sheet("Unresolved")
    _write_unresolved_sheet(unresolved, unresolved_rows, table_suffix="all")
    _set_column_widths(unresolved, UNRESOLVED_WIDTHS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
