"""Cross-club audit export: one XLSX workbook with a sheet per payment provider."""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any, Callable, Literal

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func
from sqlalchemy.orm import Session

from api.club_audit_timezone import (
    audit_day_bounds_utc as club_audit_day_bounds_utc,
    audit_day_window_utc as club_audit_day_window_utc,
    occurred_at_in_audit_day,
    union_audit_day_window_utc,
    zone_for_payment_display,
    zone_for_slug,
)
from api.club_slug import CLUB_SLUG_TO_NAME, slug_for_club_id

from api.payments_helpers import (
    apply_analytics_payment_exclusion,
    build_cashapp_payment_read,
    build_paypal_payment_read,
    build_venmo_payment_read,
    build_zelle_payment_read,
    lookup_gg_nickname,
    resolve_group_title,
    resolve_method_display,
)
from config import CLUB_SHORTHAND_TO_NAME
from db.models import (
    BonusRecord,
    CashAppPayment,
    Club,
    EarlyRakebackLine,
    EarlyRakebackSnapshot,
    PayPalPayment,
    StripeCheckoutSession,
    StripeCustomer,
    VenmoPayment,
    ZellePayment,
)

STRIPE_LAYOUT = ["Amount", "Player", "Method", "Group", "Club", "Time"]
MANUAL_LAYOUT = ["Amount", "Name", "Group", "Club", "Time"]
TAGGED_MANUAL_LAYOUT = ["Amount", "Name", "Tag", "Group", "Club", "Time"]

SheetLayout = Literal["stripe", "manual", "tagged_manual"]

_HEADER_FILL = PatternFill("solid", fgColor="38761D")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_CURRENCY_FORMAT = "$#,##0.00"
_COLUMN_WIDTHS: dict[SheetLayout, list[float]] = {
    "stripe": [14, 28, 22, 40, 18, 28],
    "manual": [14, 28, 40, 18, 28],
    "tagged_manual": [14, 28, 22, 40, 18, 28],
}


@dataclass(frozen=True)
class SheetSpec:
    title: str
    headers: list[str]
    layout: SheetLayout


SHEET_SPECS: list[SheetSpec] = [
    SheetSpec("Stripe", STRIPE_LAYOUT, "stripe"),
    SheetSpec("Zelle", TAGGED_MANUAL_LAYOUT, "tagged_manual"),
    SheetSpec("Venmo", TAGGED_MANUAL_LAYOUT, "tagged_manual"),
    SheetSpec("Cash App", TAGGED_MANUAL_LAYOUT, "tagged_manual"),
    SheetSpec("PayPal", TAGGED_MANUAL_LAYOUT, "tagged_manual"),
    SheetSpec("Bonus", MANUAL_LAYOUT, "manual"),
    SheetSpec("Early Rakeback", STRIPE_LAYOUT, "stripe"),
]


@dataclass(frozen=True)
class StripeAuditRow:
    amount_usd: float
    player: str
    method_label: str
    group_title: str
    club_label: str
    time_label: str
    stripe_fee_usd: Decimal


@dataclass(frozen=True)
class ManualAuditRow:
    amount_usd: float
    payer_name: str
    group_title: str
    club_label: str
    time_label: str


@dataclass(frozen=True)
class TaggedManualAuditRow:
    amount_usd: float
    payer_name: str
    account_tag: str
    group_title: str
    club_label: str
    time_label: str


def _stripe_fee_usd(amount_cents: int) -> Decimal:
    return Decimal(round(amount_cents * 0.029 + 30)) / Decimal(100)


def eastern_day_bounds_utc(date_str: str) -> tuple[datetime, datetime]:
    """Return (start, end) of round-table local calendar day as UTC datetimes."""
    return club_audit_day_bounds_utc("round-table", date_str)


def eastern_audit_end_utc(date_str: str) -> datetime:
    """UTC end of round-table audit export window (day + first hour of next)."""
    _, end = club_audit_day_window_utc("round-table", date_str)
    return end


def audit_day_window_utc(date_str: str) -> tuple[datetime, datetime]:
    """UTC bounds for round-table audit export day (backward-compatible default)."""
    return club_audit_day_window_utc("round-table", date_str)


def _to_club_local(dt: datetime, club_slug: str) -> datetime:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    return dt.astimezone(zone_for_slug(club_slug))


def _to_audit_display_local(dt: datetime) -> datetime:
    """Audit export Time column: always US Eastern (America/New_York).

    Per-club timezone policies apply only to audit-day bucketing (and trade
    records), not to payment timestamps shown on the sheet.
    """
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    else:
        dt = dt.astimezone(timezone.utc)
    return dt.astimezone(zone_for_payment_display())


def _to_eastern(dt: datetime) -> datetime:
    return dt.astimezone(zone_for_payment_display())


def _ordinal_day(day: int) -> str:
    if 10 <= day % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return f"{day}{suffix}"


def _fmt_stripe_audit_time(value: datetime | None, club_slug: str = "round-table") -> str:
    if value is None:
        return ""
    dt = _to_audit_display_local(value)
    month = dt.strftime("%b")
    day = _ordinal_day(dt.day)
    year = dt.year
    clock = dt.strftime("%I:%M %p").lstrip("0")
    return f"{month} {day} {year}, {clock}"


def _fmt_manual_audit_time(value: datetime | None, club_slug: str = "round-table") -> str:
    if value is None:
        return ""
    dt = _to_audit_display_local(value)
    month = dt.strftime("%B")
    day = dt.day
    year = dt.year
    clock = dt.strftime("%I:%M %p").lstrip("0")
    return f"{month} {day}, {year} at {clock}"


def _shorthand_for_club_name(club_name: str) -> str:
    lower = (club_name or "").strip().lower()
    if lower == "clubgto":
        return "GTO"
    if lower == "creator club":
        return "CC"
    if lower == "round table":
        return "RT"
    for shorthand, full_name in CLUB_SHORTHAND_TO_NAME.items():
        if full_name.lower() == lower:
            return shorthand
    return (club_name or "").strip()


def _stripe_player_cell(
    *,
    group_title: str | None,
    club_name: str,
    gg_player_id: str | None,
    gg_nickname: str | None,
) -> str:
    title = (group_title or "").strip()
    if title:
        return title
    shorthand = _shorthand_for_club_name(club_name)
    player_id = (gg_player_id or "").strip()
    nickname = (gg_nickname or "").strip()
    parts = [part for part in (shorthand, player_id, nickname) if part]
    return " / ".join(parts)


def _manual_group_cell(data: dict) -> str:
    return str(data.get("group_title") or "").strip()


def _manual_club_name(data: dict, club_names: dict[int, str]) -> str:
    club_id = data.get("club_id")
    if club_id is not None:
        name = club_names.get(int(club_id), "")
        if name:
            return name

    title = str(data.get("group_title") or "").strip()
    if not title:
        return ""

    from bot.services.player_details import parse_group_title_parts

    parsed = parse_group_title_parts(title)
    if not parsed:
        return ""

    for token in sorted(parsed.shorthands):
        full_name = CLUB_SHORTHAND_TO_NAME.get(token)
        if full_name:
            return full_name
    return ""


def _club_name_map(session: Session) -> dict[int, str]:
    return {int(row.id): str(row.name) for row in session.query(Club.id, Club.name).all()}


def _club_name(club_names: dict[int, str], club_id: int | None) -> str:
    if club_id is None:
        return ""
    return club_names.get(int(club_id), "")


def _apply_audit_manual_filters(
    session: Session,
    query,
    payment_cls,
    *,
    from_dt: datetime,
    to_dt: datetime,
):
    query = query.filter(
        payment_cls.is_test.is_(False),
        payment_cls.created_at >= from_dt,
        payment_cls.created_at <= to_dt,
    )
    return apply_analytics_payment_exclusion(
        session, query, payment_cls.telegram_chat_id
    )


def _apply_audit_stripe_filters(query, *, from_dt: datetime, to_dt: datetime):
    effective_dt = func.coalesce(
        StripeCheckoutSession.completed_at,
        StripeCheckoutSession.created_at,
    )
    return query.filter(
        StripeCheckoutSession.status == "complete",
        effective_dt >= from_dt,
        effective_dt <= to_dt,
    )


def _write_formatted_sheet(
    ws: Worksheet,
    spec: SheetSpec,
    stripe_rows: list[StripeAuditRow] | None = None,
    manual_rows: list[ManualAuditRow] | None = None,
    tagged_manual_rows: list[TaggedManualAuditRow] | None = None,
) -> None:
    ws.append(spec.headers)
    for col_idx, header in enumerate(spec.headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    data_row_start = 2
    if spec.layout == "stripe":
        rows = stripe_rows or []
        for row_idx, row in enumerate(rows, start=data_row_start):
            ws.append(
                [
                    row.amount_usd,
                    row.player,
                    row.method_label,
                    row.group_title,
                    row.club_label,
                    row.time_label,
                ]
            )
            amount_cell = ws.cell(row=row_idx, column=1)
            amount_cell.number_format = _CURRENCY_FORMAT
            amount_cell.alignment = Alignment(horizontal="right")
            fee = row.stripe_fee_usd
            amount_cell.comment = Comment(
                f"Stripe fee: ${fee:.2f}",
                "audit-export",
            )
    elif spec.layout == "tagged_manual":
        rows = tagged_manual_rows or []
        for row_idx, row in enumerate(rows, start=data_row_start):
            ws.append(
                [
                    row.amount_usd,
                    row.payer_name,
                    row.account_tag,
                    row.group_title,
                    row.club_label,
                    row.time_label,
                ]
            )
            amount_cell = ws.cell(row=row_idx, column=1)
            amount_cell.number_format = _CURRENCY_FORMAT
            amount_cell.alignment = Alignment(horizontal="right")
    else:
        rows = manual_rows or []
        for row_idx, row in enumerate(rows, start=data_row_start):
            ws.append(
                [
                    row.amount_usd,
                    row.payer_name,
                    row.group_title,
                    row.club_label,
                    row.time_label,
                ]
            )
            amount_cell = ws.cell(row=row_idx, column=1)
            amount_cell.number_format = _CURRENCY_FORMAT
            amount_cell.alignment = Alignment(horizontal="right")

    for col_idx, width in enumerate(_COLUMN_WIDTHS[spec.layout], start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = width

    if spec.headers:
        ws.auto_filter.ref = ws.dimensions


def _slug_for_payment_club(
    session: Session,
    club_id: int | None,
    data: dict | None = None,
) -> str:
    if club_id is not None:
        slug = slug_for_club_id(session, int(club_id))
        if slug:
            return slug
    if data:
        title = str(data.get("group_title") or "").strip()
        if title:
            from bot.services.player_details import parse_group_title_parts

            parsed = parse_group_title_parts(title)
            if parsed:
                for token in sorted(parsed.shorthands):
                    if token == "AT":
                        return "aces-table"
                    if token == "RT":
                        return "round-table"
                    if token == "GTO":
                        return "clubgto"
                    if token == "CC":
                        return "creator-club"
    return "round-table"


def _payment_in_audit_day(
    session: Session,
    *,
    audit_date: str,
    club_id: int | None,
    occurred_at: datetime | None,
    data: dict | None = None,
) -> bool:
    if occurred_at is None:
        return False
    slug = _slug_for_payment_club(session, club_id, data)
    return occurred_at_in_audit_day(occurred_at, slug, audit_date)


def _parse_audit_date_str(audit_date: str):
    from datetime import date as date_cls

    return date_cls.fromisoformat(str(audit_date).strip()[:10])


def _fetch_early_rakeback_rows(
    session: Session,
    audit_date: str,
) -> list[StripeAuditRow]:
    """Read synced early-rakeback lines from Postgres (no live aon-beta call)."""
    parsed_date = _parse_audit_date_str(audit_date)
    snapshots = (
        session.query(EarlyRakebackSnapshot)
        .filter(EarlyRakebackSnapshot.audit_date == parsed_date)
        .all()
    )
    if not snapshots:
        return []

    snapshot_ids = [s.id for s in snapshots]
    slug_by_snapshot = {s.id: (s.club_slug or "round-table") for s in snapshots}

    lines = (
        session.query(EarlyRakebackLine)
        .filter(EarlyRakebackLine.snapshot_id.in_(snapshot_ids))
        .order_by(
            EarlyRakebackLine.occurred_at.desc().nullslast(),
            EarlyRakebackLine.id.desc(),
        )
        .all()
    )

    out: list[StripeAuditRow] = []
    for line in lines:
        club_slug = slug_by_snapshot.get(line.snapshot_id, "round-table")
        club_label = CLUB_SLUG_TO_NAME.get(club_slug, club_slug)
        nickname = (line.member_nickname or "").strip()
        gg_id = (line.gg_player_id or "").strip()
        unmapped = not gg_id
        if unmapped:
            shorthand = _shorthand_for_club_name(club_label)
            parts = [part for part in (shorthand, "UNMAPPED", nickname) if part]
            player = " / ".join(parts) if parts else "UNMAPPED"
            method_label = "Early RB (unmapped)"
        else:
            player = _stripe_player_cell(
                group_title="",
                club_name=club_label,
                gg_player_id=gg_id,
                gg_nickname=nickname or None,
            )
            method_label = "Early RB"
        amount = float(line.amount_usd) if line.amount_usd is not None else 0.0
        out.append(
            StripeAuditRow(
                amount_usd=amount,
                player=player,
                method_label=method_label,
                group_title=nickname,
                club_label=club_label,
                time_label=_fmt_stripe_audit_time(line.occurred_at, club_slug),
                stripe_fee_usd=Decimal(0),
            )
        )
    return out


def build_audit_workbook(session: Session, audit_date: str) -> bytes:
    from_dt, to_dt = union_audit_day_window_utc(audit_date)
    club_names = _club_name_map(session)
    wb = Workbook()
    wb.remove(wb.active)

    stripe_rows = _fetch_stripe_rows(
        session, club_names, from_dt, to_dt, audit_date=audit_date
    )
    zelle_rows = _fetch_tagged_manual_rows(
        session,
        ZellePayment,
        build_zelle_payment_read,
        club_names,
        from_dt,
        to_dt,
        audit_date=audit_date,
        tag_field="zelle_recipient",
    )
    venmo_rows = _fetch_tagged_manual_rows(
        session,
        VenmoPayment,
        build_venmo_payment_read,
        club_names,
        from_dt,
        to_dt,
        audit_date=audit_date,
        tag_field="venmo_handle",
    )
    cashapp_rows = _fetch_tagged_manual_rows(
        session,
        CashAppPayment,
        build_cashapp_payment_read,
        club_names,
        from_dt,
        to_dt,
        audit_date=audit_date,
        tag_field="cashapp_handle",
    )
    paypal_rows = _fetch_tagged_manual_rows(
        session,
        PayPalPayment,
        build_paypal_payment_read,
        club_names,
        from_dt,
        to_dt,
        audit_date=audit_date,
        tag_field="paypal_email",
    )
    bonus_rows = _fetch_bonus_rows(
        session, club_names, from_dt, to_dt, audit_date=audit_date
    )
    early_rb_rows = _fetch_early_rakeback_rows(session, audit_date)

    sheet_rows: list[
        list[StripeAuditRow] | list[ManualAuditRow] | list[TaggedManualAuditRow]
    ] = [
        stripe_rows,
        zelle_rows,
        venmo_rows,
        cashapp_rows,
        paypal_rows,
        bonus_rows,
        early_rb_rows,
    ]

    for spec, rows in zip(SHEET_SPECS, sheet_rows):
        ws = wb.create_sheet(title=spec.title)
        if spec.layout == "stripe":
            _write_formatted_sheet(ws, spec, stripe_rows=rows)  # type: ignore[arg-type]
        elif spec.layout == "tagged_manual":
            _write_formatted_sheet(ws, spec, tagged_manual_rows=rows)  # type: ignore[arg-type]
        else:
            _write_formatted_sheet(ws, spec, manual_rows=rows)  # type: ignore[arg-type]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fetch_stripe_rows(
    session: Session,
    club_names: dict[int, str],
    from_dt: datetime,
    to_dt: datetime,
    *,
    audit_date: str,
) -> list[StripeAuditRow]:
    query = _apply_audit_stripe_filters(
        session.query(StripeCheckoutSession),
        from_dt=from_dt,
        to_dt=to_dt,
    )
    effective_dt = func.coalesce(
        StripeCheckoutSession.completed_at,
        StripeCheckoutSession.created_at,
    )
    rows = query.order_by(effective_dt.desc(), StripeCheckoutSession.id.desc()).all()

    customer_by_stripe_id: dict[str, StripeCustomer] = {}
    if rows:
        customer_ids = {row.stripe_customer_id for row in rows}
        for cust in (
            session.query(StripeCustomer)
            .filter(StripeCustomer.stripe_customer_id.in_(customer_ids))
            .all()
        ):
            customer_by_stripe_id[cust.stripe_customer_id] = cust

    out: list[StripeAuditRow] = []
    for row in rows:
        completed = row.completed_at or row.created_at
        if not _payment_in_audit_day(
            session,
            audit_date=audit_date,
            club_id=row.club_id,
            occurred_at=completed,
        ):
            continue

        cust = customer_by_stripe_id.get(row.stripe_customer_id)
        title, gg_id = resolve_group_title(
            session,
            row.telegram_chat_id,
            fallback_gg_player_id=cust.gg_player_id if cust else None,
        )
        nickname = lookup_gg_nickname(session, row.club_id, gg_id) or ""
        club_label = _club_name(club_names, row.club_id)
        club_slug = _slug_for_payment_club(session, row.club_id)
        method_label, _ = resolve_method_display(
            session, int(row.club_id), row.payment_method_id
        )
        out.append(
            StripeAuditRow(
                amount_usd=float(row.amount_cents) / 100.0,
                player=_stripe_player_cell(
                    group_title=title,
                    club_name=club_label,
                    gg_player_id=gg_id,
                    gg_nickname=nickname,
                ),
                method_label=(method_label or "").strip(),
                group_title=(title or "").strip(),
                club_label=club_label,
                time_label=_fmt_stripe_audit_time(completed, club_slug),
                stripe_fee_usd=_stripe_fee_usd(row.amount_cents),
            )
        )
    return out


def _fetch_tagged_manual_rows(
    session: Session,
    payment_cls,
    build_read: Callable,
    club_names: dict[int, str],
    from_dt: datetime,
    to_dt: datetime,
    *,
    audit_date: str,
    tag_field: str,
) -> list[TaggedManualAuditRow]:
    query = _apply_audit_manual_filters(
        session,
        session.query(payment_cls),
        payment_cls,
        from_dt=from_dt,
        to_dt=to_dt,
    )
    rows = query.order_by(payment_cls.created_at.desc(), payment_cls.id.desc()).all()
    out: list[TaggedManualAuditRow] = []
    for row in rows:
        data = build_read(session, row)
        if not _payment_in_audit_day(
            session,
            audit_date=audit_date,
            club_id=data.get("club_id"),
            occurred_at=data.get("created_at"),
            data=data,
        ):
            continue
        out.append(_tagged_manual_row(session, data, club_names, tag_field=tag_field))
    return out


def _tagged_manual_row(
    session: Session,
    data: dict,
    club_names: dict[int, str],
    *,
    tag_field: str,
) -> TaggedManualAuditRow:
    amount = data["amount_usd"]
    if isinstance(amount, Decimal):
        amount_usd = float(amount)
    else:
        amount_usd = float(amount)
    club_slug = _slug_for_payment_club(session, data.get("club_id"), data)
    return TaggedManualAuditRow(
        amount_usd=amount_usd,
        payer_name=str(data["payer_name"]),
        account_tag=str(data.get(tag_field) or "").strip(),
        group_title=_manual_group_cell(data),
        club_label=_manual_club_name(data, club_names),
        time_label=_fmt_manual_audit_time(data["created_at"], club_slug),
    )


def _fetch_manual_rows(
    session: Session,
    payment_cls,
    build_read: Callable,
    club_names: dict[int, str],
    from_dt: datetime,
    to_dt: datetime,
    *,
    audit_date: str,
) -> list[ManualAuditRow]:
    query = _apply_audit_manual_filters(
        session,
        session.query(payment_cls),
        payment_cls,
        from_dt=from_dt,
        to_dt=to_dt,
    )
    rows = query.order_by(payment_cls.created_at.desc(), payment_cls.id.desc()).all()
    out: list[ManualAuditRow] = []
    for row in rows:
        data = build_read(session, row)
        if not _payment_in_audit_day(
            session,
            audit_date=audit_date,
            club_id=data.get("club_id"),
            occurred_at=data.get("created_at"),
            data=data,
        ):
            continue
        out.append(_manual_row(session, data, club_names))
    return out


def _bonus_payer_display(record: BonusRecord) -> str:
    title = (record.group_title or "").strip()
    if title:
        from cashier.services.zapier import build_zapier_name

        formatted = build_zapier_name(title)
        return formatted or title
    return str(record.player_username).strip()


def _bonus_group_cell(record: BonusRecord) -> str:
    type_name = (record.bonus_type.name if record.bonus_type else "").strip()
    desc = (record.custom_description or "").strip()
    if type_name and desc:
        return f"{type_name} — {desc}"
    return type_name or desc


def _fetch_bonus_rows(
    session: Session,
    club_names: dict[int, str],
    from_dt: datetime,
    to_dt: datetime,
    *,
    audit_date: str,
) -> list[ManualAuditRow]:
    rows = (
        session.query(BonusRecord)
        .filter(
            BonusRecord.created_at >= from_dt,
            BonusRecord.created_at <= to_dt,
        )
        .order_by(BonusRecord.created_at.desc(), BonusRecord.id.desc())
        .all()
    )
    out: list[ManualAuditRow] = []
    for row in rows:
        if not _payment_in_audit_day(
            session,
            audit_date=audit_date,
            club_id=row.club_id,
            occurred_at=row.created_at,
        ):
            continue
        club_slug = _slug_for_payment_club(session, row.club_id)
        amount = row.amount
        amount_usd = float(amount) if amount is not None else 0.0
        out.append(
            ManualAuditRow(
                amount_usd=amount_usd,
                payer_name=_bonus_payer_display(row),
                group_title=_bonus_group_cell(row),
                club_label=_club_name(club_names, row.club_id),
                time_label=_fmt_manual_audit_time(row.created_at, club_slug),
            )
        )
    return out


def _manual_row(session: Session, data: dict, club_names: dict[int, str]) -> ManualAuditRow:
    amount = data["amount_usd"]
    if isinstance(amount, Decimal):
        amount_usd = float(amount)
    else:
        amount_usd = float(amount)
    club_slug = _slug_for_payment_club(session, data.get("club_id"), data)
    return ManualAuditRow(
        amount_usd=amount_usd,
        payer_name=str(data["payer_name"]),
        group_title=_manual_group_cell(data),
        club_label=_manual_club_name(data, club_names),
        time_label=_fmt_manual_audit_time(data["created_at"], club_slug),
    )
