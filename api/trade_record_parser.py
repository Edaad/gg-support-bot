"""Parse ClubGG trade record XLSX (Trade Record tab) for audit ingest."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, BinaryIO, Literal

from openpyxl import load_workbook

from api.club_audit_timezone import (
    audit_timezone_for_slug,
    parse_row_datetime as audit_parse_row_datetime,
    period_timezone_warning,
)
from api.club_slug import CLUB_LABEL_TO_SLUG, CLUB_SLUG_TO_NAME, slug_for_club_name

SHEET_NAME = "Trade Record"
DATA_START_ROW = 6
_GG_ID_RE = re.compile(r"^[0-9]{1,48}-[0-9]{1,48}$")

# Column indices (1-based) — Aces-21.xlsx Trade Record layout
COL_TIME = 1  # A Date
COL_AMOUNT = 7  # G Amount (F=Before, H=After)
COL_SA_ID = 11  # K
COL_SA_NICK = 12  # L
COL_AGENT_ID = 13  # M
COL_AGENT_NICK = 14  # N
COL_MEMBER_ID = 15  # O
COL_MEMBER_NICK = 16  # P

Role = Literal["member", "agent", "superAgent"]

ROLE_COLUMNS: dict[Role, tuple[int, int]] = {
    "superAgent": (COL_SA_ID, COL_SA_NICK),
    "agent": (COL_AGENT_ID, COL_AGENT_NICK),
    "member": (COL_MEMBER_ID, COL_MEMBER_NICK),
}


@dataclass(frozen=True)
class TradeRecordMetadata:
    club_text: str
    club_id_text: str
    date_text: str


@dataclass(frozen=True)
class ParsedIdentity:
    role: Role
    gg_player_id: str
    nickname: str


@dataclass(frozen=True)
class ParsedTransaction:
    sheet_row: int
    occurred_at: datetime | None
    amount: Decimal
    member_gg_player_id: str | None
    member_nickname: str | None
    agent_gg_player_id: str | None
    super_agent_gg_player_id: str | None


@dataclass
class TradeRecordParseResult:
    metadata: TradeRecordMetadata
    audit_date: date
    club_slug: str
    identities: list[ParsedIdentity] = field(default_factory=list)
    transactions: list[ParsedTransaction] = field(default_factory=list)
    skipped_rows: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


class TradeRecordParseError(ValueError):
    pass


class TradeRecordValidationError(ValueError):
    pass


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def _normalize_gg_id(raw: Any) -> str | None:
    s = _cell_str(raw)
    if not s or s == "-":
        return None
    if _GG_ID_RE.match(s):
        return s
    return None


def _parse_amount(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    s = _cell_str(raw).replace(",", "")
    if not s or s == "-":
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _row_cells(ws, row: int, max_col: int = 20) -> list[str]:
    return [_cell_str(ws.cell(row=row, column=c).value) for c in range(1, max_col + 1)]


def _split_label_value(text: str) -> tuple[str, str]:
    raw = text.strip()
    if not raw:
        return "", ""
    if ":" in raw:
        label, _, value = raw.partition(":")
        return label.strip(), value.strip()
    return "", raw


def _metadata_from_sheet(ws) -> TradeRecordMetadata:
    club_text = ""
    club_id_text = ""
    date_text = ""

    for row in range(1, 4):
        cells = _row_cells(ws, row)
        non_empty = [v for v in cells if v]
        if not non_empty:
            continue

        label = ""
        value = ""
        if len(non_empty) == 1:
            label, value = _split_label_value(non_empty[0])
            if not label:
                value = non_empty[0]
        else:
            label = non_empty[0]
            value = non_empty[1]

        ll = label.lower()
        if "club name" in ll or ll == "club":
            club_text = value or club_text
        elif "club id" in ll:
            club_id_text = value or club_id_text
        elif "period" in ll or "date" in ll:
            date_text = value or date_text

        if not club_text and row == 1 and not label and value:
            club_text = value

    if not club_text:
        raise TradeRecordParseError("Trade Record metadata rows 1–3 are empty")
    if not date_text:
        raise TradeRecordParseError("Trade Record Period metadata is missing in rows 1–3")

    return TradeRecordMetadata(
        club_text=club_text,
        club_id_text=club_id_text,
        date_text=date_text,
    )


# Display labels for trade-record metadata validation (match dashboard clubMap.ts)
CLUB_SLUG_DISPLAY_LABELS: dict[str, str] = {
    "clubgto": "ClubGTO",
    "round-table": "Round Table",
    "aces-table": "Aces Table",
    "creator-club": "Creator Club",
}


def _parse_metadata_date(text: str) -> date | None:
    raw = text.strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(raw[:10] if fmt == "%Y-%m-%d" else raw, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", raw)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", raw)
    if m:
        return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    return None


def extract_audit_date_from_metadata(metadata: TradeRecordMetadata) -> date:
    """Parse ET audit day from Period row (e.g. 2026-06-21 ~ 2026-06-21 (UTC-5:00))."""
    combined = metadata.date_text.strip()
    if not combined:
        raise TradeRecordValidationError("Trade Record Period metadata is empty")

    dates_in_text = re.findall(r"(\d{4})-(\d{2})-(\d{2})", combined)
    if dates_in_text:
        found = [date(int(y), int(m), int(d)) for y, m, d in dates_in_text]
        if len(found) >= 2 and found[0] != found[-1]:
            raise TradeRecordValidationError(
                f"Period spans multiple days ({found[0].isoformat()} to {found[-1].isoformat()}); "
                "upload one calendar day at a time."
            )
        return found[0]

    parsed = _parse_metadata_date(combined)
    if parsed:
        return parsed

    raise TradeRecordValidationError(
        f"Could not parse audit date from Period metadata {metadata.date_text!r}."
    )


def resolve_club_slug_from_metadata(metadata: TradeRecordMetadata) -> str:
    """Map Club Name metadata to a known gg-computer club slug."""
    raw = metadata.club_text.strip()
    if not raw:
        raise TradeRecordValidationError("Trade Record Club Name metadata is empty")

    key = raw.lower()
    if key in CLUB_LABEL_TO_SLUG:
        return CLUB_LABEL_TO_SLUG[key]

    slug = slug_for_club_name(raw)
    if slug and slug in CLUB_SLUG_TO_NAME:
        return slug

    for label, mapped in CLUB_LABEL_TO_SLUG.items():
        if label in key or key in label:
            return mapped

    raise TradeRecordValidationError(
        f"Unknown club in file metadata: {raw!r}. "
        f"Expected one of: {', '.join(sorted(CLUB_SLUG_DISPLAY_LABELS.values()))}."
    )


def _row_is_blank(ws, row: int) -> bool:
    member_id = _cell_str(ws.cell(row=row, column=COL_MEMBER_ID).value)
    amount = ws.cell(row=row, column=COL_AMOUNT).value
    time_val = ws.cell(row=row, column=COL_TIME).value
    return not member_id and amount in (None, "") and time_val in (None, "")


def parse_trade_record_workbook(
    source: BinaryIO | bytes,
) -> TradeRecordParseResult:
    if isinstance(source, bytes):
        stream: BinaryIO = io.BytesIO(source)
    else:
        stream = source
        stream.seek(0)

    wb = load_workbook(stream, read_only=False, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise TradeRecordParseError(f'Missing "{SHEET_NAME}" sheet')

    ws = wb[SHEET_NAME]
    metadata = _metadata_from_sheet(ws)
    audit_date = extract_audit_date_from_metadata(metadata)
    club_slug = resolve_club_slug_from_metadata(metadata)
    timezone_policy = audit_timezone_for_slug(club_slug)
    warnings: list[str] = []
    tz_warning = period_timezone_warning(metadata.date_text, club_slug)
    if tz_warning:
        warnings.append(tz_warning)

    identities_by_key: dict[tuple[Role, str], ParsedIdentity] = {}
    transactions: list[ParsedTransaction] = []
    skipped_rows: list[str] = []

    row = DATA_START_ROW
    while row <= (ws.max_row or DATA_START_ROW):
        if _row_is_blank(ws, row):
            break

        member_id = _normalize_gg_id(ws.cell(row=row, column=COL_MEMBER_ID).value)
        member_nick = _cell_str(ws.cell(row=row, column=COL_MEMBER_NICK).value)
        agent_id = _normalize_gg_id(ws.cell(row=row, column=COL_AGENT_ID).value)
        sa_id = _normalize_gg_id(ws.cell(row=row, column=COL_SA_ID).value)

        amount = _parse_amount(ws.cell(row=row, column=COL_AMOUNT).value)
        if amount is None:
            skipped_rows.append(f"row {row}: missing or invalid amount")
            row += 1
            continue

        occurred_at = audit_parse_row_datetime(
            ws.cell(row=row, column=COL_TIME).value,
            audit_date,
            timezone_policy,
        )

        for role, (id_col, nick_col) in ROLE_COLUMNS.items():
            gid = _normalize_gg_id(ws.cell(row=row, column=id_col).value)
            nick = _cell_str(ws.cell(row=row, column=nick_col).value)
            if not gid:
                continue
            if not nick:
                nick = gid
            key = (role, gid)
            identities_by_key[key] = ParsedIdentity(
                role=role,
                gg_player_id=gid,
                nickname=nick,
            )

        transactions.append(
            ParsedTransaction(
                sheet_row=row,
                occurred_at=occurred_at,
                amount=amount,
                member_gg_player_id=member_id,
                member_nickname=member_nick or None,
                agent_gg_player_id=agent_id,
                super_agent_gg_player_id=sa_id,
            )
        )
        row += 1

    wb.close()

    if not transactions:
        raise TradeRecordParseError("No transaction rows found from row 6")

    return TradeRecordParseResult(
        metadata=metadata,
        audit_date=audit_date,
        club_slug=club_slug,
        identities=list(identities_by_key.values()),
        transactions=transactions,
        skipped_rows=skipped_rows,
        warnings=warnings,
    )
