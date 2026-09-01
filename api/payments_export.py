"""XLSX export for unified payments (Summary + per-method sheets)."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from sqlalchemy.orm import Session

from api.schemas_payments import UnifiedPaymentRowRead
from db.models import Club


def _club_name(club_names: dict[int, str], club_id: int | None) -> str:
    if club_id is None:
        return "Unbound"
    return club_names.get(int(club_id), f"Club {club_id}")


def _fmt_status(status: str | None) -> str:
    if status is None:
        return "—"
    return status


def _fmt_dt(value: datetime | str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return value.isoformat()


def _summary_row(
    row: UnifiedPaymentRowRead,
    club_names: dict[int, str],
) -> list[Any]:
    return [
        _fmt_dt(row.occurred_at),
        float(row.amount_usd),
        row.group_title or "",
        row.gg_nickname or "Not available",
        row.method_label,
        row.owner_label,
        _club_name(club_names, row.club_id),
        _fmt_status(row.status),
    ]


def _stripe_detail_rows(
    rows: list[UnifiedPaymentRowRead],
    club_names: dict[int, str],
) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "completed_at",
        "group_title",
        "gg_nickname",
        "gg_player_id",
        "method_name",
        "owner",
        "club",
        "amount_usd",
        "stripe_fee_usd",
        "currency",
        "stripe_payment_intent_id",
        "stripe_checkout_session_id",
    ]
    data = []
    for row in rows:
        d = row.detail
        data.append(
            [
                d.get("completed_at") or d.get("created_at") or "",
                d.get("group_title") or "",
                d.get("gg_nickname") or "",
                d.get("gg_player_id") or "",
                d.get("method_name") or "",
                row.owner_label,
                _club_name(club_names, row.club_id),
                str(d.get("amount_usd", "")),
                str(d.get("stripe_fee_usd", "")),
                d.get("currency") or "",
                d.get("stripe_payment_intent_id") or "",
                d.get("stripe_checkout_session_id") or "",
            ]
        )
    return headers, data


def _crypto_detail_rows(
    rows: list[UnifiedPaymentRowRead],
    club_names: dict[int, str],
) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "created_at",
        "from_label",
        "chain",
        "token_symbol",
        "to_address",
        "transaction_hash",
        "group_title",
        "gg_nickname",
        "gg_player_id",
        "owner",
        "club",
        "amount_usd",
        "status",
    ]
    data = []
    for row in rows:
        d = row.detail
        data.append(
            [
                d.get("created_at") or "",
                d.get("from_label") or "",
                d.get("chain") or "",
                d.get("token_symbol") or "",
                d.get("to_address") or "",
                d.get("transaction_hash") or "",
                d.get("group_title") or "",
                d.get("gg_nickname") or "",
                d.get("gg_player_id") or "",
                row.owner_label,
                _club_name(club_names, row.club_id),
                str(d.get("amount_usd", "")),
                d.get("status") or "",
            ]
        )
    return headers, data


def _manual_ingest_detail_rows(
    rows: list[UnifiedPaymentRowRead],
    club_names: dict[int, str],
    method_slug: str,
) -> tuple[list[str], list[list[Any]]]:
    account_key = {
        "zelle": "zelle_recipient",
        "cashapp": "cashapp_handle",
        "paypal": "paypal_email",
        "venmo": "venmo_handle",
    }.get(method_slug, "venmo_handle")
    headers = [
        "created_at",
        "payer_name",
        account_key,
        "group_title",
        "gg_nickname",
        "gg_player_id",
        "owner",
        "club",
        "amount_usd",
        "status",
        "auto_bound",
    ]
    if method_slug == "venmo":
        headers.append("goods_or_services")
    data = []
    for row in rows:
        d = row.detail
        base = [
            d.get("created_at") or "",
            d.get("payer_name") or "",
            d.get(account_key) or "",
            d.get("group_title") or "",
            d.get("gg_nickname") or "",
            d.get("gg_player_id") or "",
            row.owner_label,
            _club_name(club_names, row.club_id),
            str(d.get("amount_usd", "")),
            d.get("status") or "",
            str(d.get("auto_bound", "")),
        ]
        if method_slug == "venmo":
            base.append(str(d.get("goods_or_services", "")))
        data.append(base)
    return headers, data


def _union_detail_rows(
    rows: list[UnifiedPaymentRowRead],
    club_names: dict[int, str],
) -> tuple[list[str], list[list[Any]]]:
    headers = [
        "created_at",
        "method_name",
        "variant_name",
        "club",
        "group_title",
        "player",
        "amount",
    ]
    data = []
    for row in rows:
        d = row.detail
        club = d.get("club") or {}
        data.append(
            [
                d.get("created_at") or "",
                d.get("method_name") or "",
                d.get("variant_name") or "",
                club.get("name") or _club_name(club_names, row.club_id),
                d.get("group_title") or "",
                row.gg_nickname or "Not available",
                str(d.get("amount", row.amount_usd)),
            ]
        )
    return headers, data


_SHEET_BUILDERS = {
    "stripe": lambda rows, clubs: _stripe_detail_rows(rows, clubs),
    "crypto": lambda rows, clubs: _crypto_detail_rows(rows, clubs),
    "venmo": lambda rows, clubs: _manual_ingest_detail_rows(rows, clubs, "venmo"),
    "zelle": lambda rows, clubs: _manual_ingest_detail_rows(rows, clubs, "zelle"),
    "cashapp": lambda rows, clubs: _manual_ingest_detail_rows(rows, clubs, "cashapp"),
    "paypal": lambda rows, clubs: _manual_ingest_detail_rows(rows, clubs, "paypal"),
    "applepay": lambda rows, clubs: _union_detail_rows(rows, clubs),
}


def _sheet_title(method_slug: str) -> str:
    titles = {
        "stripe": "Stripe",
        "venmo": "Venmo",
        "zelle": "Zelle",
        "cashapp": "Cash App",
        "paypal": "PayPal",
        "crypto": "Crypto",
        "applepay": "Apple Pay",
        "union_manual": "Union",
    }
    return titles.get(method_slug, method_slug.title())[:31]


def build_payments_workbook(
    db: Session,
    rows: list[UnifiedPaymentRowRead],
) -> bytes:
    club_names = {int(c.id): c.name for c in db.query(Club).all()}

    wb = Workbook()
    wb.remove(wb.active)

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(
        ["time", "amount", "group", "player", "method", "owner", "club", "status"]
    )
    sorted_rows = sorted(
        rows,
        key=lambda r: (
            -r.occurred_at.timestamp()
            if r.occurred_at.tzinfo
            else -r.occurred_at.replace(tzinfo=None).timestamp(),
            r.id,
        ),
    )
    for row in sorted_rows:
        summary_ws.append(_summary_row(row, club_names))

    by_method: dict[str, list[UnifiedPaymentRowRead]] = {}
    for row in rows:
        key = "union_manual" if row.source == "union_manual" else row.method_slug
        by_method.setdefault(key, []).append(row)

    for method_slug in sorted(by_method.keys()):
        method_rows = by_method[method_slug]
        if method_slug == "union_manual":
            builder = _union_detail_rows
            title = "Union"
        else:
            builder = _SHEET_BUILDERS.get(method_slug)
            title = _sheet_title(method_slug)
        if builder is None:
            continue
        headers, data = builder(method_rows, club_names)
        ws = wb.create_sheet(title)
        ws.append(headers)
        for data_row in data:
            ws.append(data_row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
